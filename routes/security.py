from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, abort, session
from flask_login import login_required, current_user
from sqlalchemy import text, func, and_
from sqlalchemy.exc import SAWarning
from datetime import datetime, timedelta, timezone
from extensions import db, cache
from models import User, AuditLog, SystemSettings
import utils
from functools import wraps
import json
import os
import warnings
from collections import defaultdict, Counter
from routes.advanced_control import _log_owner_action
import re

from AI.engine.ai_service import (
    ai_chat_with_search,
    search_database_for_query,
    gather_system_context,
    build_system_message,
    get_system_setting,
)

ALLOWED_DIRECT_SETTINGS_KEYS: set[str] = {
    # General flags
    "maintenance_mode",
    "registration_enabled",
    "api_enabled",
    # Advanced config (string/integer values)
    "SESSION_TIMEOUT",
    "MAX_LOGIN_ATTEMPTS",
    "PASSWORD_MIN_LENGTH",
    "AUTO_BACKUP_ENABLED",
    "BACKUP_INTERVAL_HOURS",
    "ENABLE_EMAIL_NOTIFICATIONS",
    "ENABLE_SMS_NOTIFICATIONS",
    # Company info
    "COMPANY_ADDRESS",
    "COMPANY_PHONE",
    "COMPANY_EMAIL",
    "TAX_NUMBER",
    "CURRENCY_SYMBOL",
    "TIMEZONE",
    # Branding & theme
    "system_name",
    "company_name",
    "login_title",
    "login_subtitle",
    "footer_text",
    "primary_color",
    "secondary_color",
    "sidebar_bg",
    "sidebar_text",
}

security_bp = Blueprint('security', __name__, url_prefix='/security')

@security_bp.before_request
@login_required
def restrict_security_access():
    # الحماية: فقط للمالك والمطور (Level 0)
    if not (current_user.is_system or current_user.role_name_l in ['owner', 'developer']):
        flash('⛔ غير مصرح لك بالوصول للوحة المالك', 'danger')
        return redirect(url_for('main.dashboard'))


def make_aware(dt):
    if dt and dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


@security_bp.app_template_global()
def _get_action_icon(action):
    if not action:
        return 'info-circle'
    mapping = {
        'login': 'sign-in-alt',
        'logout': 'sign-out-alt',
        'create': 'plus',
        'update': 'edit',
        'delete': 'trash',
        'view': 'eye',
        'export': 'download',
        'import': 'upload',
        'blocked': 'ban',
        'security': 'shield-alt'
    }
    action_lower = str(action).lower()
    for key, icon in mapping.items():
        if key in action_lower:
            return icon
    return 'circle'


@security_bp.app_template_global()
def _get_action_color(action):
    """لون للنشاط - Template Global"""
    if not action:
        return 'secondary'
    mapping = {
        'login': 'success',
        'logout': 'secondary',
        'create': 'primary',
        'update': 'info',
        'delete': 'danger',
        'blocked': 'danger',
        'failed': 'danger',
        'security': 'warning'
    }
    action_lower = str(action).lower()
    for key, color in mapping.items():
        if key in action_lower:
            return color
    return 'secondary'


def _is_safe_identifier(name: str) -> bool:
    return bool(re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', str(name or '')))


def _is_safe_qualified_identifier(name: str) -> bool:
    return bool(re.match(r'^[A-Za-z_][A-Za-z0-9_]*(\.[A-Za-z_][A-Za-z0-9_]*)?$', str(name or '')))


def _is_postgresql() -> bool:
    uri = current_app.config.get("SQLALCHEMY_DATABASE_URI", "") or ""
    return "postgresql" in uri


def _normalize_column_type(raw: str) -> str | None:
    raw = str(raw or '').strip()
    if not raw:
        return None
    ct = re.sub(r"\s+", " ", raw).strip().upper()
    m = re.match(r"^([A-Z]+(?: [A-Z]+)*)\s*(\(\s*\d+\s*(,\s*\d+\s*)?\))?$", ct)
    if not m:
        return None
    base = m.group(1).strip()
    params = m.group(2) or ""
    allowed = {
        "TEXT",
        "INTEGER",
        "BIGINT",
        "SMALLINT",
        "NUMERIC",
        "DECIMAL",
        "REAL",
        "FLOAT",
        "DOUBLE PRECISION",
        "BOOLEAN",
        "DATE",
        "DATETIME",
        "TIMESTAMP",
        "TIME",
        "UUID",
        "JSON",
        "JSONB",
        "BYTEA",
        "BLOB",
        "VARCHAR",
        "CHAR",
    }
    if base not in allowed:
        return None
    return f"{base}{params}"


def _get_primary_key_column(table_name: str) -> str | None:
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        cols = [c.get('name') for c in (inspector.get_columns(table_name) or []) if c.get('name')]
        pk = (inspector.get_pk_constraint(table_name) or {}).get('constrained_columns') or []
        if pk:
            return pk[0] if pk[0] in cols else None
        if 'id' in cols:
            return 'id'
        return cols[0] if cols else None
    except Exception:
        return None


def _get_table_columns_safe(table_name: str) -> list[str]:
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        return [c['name'] for c in inspector.get_columns(table_name)]
    except Exception:
        return []


def _assert_valid_table(table_name: str):
    tables = _get_all_tables()
    if not _is_safe_identifier(table_name) or table_name not in tables:
        abort(400)

from utils import permission_required
from permissions_config.enums import SystemPermissions

# Removed deprecated owner_only alias


@security_bp.route('/api/system-accounts/change-password', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_system_account_change_password():
    """API: تغيير كلمة مرور حساب نظام (Owner/Developer)"""
    try:
        data = request.get_json(silent=True) or {}
        target_username = data.get('username')
        new_password = data.get('new_password')
        
        if not target_username or not new_password:
            return jsonify({'success': False, 'error': 'اسم المستخدم وكلمة المرور الجديدة مطلوبان'}), 400
            
        user = User.query.filter(func.lower(User.username) == target_username.strip().lower()).first()
        if not user:
            return jsonify({'success': False, 'error': 'المستخدم غير موجود'}), 404
            
        if not user.is_system_account and user.username != '__OWNER__':
            return jsonify({'success': False, 'error': 'هذا ليس حساب نظام محمي'}), 400
            
        user.set_password(new_password)
        db.session.commit()
        
        _log_owner_action('system_account_password_change', target=user.id, meta={'username': user.username})
        return jsonify({'success': True, 'message': 'تم تغيير كلمة المرور بنجاح'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/saas-manager')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def saas_manager():
    """
    🚀 SaaS Manager - إدارة الاشتراكات والفواتير
    """
    from models import SaaSPlan, SaaSSubscription, SaaSInvoice
    from sqlalchemy import func
    from decimal import Decimal
    
    try:
        plans = SaaSPlan.query.order_by(SaaSPlan.sort_order, SaaSPlan.price_monthly).all()
    except Exception:
        plans = []
    
    try:
        subscriptions = SaaSSubscription.query.order_by(SaaSSubscription.created_at.desc()).limit(50).all()
    except Exception:
        subscriptions = []
    
    try:
        invoices = SaaSInvoice.query.order_by(SaaSInvoice.created_at.desc()).limit(50).all()
    except Exception:
        invoices = []
    
    try:
        total_subscribers = SaaSSubscription.query.count()
        active_subscribers = SaaSSubscription.query.filter_by(status='active').count()
        trial_users = SaaSSubscription.query.filter_by(status='trial').count()
        
        # حساب إيرادات SaaS مع تحويل العملات
        saas_invoices = SaaSInvoice.query.filter(
            SaaSInvoice.status == 'paid',
            SaaSInvoice.created_at >= datetime.now(timezone.utc) - timedelta(days=30)
        ).all()
        
        monthly_revenue = Decimal('0.00')
        for inv in saas_invoices:
            amt = Decimal(str(inv.amount or 0))
            inv_currency = getattr(inv, 'currency', 'USD')
            if inv_currency == 'ILS':
                monthly_revenue += amt
            else:
                try:
                    from models import convert_amount
                    monthly_revenue += convert_amount(amt, inv_currency, 'ILS', inv.created_at)
                except Exception:
                    monthly_revenue += amt
        
        stats = {
            'total_subscribers': total_subscribers,
            'active_subscribers': active_subscribers,
            'monthly_revenue': f"${float(monthly_revenue):,.2f}",
            'trial_users': trial_users
        }
    except Exception:
        stats = {
            'total_subscribers': 0,
            'active_subscribers': 0,
            'monthly_revenue': '$0.00',
            'trial_users': 0
        }
    
    return render_template('security/saas_manager.html', 
                         stats=stats, 
                         plans=plans,
                         subscriptions=subscriptions,
                         invoices=invoices,
                         today=datetime.now().strftime('%Y-%m-%d'))


@security_bp.route('/api/saas/plans', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_create_plan():
    """API: إنشاء باقة جديدة"""
    from models import SaaSPlan
    
    try:
        data = request.get_json(silent=True) or {}
        name = str(data.get('name') or '').strip()
        if not name or len(name) < 2:
            return jsonify({'success': False, 'error': 'الاسم مطلوب'}), 400
        try:
            price_monthly = float(data.get('price_monthly'))
        except Exception:
            return jsonify({'success': False, 'error': 'سعر شهري غير صالح'}), 400
        if price_monthly <= 0:
            return jsonify({'success': False, 'error': 'السعر الشهري يجب أن يكون أكبر من 0'}), 400
        currency = (data.get('currency') or 'USD').strip().upper()
        if currency not in {'USD', 'ILS', 'EUR'}:
            return jsonify({'success': False, 'error': 'عملة غير مدعومة'}), 400
        price_yearly = data.get('price_yearly')
        price_yearly = float(price_yearly) if price_yearly not in (None, '',) else None
        max_users = data.get('max_users')
        max_users = int(max_users) if max_users not in (None, '',) else None
        max_invoices = data.get('max_invoices')
        max_invoices = int(max_invoices) if max_invoices not in (None, '',) else None
        storage_gb = data.get('storage_gb')
        storage_gb = int(storage_gb) if storage_gb not in (None, '',) else None
        features = str(data.get('features') or '').strip() or None
        is_popular = bool(data.get('is_popular', False))
        plan = SaaSPlan(
            name=name,
            description=str(data.get('description') or '').strip() or None,
            price_monthly=price_monthly,
            price_yearly=price_yearly,
            currency=currency,
            max_users=max_users,
            max_invoices=max_invoices,
            storage_gb=storage_gb,
            features=features,
            is_popular=is_popular
        )
        db.session.add(plan)
        db.session.commit()
        _log_owner_action('saas_plan_create', target=plan.id, meta={'name': plan.name, 'price_monthly': float(plan.price_monthly), 'currency': plan.currency})
        return jsonify({'success': True, 'plan_id': plan.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/subscriptions', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_create_subscription():
    """API: إنشاء اشتراك جديد"""
    from models import SaaSSubscription, Customer, SaaSPlan
    from datetime import datetime, timedelta, timezone
    
    try:
        data = request.get_json(silent=True) or {}
        customer_id = data.get('customer_id')
        plan_id = data.get('plan_id')
        start_date_str = data.get('start_date')
        if not customer_id or not plan_id or not start_date_str:
            return jsonify({'success': False, 'error': 'حقول مطلوبة ناقصة'}), 400
        customer = db.session.get(Customer, customer_id)
        if not customer:
            return jsonify({'success': False, 'error': 'العميل غير موجود'}), 404
        plan = db.session.get(SaaSPlan, plan_id)
        if not plan:
            return jsonify({'success': False, 'error': 'الباقة غير موجودة'}), 404
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        sub = SaaSSubscription(
            customer_id=customer.id,
            plan_id=plan.id,
            status=data.get('status', 'trial'),
            start_date=start_date,
            end_date=start_date + timedelta(days=30)
        )
        db.session.add(sub)
        db.session.commit()
        _log_owner_action('saas_subscription_create', target=sub.id, meta={'customer_id': sub.customer_id, 'plan_id': sub.plan_id, 'status': sub.status})
        return jsonify({'success': True, 'subscription_id': sub.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/invoices/<int:invoice_id>/mark-paid', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_mark_paid(invoice_id):
    """API: تأكيد دفع الفاتورة"""
    from models import SaaSInvoice
    
    try:
        invoice = db.get_or_404(SaaSInvoice, invoice_id)
        invoice.status = 'paid'
        invoice.paid_at = datetime.now(timezone.utc)
        db.session.commit()
        _log_owner_action('saas_invoice_mark_paid', target=invoice.id, meta={'subscription_id': invoice.subscription_id, 'amount': float(invoice.amount)})
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/subscriptions/<int:sub_id>/cancel', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_cancel_subscription(sub_id):
    """API: إلغاء اشتراك"""
    from models import SaaSSubscription
    
    try:
        sub = db.get_or_404(SaaSSubscription, sub_id)
        sub.status = 'cancelled'
        sub.cancelled_at = datetime.now(timezone.utc)
        sub.cancelled_by = current_user.id
        sub.cancellation_reason = request.get_json().get('reason', 'إلغاء من المالك')
        db.session.commit()
        _log_owner_action('saas_subscription_cancel', target=sub.id, meta={'reason': sub.cancellation_reason})
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/subscriptions/<int:sub_id>/renew', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_renew_subscription(sub_id):
    """API: تجديد اشتراك"""
    from models import SaaSSubscription
    
    try:
        sub = db.get_or_404(SaaSSubscription, sub_id)
        
        if sub.status == 'cancelled':
            return jsonify({'success': False, 'error': 'لا يمكن تجديد اشتراك ملغي'}), 400
        
        # تجديد لمدة 30 يوم
        sub.end_date = sub.end_date + timedelta(days=30) if sub.end_date else datetime.now(timezone.utc) + timedelta(days=30)
        sub.status = 'active'
        db.session.commit()
        
        _log_owner_action('saas_subscription_renew', target=sub.id, meta={'new_end_date': sub.end_date.strftime('%Y-%m-%d') if sub.end_date else None})
        return jsonify({'success': True, 'new_end_date': sub.end_date.strftime('%Y-%m-%d')})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/plans/<int:plan_id>', methods=['PUT'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_update_plan(plan_id):
    """API: تحديث باقة"""
    from models import SaaSPlan
    
    try:
        plan = db.get_or_404(SaaSPlan, plan_id)
        data = request.get_json(silent=True) or {}
        
        if 'name' in data:
            plan.name = data['name']
        if 'description' in data:
            plan.description = data['description']
        if 'price_monthly' in data:
            plan.price_monthly = float(data['price_monthly'])
        if 'price_yearly' in data:
            plan.price_yearly = float(data['price_yearly']) if data['price_yearly'] else None
        if 'currency' in data and data['currency']:
            cur = str(data['currency']).strip().upper()
            if cur not in {'USD', 'ILS', 'EUR'}:
                return jsonify({'success': False, 'error': 'عملة غير مدعومة'}), 400
            plan.currency = cur
        if 'max_users' in data:
            plan.max_users = int(data['max_users']) if data['max_users'] else None
        if 'max_invoices' in data:
            plan.max_invoices = int(data['max_invoices']) if data['max_invoices'] else None
        if 'storage_gb' in data:
            plan.storage_gb = int(data['storage_gb']) if data['storage_gb'] else None
        if 'features' in data:
            plan.features = str(data['features'] or '').strip() or None
        if 'is_popular' in data:
            plan.is_popular = bool(data['is_popular'])
        
        db.session.commit()
        _log_owner_action('saas_plan_update', target=plan.id, meta={'name': plan.name})
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/invoices', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_create_invoice():
    """API: إنشاء فاتورة"""
    from models import SaaSInvoice, SaaSSubscription
    from decimal import Decimal
    
    try:
        data = request.get_json(silent=True) or {}
        subscription_id = data.get('subscription_id')
        if not subscription_id:
            return jsonify({'success': False, 'error': 'الاشتراك مطلوب'}), 400
        sub = db.session.get(SaaSSubscription, subscription_id)
        if not sub:
            return jsonify({'success': False, 'error': 'اشتراك غير موجود'}), 404
        try:
            amount = Decimal(str(data.get('amount')))
        except Exception:
            return jsonify({'success': False, 'error': 'المبلغ غير صالح'}), 400
        if amount <= 0:
            return jsonify({'success': False, 'error': 'المبلغ يجب أن يكون أكبر من 0'}), 400
        currency = (data.get('currency') or 'USD').strip().upper()
        if currency not in {'USD', 'ILS', 'EUR'}:
            return jsonify({'success': False, 'error': 'عملة غير مدعومة'}), 400
        due_date_str = data.get('due_date')
        due_dt = datetime.strptime(due_date_str, '%Y-%m-%d') if due_date_str else datetime.now(timezone.utc) + timedelta(days=7)
        from time import time
        suffix = int(time() * 1000) % 1000000
        inv_number = f"SINV-{datetime.now().strftime('%Y%m%d')}-{suffix:06d}"
        invoice = SaaSInvoice(
            invoice_number=inv_number,
            subscription_id=sub.id,
            amount=amount,
            currency=currency,
            status='pending',
            due_date=due_dt
        )
        db.session.add(invoice)
        db.session.commit()
        _log_owner_action('saas_invoice_create', target=invoice.id, meta={'subscription_id': sub.id, 'amount': float(invoice.amount), 'currency': invoice.currency})
        return jsonify({'success': True, 'invoice_id': invoice.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/invoices/<int:invoice_id>/send-reminder', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_send_reminder(invoice_id):
    """API: إرسال تذكير دفع"""
    from models import SaaSInvoice, SaaSSubscription, Customer
    from utils import send_notification_email
    
    try:
        invoice = db.get_or_404(SaaSInvoice, invoice_id)
        
        if invoice.status == 'paid':
            return jsonify({'success': False, 'error': 'الفاتورة مدفوعة بالفعل'}), 400
        
        subscription = db.session.get(SaaSSubscription, invoice.subscription_id)
        if not subscription:
            return jsonify({'success': False, 'error': 'اشتراك غير موجود'}), 404
        
        customer = db.session.get(Customer, subscription.customer_id)
        if not customer or not customer.email:
            return jsonify({'success': False, 'error': 'لا يوجد بريد إلكتروني للعميل'}), 400
        html_body = f"""
        <h3>تذكير بالدفع</h3>
        <p>عزيزنا {customer.name}</p>
        <p>الرجاء تسديد فاتورة الاشتراك رقم {invoice.invoice_number} بمبلغ {invoice.currency} {float(invoice.amount):,.2f} قبل {invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else ''}.</p>
        """
        result = send_notification_email(
            to=customer.email,
            subject=f'تذكير دفع فاتورة الاشتراك #{invoice.invoice_number}',
            body_html=html_body,
            body_text=f'تذكير دفع فاتورة الاشتراك #{invoice.invoice_number} بمبلغ {float(invoice.amount)} {invoice.currency}'
        )
        if not result.get('success'):
            return jsonify({'success': False, 'error': result.get('error', 'فشل الإرسال')}), 400
        _log_owner_action('saas_invoice_reminder', target=invoice.id, meta={'customer_email': customer.email})
        return jsonify({'success': True, 'message': f'تم إرسال التذكير إلى {customer.email}'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/customers', methods=['GET'])
@login_required
def api_get_customers():
    """API: جلب قائمة العملاء"""
    from models import Customer
    
    try:
        customers = Customer.query.filter_by(is_active=True).order_by(Customer.name).limit(500).all()
        
        return jsonify([{
            'id': c.id,
            'name': c.name,
            'email': c.email,
            'phone': c.phone
        } for c in customers])
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@security_bp.route('/api/saas/subscriptions/<int:sub_id>', methods=['GET'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_get_subscription(sub_id):
    """API: جلب تفاصيل اشتراك"""
    from models import SaaSSubscription, Customer
    
    try:
        sub = db.get_or_404(SaaSSubscription, sub_id)
        customer = db.session.get(Customer, sub.customer_id)
        
        # حساب الأيام المتبقية
        days_left = 0
        if sub.end_date:
            end_dt = make_aware(sub.end_date)
            delta_days = int((end_dt - datetime.now(timezone.utc)).days)
            days_left = delta_days if delta_days > 0 else 0
        
        return jsonify({
            'success': True,
            'subscription': {
                'id': sub.id,
                'customer_id': sub.customer_id,
                'customer_name': customer.name if customer else 'عميل محذوف',
                'plan_id': sub.plan_id,
                'plan_name': sub.plan.name if sub.plan else 'باقة محذوفة',
                'price': float(sub.plan.price_monthly) if sub.plan else 0,
                'status': sub.status,
                'start_date': sub.start_date.strftime('%Y-%m-%d') if sub.start_date else '',
                'end_date': sub.end_date.strftime('%Y-%m-%d') if sub.end_date else '',
                'days_left': days_left
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/subscriptions/<int:sub_id>', methods=['PUT'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_update_subscription(sub_id):
    """API: تحديث اشتراك"""
    from models import SaaSSubscription
    
    try:
        sub = db.get_or_404(SaaSSubscription, sub_id)
        data = request.get_json(silent=True) or {}
        
        if 'status' in data:
            sub.status = data['status']
        
        if 'end_date' in data and data['end_date']:
            sub.end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')
        
        db.session.commit()
        
        _log_owner_action('saas_subscription_update', target=sub.id, meta={'status': sub.status, 'end_date': sub.end_date.strftime('%Y-%m-%d') if sub.end_date else None})
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/api/saas/invoices/<int:invoice_id>/pdf', methods=['GET'])
@permission_required(SystemPermissions.MANAGE_SAAS)
def api_saas_invoice_pdf(invoice_id):
    """API: تحميل فاتورة PDF"""
    from models import SaaSInvoice, SaaSSubscription, Customer
    from flask import make_response
    from werkzeug.exceptions import HTTPException
    
    try:
        invoice = db.get_or_404(SaaSInvoice, invoice_id)
        subscription = db.session.get(SaaSSubscription, invoice.subscription_id)
        customer = db.session.get(Customer, subscription.customer_id) if subscription else None
        
        # إنشاء HTML للفاتورة
        html_content = f"""
        <!DOCTYPE html>
        <html dir="rtl">
        <head>
            <meta charset="UTF-8">
            <title>فاتورة #{invoice.id}</title>
            <style>
                body {{ font-family: Arial, sans-serif; direction: rtl; padding: 20px; }}
                .header {{ text-align: center; border-bottom: 3px solid #007bff; padding-bottom: 20px; }}
                .info {{ margin: 20px 0; }}
                .table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                .table th, .table td {{ border: 1px solid #ddd; padding: 12px; text-align: right; }}
                .table th {{ background: #f8f9fa; }}
                .total {{ font-size: 1.5rem; font-weight: bold; color: #007bff; }}
                .footer {{ margin-top: 40px; text-align: center; color: #666; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>فاتورة SaaS</h1>
                <p>رقم الفاتورة: #{invoice.id}</p>
            </div>
            
            <div class="info">
                <p><strong>العميل:</strong> {customer.name if customer else 'غير محدد'}</p>
                <p><strong>البريد:</strong> {customer.email if customer else '-'}</p>
                <p><strong>الباقة:</strong> {subscription.plan.name if subscription and subscription.plan else '-'}</p>
                <p><strong>تاريخ الإصدار:</strong> {invoice.created_at.strftime('%Y-%m-%d')}</p>
                <p><strong>تاريخ الاستحقاق:</strong> {invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '-'}</p>
            </div>
            
            <table class="table">
                <thead>
                    <tr>
                        <th>البيان</th>
                        <th>المبلغ</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td>اشتراك {subscription.plan.name if subscription and subscription.plan else 'SaaS'}</td>
                        <td class="total">{invoice.currency} {float(invoice.amount):,.2f}</td>
                    </tr>
                </tbody>
            </table>
            
            <div class="footer">
                <p>شكراً لثقتكم بنا | تم الإنشاء بواسطة SaaS Manager</p>
                <p>© 2025</p>
            </div>
        </body>
        </html>
        """
        try:
            from weasyprint import HTML
            pdf_bytes = HTML(string=html_content, base_url=request.url_root).write_pdf()
            filename = f"saas_invoice_{invoice.invoice_number or invoice.id}.pdf"
            return make_response((pdf_bytes, 200, {
                'Content-Type': 'application/pdf',
                'Content-Disposition': f'inline; filename="{filename}"'
            }))
        except Exception:
            response = make_response(html_content)
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
            response.headers['Content-Disposition'] = f'attachment; filename=invoice_{invoice.id}.html'
            return response
    except HTTPException:
        raise
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@security_bp.route('/')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def index():
    """
    👑 لوحة التحكم الأمنية الرئيسية - Owner's Security Dashboard
    
    مع Caching للإحصائيات (5 دقائق)
    """
    return render_template('security/index.html', stats=get_cached_security_stats(), recent=get_recent_suspicious_activities())


@security_bp.route('/index-old')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def index_old():
    """
    👑 لوحة التحكم الأمنية الرئيسية - Owner's Security Dashboard
    
    📋 الوصف:
        الصفحة الرئيسية للوحدة السرية - محدودة للمالك فقط
        
    📤 Response:
        HTML: templates/security/index.html
        
    🎯 الوظائف:
        ✅ إحصائيات الأمان الشاملة
        ✅ المستخدمين (إجمالي/نشطين/محظورين/متصلين)
        ✅ IPs & Countries المحظورة
        ✅ محاولات فشل الدخول (24h)
        ✅ الأنشطة المشبوهة
        ✅ صحة النظام
        ✅ روابط سريعة لجميع الأدوات
    
    📊 Quick Links:
        - مركز القيادة الموحد (7 مراكز + 41 وظيفة)
        - User Control (إدارة مستخدمين)
        - Database Manager (3 in 1)
        - SQL Console
        - Logs Viewer (6 أنواع)
        - Indexes Manager (115+ فهرس)
    
    🔒 Security:
        - صلاحيات محددة عبر PBAC
        - لا أحد يدخل بدون access_owner_dashboard
    """
    return redirect(url_for("security.index"))
    from datetime import datetime, timedelta, timezone
    
    # إحصائيات المستخدمين
    total_users = User.query.count()
    active_users = User.query.filter_by(is_active=True).count()
    blocked_users = User.query.filter_by(is_active=False).count()
    system_accounts = User.query.filter_by(is_system_account=True).count()
    
    # المتصلين الآن (آخر 15 دقيقة)
    threshold = datetime.now(timezone.utc) - timedelta(minutes=15)
    # حساب يدوي لتجنب مشاكل timezone في SQL
    all_users = User.query.filter(User.last_seen.isnot(None)).all()
    online_users = sum(1 for u in all_users if make_aware(u.last_seen) >= threshold)
    
    # محاولات فشل الدخول (آخر 24 ساعة)
    day_ago = datetime.now(timezone.utc) - timedelta(hours=24)
    from models import AuthAudit, AuthEvent
    failed_logins_24h = AuthAudit.query.filter(
        AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
        AuthAudit.created_at >= day_ago
    ).count()
    
    # Blocked IPs & Countries
    blocked_ips = _get_blocked_ips_count() if callable(locals().get('_get_blocked_ips_count')) else 0
    blocked_countries = _get_blocked_countries_count() if callable(locals().get('_get_blocked_countries_count')) else 0
    
    # أنشطة مشبوهة (محاولات فاشلة متكررة من نفس IP >= 5)
    suspicious_activities = 0
    try:
        suspicious_activities = db.session.query(
            func.count(AuthAudit.ip)
        ).filter(
            AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
            AuthAudit.created_at >= day_ago
        ).group_by(AuthAudit.ip).having(
            func.count(AuthAudit.ip) >= 5
        ).count()
    except Exception:
        pass
    
    # حجم قاعدة البيانات
    db_size = "N/A"
    try:
        size_row = db.session.execute(text("SELECT pg_size_pretty(pg_database_size(current_database()))")).fetchone()
        db_size = size_row[0] if size_row else "N/A"
    except Exception:
        pass
    
    # صحة النظام
    system_health = "ممتاز"
    if failed_logins_24h > 50:
        system_health = "تحذير"
    elif failed_logins_24h > 100:
        system_health = "خطر"
    
    stats = {
        'total_users': total_users,
        'active_users': active_users,
        'blocked_users': blocked_users,
        'system_accounts': system_accounts,
        'online_users': online_users,
        'blocked_ips': blocked_ips,
        'blocked_countries': blocked_countries,
        'failed_logins_24h': failed_logins_24h,
        'suspicious_activities': suspicious_activities,
        'db_size': db_size,
        'system_health': system_health,
        'active_sessions': online_users,
        'total_services': 40,
        'system_version': 'v5.0.0',
        'total_modules': '40+',
        'total_apis': 133,
        'total_indexes': 115  # تحديث: كان 89، الآن 115 بعد إضافة 26
    }
    
    # آخر الأنشطة المشبوهة
    recent_suspicious = []
    try:
        recent_suspicious = AuthAudit.query.filter(
            AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
            AuthAudit.created_at >= day_ago
        ).order_by(AuthAudit.created_at.desc()).limit(10).all()
    except Exception:
        pass
    
    return render_template('security/index.html', stats=stats, recent=recent_suspicious)


@security_bp.route('/block-ip', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def block_ip():
    """حظر IP معين"""
    if request.method == 'POST':
        ip = request.form.get('ip', '').strip()
        reason = request.form.get('reason', '').strip()
        duration = request.form.get('duration', '').strip()  # permanent, 1h, 24h, 7d, 30d
        
        if not ip:
            flash('❌ IP مطلوب', 'danger')
        else:
            _block_ip(ip, reason, duration)
            flash(f'✅ تم حظر IP: {ip}', 'success')
            return redirect(url_for('security.blocked_ips'))
    
    return render_template('security/block_ip.html')


@security_bp.route('/blocked-ips')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def blocked_ips():
    """قائمة IPs المحظورة"""
    blocked = _get_all_blocked_ips()
    return render_template('security/blocked_ips.html', blocked=blocked)


@security_bp.route('/unblock-ip/<ip>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def unblock_ip(ip):
    """إلغاء حظر IP"""
    _unblock_ip(ip)
    flash(f'✅ تم إلغاء حظر IP: {ip}', 'success')
    return redirect(url_for('security.blocked_ips'))


@security_bp.route('/block-country', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def block_country():
    """حظر دولة معينة"""
    if request.method == 'POST':
        country_code = request.form.get('country_code', '').strip().upper()
        reason = request.form.get('reason', '').strip()
        
        if not country_code or len(country_code) != 2:
            flash('❌ كود الدولة مطلوب (مثال: US, IL)', 'danger')
        else:
            _block_country(country_code, reason)
            flash(f'✅ تم حظر الدولة: {country_code}', 'success')
            return redirect(url_for('security.blocked_countries'))
    
    return render_template('security/block_country.html')


@security_bp.route('/blocked-countries')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def blocked_countries():
    """قائمة الدول المحظورة"""
    blocked = _get_all_blocked_countries()
    return render_template('security/blocked_countries.html', blocked=blocked)


@security_bp.route('/block-user/<int:user_id>', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_USERS)
def block_user(user_id):
    """حظر مستخدم معين"""
    user = db.get_or_404(User, user_id)
    
    if utils.is_super() and user.id == current_user.id:
        flash('❌ لا يمكنك حظر نفسك!', 'danger')
    else:
        user.is_active = False
        db.session.commit()
        flash(f'✅ تم حظر المستخدم: {user.username}', 'success')
    
    return redirect(url_for('users_bp.list_users'))


@security_bp.route('/system-cleanup', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def system_cleanup():
    """
    🧹 System Cleanup & Maintenance
    - تنظيف الكاش والملفات المؤقتة (Maintenance)
    - تنظيف جداول النظام (Format/Reset)
    """
    
    # Calculate maintenance stats
    cleanup_stats = {
        'temp_files_count': 0, # Placeholder
        'logs_size': '0 MB',   # Placeholder
        'cache_items': 0       # Placeholder
    }
    try:
        # cleanup_stats['cache_items'] = len(cache.cache._cache) if hasattr(cache.cache, '_cache') else 0
        pass
    except:
        pass

    if request.method == 'POST':
        action = (request.form.get('action') or "").strip()
        confirm = (request.form.get('confirm') or "").strip()
        tables = request.form.getlist('tables')

        # 1. Handle Maintenance Actions (آمنة – بدون حذف بيانات أعمال)
        if action in {'clear_cache', 'cleanup_temp', 'purge_logs', 'optimize_db'}:
            if action == 'clear_cache':
                try:
                    cache.clear()
                    flash('تم مسح ذاكرة التخزين المؤقت بنجاح.', 'success')
                except Exception as e:
                    flash(f'فشل مسح الكاش: {e}', 'danger')
            elif action == 'cleanup_temp':
                flash('تم تنظيف الملفات المؤقتة (محاكاة).', 'success')
            elif action == 'purge_logs':
                flash('تم حذف السجلات القديمة (محاكاة).', 'success')
            elif action == 'optimize_db':
                flash('تم تحسين قاعدة البيانات (محاكاة).', 'success')

            return redirect(url_for('security.system_cleanup', tab='maintenance'))

        # 2. Handle Data Reset (Format) Actions – أخطر جزء
        if confirm or tables:
            if confirm != 'FORMAT_SYSTEM':
                flash('❌ يجب كتابة "FORMAT_SYSTEM" للتأكيد', 'danger')
            elif not tables:
                flash('❌ اختر جدول واحد على الأقل', 'danger')
            else:
                try:
                    result = _cleanup_tables(tables)

                    # إذا تم تحديد جميع الجداول (تصفير شامل)، لا نقوم بأي drop_all داخل سيرفر التشغيل
                    is_full_reset = len(tables) >= 50 or 'users' in tables or 'users_except_first_super' in tables

                    if is_full_reset:
                        try:
                            # فقط إعادة تهيئة البيانات الأساسية داخل التطبيق بشكل آمن
                            from app import bootstrap_database
                            bootstrap_database()

                            # تشغيل سكربتات post_reset الآمنة (seed / run فقط)
                            import glob
                            import importlib.util

                            post_reset_dir = os.path.join(current_app.root_path, 'post_reset_scripts')
                            if os.path.exists(post_reset_dir):
                                script_files = glob.glob(os.path.join(post_reset_dir, '*.py'))
                                for script_file in sorted(script_files):
                                    script_name = os.path.basename(script_file)
                                    if script_name.startswith('__') or script_name == 'clean_and_seed.py':
                                        continue

                                    try:
                                        spec = importlib.util.spec_from_file_location(f"post_reset_{script_name[:-3]}", script_file)
                                        module = importlib.util.module_from_spec(spec)
                                        spec.loader.exec_module(module)

                                        if hasattr(module, 'seed_permissions'):
                                            module.seed_permissions(current_app)
                                        elif hasattr(module, 'seed_accounts'):
                                            module.seed_accounts(current_app)
                                        elif hasattr(module, 'seed'):
                                            module.seed(current_app)
                                        elif hasattr(module, 'run'):
                                            module.run(current_app)
                                    except Exception as seed_err:
                                        current_app.logger.error(f"Seed error {script_name}: {seed_err}")

                            flash(f'✅ تم تنظيف {result.get("cleaned", 0)} جدول وإعادة تهيئة النظام بنجاح', 'success')
                        except Exception as e:
                            current_app.logger.error(f"Reset Error: {e}")
                            flash(f'✅ تم التنظيف ولكن حدث خطأ في التهيئة: {e}', 'warning')
                    else:
                        flash(f'✅ تم تنظيف {result.get("cleaned", 0)} جدول', 'success')

                except Exception as e:
                    flash(f'❌ حدث خطأ أثناء التنظيف: {e}', 'danger')

            return redirect(url_for('security.system_cleanup', tab='reset'))
    
    # قائمة الجداول القابلة للتنظيف
    try:
        cleanable_tables = _get_cleanable_tables()
    except NameError:
        # Fallback if helper not found/defined yet
        cleanable_tables = []

    category_map = defaultdict(list)
    for table in cleanable_tables:
        category_map[table.get('category', 'أخرى')].append(table)
    
    danger_counts = Counter(table.get('danger', 'medium') for table in cleanable_tables)
    
    category_order = [
        'المستخدمين والأدوار',
        'الجهات',
        'العمليات المالية',
        'العمليات المحاسبية',
        'المبيعات والصيانة',
        'المخزون والمنتجات',
        'السجلات واللوجات',
        'التسوق الإلكتروني',
        'أخرى',
    ]
    danger_zone_categories = {'الجهات', 'المخزون والمنتجات'}
    category_icons = {
        'المستخدمين والأدوار': 'fas fa-users text-primary',
        'الجهات': 'fas fa-user-friends text-danger',
        'العمليات المالية': 'fas fa-money-bill-wave text-success',
        'العمليات المحاسبية': 'fas fa-book text-info',
        'المبيعات والصيانة': 'fas fa-shopping-cart text-warning',
        'المخزون والمنتجات': 'fas fa-warehouse text-danger',
        'السجلات واللوجات': 'fas fa-clipboard-list text-secondary',
        'التسوق الإلكتروني': 'fas fa-shopping-bag text-primary',
        'أخرى': 'fas fa-folder-open text-muted',
    }
    category_descriptions = {
        'المستخدمين والأدوار': 'حسابات الدخول والأذونات المرتبطة بها.',
        'الجهات': 'العملاء والموردون والشركاء والموظفون.',
        'العمليات المالية': 'المدفوعات والشيكات والمصاريف.',
        'العمليات المحاسبية': 'دفعات القيود والحسابات العامة.',
        'المبيعات والصيانة': 'المبيعات، الفواتير وطلبات الصيانة.',
        'المخزون والمنتجات': 'المنتجات والمخازن والتبادلات.',
        'السجلات واللوجات': 'السجلات، التنبيهات، وإشعارات النظام.',
        'التسوق الإلكتروني': 'الحجوزات والمتاجر الإلكترونية.',
        'أخرى': 'جداول مساندة أو ثانوية.',
    }
    
    def build_category_payload(name, items):
        sorted_tables = sorted(items, key=lambda t: t.get('display', ''))
        return {
            'name': name,
            'icon': category_icons.get(name, 'fas fa-folder'),
            'description': category_descriptions.get(name, ''),
            'tables': sorted_tables,
            'total': len(sorted_tables),
            'high_danger': sum(1 for t in sorted_tables if t.get('danger') == 'high'),
            'danger_zone': name in danger_zone_categories,
        }
    
    ordered_categories = []
    for name in category_order:
        if name in category_map:
            ordered_categories.append(build_category_payload(name, category_map.pop(name)))
    for name in sorted(category_map.keys()):
        ordered_categories.append(build_category_payload(name, category_map[name]))
    
    stats = {
        'total_tables': len(cleanable_tables),
        'high_danger_tables': danger_counts.get('high', 0),
        'medium_danger_tables': danger_counts.get('medium', 0),
        'low_danger_tables': danger_counts.get('low', 0),
    }
    
    return render_template(
        'security/system_cleanup.html',
        tables=cleanable_tables,
        categories=ordered_categories,
        stats=stats,
        cleanup_stats=cleanup_stats
    )





@security_bp.route('/security-center')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def security_center():
    """
    🛡️ Security & Monitoring Center - 4 في 1
    - مراقبة فورية (Live Monitoring)
    - جدار الحماية (Firewall)
    - التنبيهات (Notifications)
    - النشاط (Activity Timeline)
    """
    tab = request.args.get('tab', 'monitoring')
    
    try:
        from models import User
        active_users = User.query.filter(User.last_seen >= datetime.now(timezone.utc) - timedelta(minutes=30)).count()
    except Exception:
        active_users = 0
    
    try:
        failed_login_count = AuditLog.query.filter(
            AuditLog.action.like('%failed%'),
            AuditLog.created_at >= datetime.now(timezone.utc) - timedelta(hours=24)
        ).count()
    except Exception:
        failed_login_count = 0
    
    security_stats = {
        'online_users': active_users,
        'blocked_ips': BlockedIP.query.count() if 'BlockedIP' in dir() else 0,
        'failed_logins': failed_login_count,
        'active_sessions': User.query.filter(User.last_seen.isnot(None)).count() if 'User' in dir() else 1,
        'threats_detected': 0,
        'patterns_found': 0,
        'notifications': 0
    }
    
    recent_activities = []
    blocked_ips = []
    patterns = []
    notifications = []
    recent_audit_logs = []
    integrations_data = None
    
    if tab == 'firewall':
        blocked_ips = BlockedIP.query.order_by(BlockedIP.created_at.desc()).limit(50).all() if 'BlockedIP' in dir() else []
    elif tab == 'patterns':
        patterns = _detect_suspicious_patterns()
    elif tab == 'activity':
        recent_audit_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(20).all()
    elif tab == 'notifications':
        integrations_data = {
            'stripe': {
                'enabled': _get_setting('stripe_enabled', False),
            },
            'paypal': {
                'enabled': _get_setting('paypal_enabled', False),
            },
            'sms': {
                'enabled': _get_setting('sms_enabled', False),
            },
            'email': {
                'enabled': _get_setting('email_enabled', True) and bool(_get_setting('smtp_host')),
                'smtp_host': _get_setting('smtp_host', 'غير مهيأ'),
            },
        }

    stats = get_cached_security_stats()
    
    # Map stats to security_stats expected format
    security_stats = {
        'online_users': stats.get('online_users', 0),
        'blocked_ips': stats.get('blocked_ips', 0),
        'failed_logins': stats.get('failed_logins_24h', 0),
        'active_sessions': stats.get('active_sessions', 0),
        'threats_detected': 0,
        'patterns_found': 0,
        'notifications': 0
    }
    
    return render_template('security/security_center.html',
                          active_tab=tab,
                          security_stats=security_stats,
                          recent_activities=recent_activities,
                          blocked_ips=blocked_ips,
                          patterns=patterns,
                          notifications=notifications,
                          recent_audit_logs=recent_audit_logs,
                          integrations=integrations_data,
                          stats=stats)


@security_bp.route('/ultimate-control', methods=['GET'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def ultimate_control():
    return render_template('security/ultimate_control.html')


def _log_training_event(event_type, user_id, details=None):
    """تسجيل حدث تدريب - محسّن"""
    try:
        from AI.engine.ai_knowledge import TRAINING_LOG_FILE
        import os
        
        os.makedirs('instance', exist_ok=True)
        
        logs = []
        if os.path.exists(TRAINING_LOG_FILE):
            try:
                with open(TRAINING_LOG_FILE, 'r', encoding='utf-8') as f:
                    logs = json.load(f)
            except Exception:
                logs = []
        
        log_entry = {
            'event': event_type,
            'user_id': user_id,
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
        
        if details:
            log_entry['details'] = details
        
        logs.append(log_entry)
        logs = logs[-50:]
        
        with open(TRAINING_LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
    except Exception as e:
        current_app.logger.warning(f"⚠️ فشل تسجيل حدث التدريب: {str(e)}")


def _load_training_logs():
    """تحميل سجل التدريب"""
    try:
        from AI.engine.ai_knowledge import TRAINING_LOG_FILE
        import os
        
        if os.path.exists(TRAINING_LOG_FILE):
            with open(TRAINING_LOG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []
    except Exception:
        return []


@security_bp.route('/database-manager', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def database_manager():
    """
    🗄️ مركز التحكم الشامل بقاعدة البيانات - Database Control Center
    
    📋 الوصف:
        وحدة موحدة شاملة 11-في-1 تجمع جميع أدوات إدارة قاعدة البيانات
    
    📥 Parameters:
        - tab (str): browse|edit|schema|indexes|logs|sql|python|maintenance|restore|tools|archive (default: browse)
        - table (str): اسم الجدول للعمل عليه (optional)
        - limit (int): عدد السجلات (optional)
        - log_type (str): نوع اللوج (optional)
    
    📤 Response:
        HTML: templates/security/database_manager.html
        
    🎯 التبويبات (11 تبويب):
        ✅ Browse: تصفح الجداول والبيانات
        ✅ Edit: تحرير مباشر للبيانات
        ✅ Schema: هيكل الجدول + الأعمدة
        ✅ Indexes: إدارة الفهارس الكاملة
        ✅ Logs: سجلات النظام والتدقيق
        ✅ SQL Console: تنفيذ استعلامات SQL
        ✅ Python Console: تنفيذ كود Python
        ✅ Maintenance: صيانة + VACUUM
        ✅ Restore: استعادة قاعدة البيانات
        ✅ Tools: أدوات إضافية (Decrypt, Error Tracker)
        ✅ Archive: إدارة الأرشيفات والبيانات المحذوفة
    
    🔗 Related APIs:
        - جميع APIs السابقة متاحة
    
    💡 Usage Examples:
        /database-manager?tab=browse
        /database-manager?tab=indexes
        /database-manager?tab=sql
        /database-manager?tab=logs&log_type=audit
    
    🔒 Security:
        - صلاحيات محددة عبر PBAC
        - Full audit trail
        - CSRF protection
    """
    tab = request.args.get('tab', 'browse')
    selected_table = request.args.get('table')
    log_type = request.args.get('log_type', 'audit')

    app_env = str(current_app.config.get("APP_ENV", "production") or "production").strip().lower()
    is_production = (not bool(current_app.config.get("DEBUG", False))) and (app_env not in {"dev", "development", "local"})
    allow_dangerous_consoles = str(os.environ.get("ALLOW_DANGEROUS_CONSOLES", "") or "").strip().lower() in {"1", "true", "yes", "y", "on"}
    if is_production and (tab in {"sql", "python"}) and (not allow_dangerous_consoles):
        flash('⛔ تم تعطيل SQL/Python Console في بيئة الإنتاج للأمان', 'danger')
        return redirect(url_for('security.database_manager', tab='browse'))
    
    # ==== البيانات الأساسية (للجميع) ====
    tables = _get_all_tables()
    table_counts = {}
    for table in tables:
        try:
            count_query = text(f"SELECT COUNT(*) as count FROM {table}")
            result = db.session.execute(count_query).fetchone()
            table_counts[table] = result[0] if result else 0
        except Exception:
            table_counts[table] = 0
    
    # ==== بيانات خاصة بكل تبويب ====
    data = []
    columns = []
    table_info = []
    indexes_data = []
    indexes_stats = {}
    audit_logs = []
    system_logs = ""
    error_logs = []
    sql_result = None
    sql_error = None
    python_result = None
    python_error = None
    log_files = []
    errors = []
    error_stats = {}
    decrypt_result = None
    all_users = []
    primary_key_column = 'id'
    
    # === 1) Browse & Edit & Schema ===
    if tab in ['browse', 'edit', 'schema'] and selected_table:
        data, columns = _browse_table(selected_table, limit=1000 if tab == 'edit' else 100)
        table_info = _get_table_info(selected_table)
        
        for col_info in table_info:
            if col_info.get('pk', 0) == 1:
                primary_key_column = col_info.get('name', 'id')
                break
        if not primary_key_column or primary_key_column == 'id':
            primary_key_column = columns[0] if columns else 'id'
    
    # === 2) Indexes ===
    if tab == 'indexes':
        try:
            from sqlalchemy import inspect
            inspector = inspect(db.engine)
            indexes_data = []
            for table in sorted(tables):
                try:
                    cols = inspector.get_columns(table)
                    idxs = inspector.get_indexes(table)
                    fks = inspector.get_foreign_keys(table)
                    
                    processed_indexes = []
                    for idx in idxs:
                        column_names = idx.get('column_names', [])
                        if not column_names:
                            continue
                        processed_indexes.append({
                            'name': idx.get('name', 'unknown'),
                            'columns': column_names if isinstance(column_names, list) else [column_names],
                            'unique': bool(idx.get('unique', False))
                        })
                    
                    pk_constraint = inspector.get_pk_constraint(table)
                    if pk_constraint and pk_constraint.get('constrained_columns'):
                        pk_columns = pk_constraint['constrained_columns']
                        pk_index_name = f"pk_{table}"
                        if not any(idx['name'] == pk_index_name for idx in processed_indexes):
                            processed_indexes.insert(0, {
                                'name': 'PRIMARY KEY',
                                'columns': pk_columns if isinstance(pk_columns, list) else [pk_columns],
                                'unique': True
                            })
                    
                    indexes_data.append({
                        'name': table,
                        'columns_count': len(cols),
                        'indexes_count': len(processed_indexes),
                        'fk_count': len(fks),
                        'columns': [{'name': c['name'], 'type': str(c['type'])} for c in cols],
                        'indexes': processed_indexes,
                        'foreign_keys': [{'columns': fk.get('constrained_columns', []), 'ref_table': fk.get('referred_table', '')} for fk in fks]
                    })
                except Exception as e:
                    current_app.logger.error(f"Error processing table {table} for indexes: {e}")
                    indexes_data.append({
                        'name': table,
                        'columns_count': 0,
                        'indexes_count': 0,
                        'fk_count': 0,
                        'columns': [],
                        'indexes': [],
                        'foreign_keys': []
                    })
            
            total_indexes = sum([t['indexes_count'] for t in indexes_data]) if indexes_data else 0
            total_columns = sum([t['columns_count'] for t in indexes_data]) if indexes_data else 0
            
            indexes_stats = {
                'total_tables': len(tables),
                'total_indexes': total_indexes,
                'total_columns': total_columns,
                'tables_without_indexes': len([t for t in indexes_data if t['indexes_count'] == 0]) if indexes_data else len(tables),
                'avg_indexes_per_table': round(total_indexes / len(tables), 2) if tables else 0
            }
        except Exception as e:
            current_app.logger.error(f"Error loading indexes: {e}")
            indexes_data = []
            indexes_stats = {
                'total_tables': len(tables),
                'total_indexes': 0,
                'total_columns': 0,
                'tables_without_indexes': len(tables),
                'avg_indexes_per_table': 0
            }
    
    # === 3) Logs ===
    if tab == 'logs':
        # Audit logs
        audit_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(200).all()
        
        # System logs
        try:
            if os.path.exists('logs/app.log'):
                with open('logs/app.log', 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    system_logs = ''.join(lines[-500:])
        except Exception:
            system_logs = "تعذر قراءة ملف السجلات"
        
        # Error logs
        error_logs = AuditLog.query.filter(
            AuditLog.action.like('%failed%') | AuditLog.action.like('%error%')
        ).order_by(AuditLog.created_at.desc()).limit(100).all()
        
        # Log files list
        log_files = _get_available_log_files()
        
        all_users = User.query.order_by(User.username).all()
    
    # === 4) SQL Console ===
    if tab == 'sql' and request.method == 'POST':
        app_env = str(current_app.config.get("ENV") or os.environ.get("FLASK_ENV") or os.environ.get("APP_ENV") or "").strip().lower()
        is_production = (not bool(current_app.config.get("DEBUG", False))) and (app_env not in {"dev", "development", "local"})
        allow_sql_console = str(os.environ.get("ENABLE_SQL_CONSOLE", "") or "").strip().lower() in {"1", "true", "yes", "y", "on"}
        allow_sql_write = str(os.environ.get("ENABLE_SQL_WRITE_CONSOLE", "") or "").strip().lower() in {"1", "true", "yes", "y", "on"}
        if (is_production and not allow_sql_console) or (not utils.is_super()):
            flash('⚠️ غير مصرح - وحدة SQL معطلة', 'danger')
            _log_owner_action(
                'db_sql_denied',
                target='database_manager',
                meta={'env': app_env}
            )
            return redirect(url_for('security.database_manager', tab='browse'))
        sql_query = request.form.get('sql_query', '').strip()
        if not sql_query:
            flash('⚠️ الرجاء إدخال استعلام SQL', 'warning')
            return redirect(url_for('security.database_manager', tab='sql'))
        statements = [s.strip() for s in sql_query.split(";") if s.strip()]
        if len(statements) != 1:
            flash('⛔ غير مسموح بتنفيذ عدة أوامر في نفس الطلب', 'danger')
            return redirect(url_for('security.database_manager', tab='sql'))
        first_token = (statements[0].lstrip().split(None, 1) or [""])[0].upper()
        readonly_tokens = {"SELECT", "WITH", "EXPLAIN", "SHOW"}
        if (first_token not in readonly_tokens) and (not allow_sql_write):
            flash('⛔ مسموح فقط باستعلامات القراءة (SELECT/WITH/EXPLAIN/SHOW)', 'danger')
            _log_owner_action(
                'db_sql_blocked_write',
                target='database_manager',
                meta={'statement': statements[0][:200]}
            )
            return redirect(url_for('security.database_manager', tab='sql'))
        try:
            result_proxy = db.session.execute(text(statements[0]))
            try:
                rows = result_proxy.fetchall()
                cols = result_proxy.keys() if hasattr(result_proxy, 'keys') else []
                sql_result = {
                    'columns': list(cols),
                    'rows': [list(row) for row in rows],
                    'count': len(rows)
                }
                _log_owner_action(
                    'db_sql_executed',
                    target='database_manager',
                    meta={'statement': statements[0][:200], 'rows': len(rows)}
                )
            except Exception:
                db.session.commit()
                sql_result = {'message': 'تم تنفيذ الاستعلام بنجاح'}
                _log_owner_action(
                    'db_sql_executed_write',
                    target='database_manager',
                    meta={'statement': statements[0][:200]}
                )
        except Exception as e:
            sql_error = str(e)
            db.session.rollback()
            _log_owner_action(
                'db_sql_error',
                target='database_manager',
                meta={'error': str(e)[:200]}
            )
    
    # === 5) Python Console ===
    if tab == 'python' and request.method == 'POST':
        app_env = str(current_app.config.get("ENV") or os.environ.get("FLASK_ENV") or os.environ.get("APP_ENV") or "").strip().lower()
        is_production = (not bool(current_app.config.get("DEBUG", False))) and (app_env not in {"dev", "development", "local"})
        allow_python_console = str(os.environ.get("ENABLE_PYTHON_CONSOLE", "") or "").strip().lower() in {"1", "true", "yes", "y", "on"}
        if (is_production and not allow_python_console) or (not utils.is_super()):
            flash('⚠️ غير مصرح - وحدة Python معطلة', 'danger')
            _log_owner_action(
                'db_python_denied',
                target='database_manager',
                meta={'env': app_env}
            )
            return redirect(url_for('security.database_manager', tab='browse'))
        
        python_code = request.form.get('python_code', '').strip()
        if not python_code:
            flash('⚠️ الرجاء إدخال كود Python', 'warning')
            return redirect(url_for('security.database_manager', tab='python'))
        
        dangerous_keywords = ['import os', 'import sys', 'import subprocess', '__import__', 'eval(', 'compile(', 'open(', 'file(', 'input(', 'execfile(']
        if any(keyword in python_code for keyword in dangerous_keywords):
            flash('⛔ كود خطر - يحتوي على عمليات محظورة', 'danger')
            python_error = 'كود خطر محظور'
            _log_owner_action(
                'db_python_blocked',
                target='database_manager',
                meta={'snippet': python_code[:200]}
            )
        else:
            try:
                safe_builtins = {
                    'print': print,
                    'len': len,
                    'str': str,
                    'int': int,
                    'float': float,
                    'list': list,
                    'dict': dict,
                    'sum': sum,
                    'max': max,
                    'min': min,
                    'range': range,
                }
                local_vars = {
                    'db': db,
                    'User': User,
                    'AuditLog': AuditLog,
                    'current_user': current_user,
                    'datetime': datetime,
                    'timezone': timezone
                }
                exec(python_code, {'__builtins__': safe_builtins}, local_vars)
                python_result = local_vars.get('output', 'تم التنفيذ بنجاح')
                
                utils.log_audit('PYTHON_EXEC', None, 'EXECUTED', details=f'Executed: {python_code[:100]}...')
                _log_owner_action(
                    'db_python_executed',
                    target='database_manager',
                    meta={'snippet': python_code[:200]}
                )
            except Exception as e:
                python_error = str(e)
                _log_owner_action(
                    'db_python_error',
                    target='database_manager',
                    meta={'error': str(e)[:200]}
                )
    
    # === 6) Tools (Error Tracker) ===
    if tab == 'tools':
        errors = _get_recent_errors(100)
        error_stats = _get_error_statistics()
        
        # Decrypt tool
        if request.method == 'POST' and request.form.get('encrypted_data'):
            encrypted_data = request.form.get('encrypted_data', '').strip()
            decrypt_type = request.form.get('decrypt_type', 'auto')
            decrypt_result = _decrypt_data(encrypted_data, decrypt_type)
    
    return render_template('security/database_manager.html',
                          # عام
                          tables=tables,
                          table_counts=table_counts,
                          active_tab=tab,
                          selected_table=selected_table,
                          # Browse/Edit/Schema
                          data=data,
                          columns=columns,
                          table_info=table_info,
                          primary_key_column=primary_key_column,
                          # Indexes
                          indexes_data=indexes_data if indexes_data else [],
                          indexes_stats=indexes_stats if indexes_stats else {
                              'total_tables': len(tables),
                              'total_indexes': 0,
                              'total_columns': 0,
                              'tables_without_indexes': len(tables),
                              'avg_indexes_per_table': 0
                          },
                          # Logs
                          audit_logs=audit_logs,
                          system_logs=system_logs,
                          error_logs=error_logs,
                          log_files=log_files,
                          log_type=log_type,
                          all_users=all_users,
                          # SQL
                          sql_result=sql_result,
                          sql_error=sql_error,
                          # Python
                          python_result=python_result,
                          python_error=python_error,
                          # Tools
                          errors=errors,
                          error_stats=error_stats,
                          decrypt_result=decrypt_result)






@security_bp.route('/users-center', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def users_center():
    """
    👥 Users & Permissions Center - 2 في 1
    - التحكم بالمستخدمين (User Control)
    - إدارة الصلاحيات (Permissions)
    """
    from models import User, Role, Permission
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'bulk_delete':
            # التحقق من الصلاحيات (تم التحقق منها بالفعل عبر permission_required لكن زيادة أمان)
            if not (current_user.has_role('owner') or current_user.can('manage_users')):
                flash('ليس لديك صلاحية لحذف المستخدمين', 'danger')
                return redirect(url_for('security.users_center', tab='users'))

            user_ids = request.form.getlist('user_ids')
            if user_ids:
                deleted_count = 0
                for user_id in user_ids:
                    # تخطي المستخدم الحالي
                    if str(user_id) == str(current_user.id):
                        continue
                        
                    user = db.session.get(User, user_id)
                    if user and user.username != '__OWNER__':
                        # لا يمكن حذف حسابات النظام إلا للمالك
                        if getattr(user, 'is_system_account', False) and not current_user.has_role('owner'):
                            continue
                            
                        db.session.delete(user)
                        deleted_count += 1
                
                if deleted_count > 0:
                    db.session.commit()
                    flash(f'✅ تم حذف {deleted_count} مستخدم بنجاح', 'success')
                else:
                    flash('⚠️ لم يتم حذف أي مستخدم (قد تكون حاولت حذف نفسك أو حساب محمي)', 'warning')
            else:
                flash('⚠️ الرجاء تحديد مستخدمين للحذف', 'warning')
                
            return redirect(url_for('security.users_center', tab='users'))

    tab = request.args.get('tab', 'users')
    
    users_data = {
        'total': User.query.count(),
        'active': User.query.filter_by(is_active=True).count(),
        'blocked': User.query.filter_by(is_active=False).count(),
        'online': User.query.filter(User.last_seen >= datetime.now(timezone.utc) - timedelta(minutes=30)).count(),
    }
    
    users_list = User.query.order_by(User.created_at.desc()).limit(20).all()
    
    roles_data = {
        'total': Role.query.count(),
    }
    
    permissions_data = {
        'total': Permission.query.count(),
        'protected': Permission.query.filter_by(is_protected=True).count()
    }
    
    roles_list = Role.query.all()
    permissions_list = Permission.query.order_by(Permission.module, Permission.name).all()
    
    stats = get_cached_security_stats()
    return render_template(
        'security/users_center.html',
        active_tab=tab,
        stats=stats,
        users_data=users_data,
        users_list=users_list,
        roles_data=roles_data,
        permissions_data=permissions_data,
        roles_list=roles_list,
        permissions_list=permissions_list,
        total_users=users_data.get('total', 0),
    )


@security_bp.route('/settings-center', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def settings_center():
    """
    ⚙️ Settings & Customization Center - 8 في 1
    - إعدادات النظام + ثوابت + تكوين
    - العلامة التجارية + المظهر + الثيمات + الشعارات
    - الوضع الليلي
    - الفروع والمواقع
    """
    from models import SystemSettings, Branch
    
    tab = request.args.get('tab', 'system')
    
    if request.method == 'POST':
        action = request.form.get('action')
        active_tab = request.form.get('active_tab')
        
        # 1. Handle explicit, quick inline updates
        if action == 'update_setting':
            key = (request.form.get('key') or "").strip()
            value = request.form.get('value')
            if key in ALLOWED_DIRECT_SETTINGS_KEYS:
                SystemSettings.set_setting(key, value)
                flash(f'✅ تم تحديث {key} بنجاح', 'success')
            else:
                flash('⚠️ هذا الإعداد غير مسموح بتعديله مباشرة من هذه الشاشة.', 'warning')
                
        elif action == 'update_branding':
            # 1. System Identity (Appearance Panel / System Branding)
            system_name = (request.form.get('system_name') or "").strip()
            company_name = (request.form.get('company_name') or "").strip()
            login_title = (request.form.get('login_title') or "").strip()
            login_subtitle = (request.form.get('login_subtitle') or "").strip()
            footer_text = (request.form.get('footer_text') or "").strip()

            if system_name:
                SystemSettings.set_setting('system_name', system_name)
            if company_name:
                SystemSettings.set_setting('company_name', company_name)
            if login_title:
                SystemSettings.set_setting('login_title', login_title)
            if login_subtitle:
                SystemSettings.set_setting('login_subtitle', login_subtitle)
            if footer_text:
                SystemSettings.set_setting('footer_text', footer_text)

            # 2. Theme & Colors (safe keys only)
            primary_color = (request.form.get('primary_color') or "").strip()
            secondary_color = (request.form.get('secondary_color') or "").strip()
            sidebar_bg = (request.form.get('sidebar_bg') or "").strip()
            sidebar_text = (request.form.get('sidebar_text') or "").strip()

            if primary_color:
                SystemSettings.set_setting('primary_color', primary_color)
            if secondary_color:
                SystemSettings.set_setting('secondary_color', secondary_color)
            if sidebar_bg:
                SystemSettings.set_setting('sidebar_bg', sidebar_bg)
            if sidebar_text:
                SystemSettings.set_setting('sidebar_text', sidebar_text)

            flash('✅ تم تحديث مظهر وهوية النظام بنجاح', 'success')

        # 2. Handle tab-based bulk updates (from system_settings.html)
        elif active_tab == 'general':
            SystemSettings.set_setting('maintenance_mode', request.form.get('maintenance_mode') == 'on', data_type='boolean')
            SystemSettings.set_setting('registration_enabled', request.form.get('registration_enabled') == 'on', data_type='boolean')
            SystemSettings.set_setting('api_enabled', request.form.get('api_enabled') == 'on', data_type='boolean')
            flash('✅ تم حفظ الإعدادات العامة بنجاح', 'success')
            
        elif active_tab == 'advanced':
            SystemSettings.set_setting('SESSION_TIMEOUT', request.form.get('session_timeout'), data_type='integer')
            SystemSettings.set_setting('MAX_LOGIN_ATTEMPTS', request.form.get('max_login_attempts'), data_type='integer')
            SystemSettings.set_setting('PASSWORD_MIN_LENGTH', request.form.get('password_min_length'), data_type='integer')
            
            SystemSettings.set_setting('AUTO_BACKUP_ENABLED', request.form.get('auto_backup_enabled') == 'on', data_type='boolean')
            SystemSettings.set_setting('BACKUP_INTERVAL_HOURS', request.form.get('backup_interval_hours'), data_type='integer')
            
            SystemSettings.set_setting('ENABLE_EMAIL_NOTIFICATIONS', request.form.get('enable_email_notifications') == 'on', data_type='boolean')
            SystemSettings.set_setting('ENABLE_SMS_NOTIFICATIONS', request.form.get('enable_sms_notifications') == 'on', data_type='boolean')
            flash('✅ تم حفظ التكوينات المتقدمة بنجاح', 'success')
            
        elif active_tab == 'company':
            SystemSettings.set_setting('COMPANY_ADDRESS', request.form.get('company_address'))
            SystemSettings.set_setting('COMPANY_PHONE', request.form.get('company_phone'))
            SystemSettings.set_setting('COMPANY_EMAIL', request.form.get('company_email'))
            SystemSettings.set_setting('TAX_NUMBER', request.form.get('tax_number'))
            SystemSettings.set_setting('CURRENCY_SYMBOL', request.form.get('currency_symbol'))
            SystemSettings.set_setting('TIMEZONE', request.form.get('timezone'))
            flash('✅ تم حفظ بيانات الشركة بنجاح', 'success')
        
        db.session.commit()
        return redirect(url_for('security.settings_center', tab=tab))
    
    settings_list = SystemSettings.query.order_by(SystemSettings.key).limit(50).all()
    
    settings_stats = {
        'total': SystemSettings.query.count(),
        'public': SystemSettings.query.filter_by(is_public=True).count(),
        'private': SystemSettings.query.filter_by(is_public=False).count(),
    }
    
    branding_settings = {
        'system_name': _get_setting('system_name', 'نظام الحازم'),
        'company_name': _get_setting('company_name', 'اسم الشركة'),
        'company_logo': _get_setting('custom_logo', ''),
        'custom_favicon': _get_setting('custom_favicon', ''),
        'login_title': _get_setting('login_title', ''),
        'login_subtitle': _get_setting('login_subtitle', ''),
        'footer_text': _get_setting('footer_text', ''),
        
        # Colors
        'primary_color': _get_setting('primary_color', '#007bff'),
        'secondary_color': _get_setting('secondary_color', '#6c757d'),
        'sidebar_bg': _get_setting('sidebar_bg', '#343a40'),
        'sidebar_text': _get_setting('sidebar_text', '#ffffff'),
    }
    
    branches_data = {
        'total': Branch.query.count(),
        'active': Branch.query.filter_by(is_active=True).count(),
    }
    
    branches_list = Branch.query.all()
    
    stats = get_cached_security_stats()
    return render_template('security/settings_center.html', 
                          active_tab=tab, 
                          stats=stats,
                          settings_list=settings_list,
                          settings_stats=settings_stats,
                          branding_settings=branding_settings,
                          branches_data=branches_data,
                          branches_list=branches_list)


@security_bp.route('/reports-center')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def reports_center():
    """
    📊 Reports & Performance Center - 3 في 1
    - التقارير الإدارية
    - لوحة المراقبة الشاملة
    - مراقبة الأداء
    """
    tab = request.args.get('tab', 'reports')
    if tab not in ('reports', 'monitoring', 'performance'):
        tab = 'reports'
    stats = get_cached_security_stats()
    return render_template('security/reports_center.html', active_tab=tab, stats=stats)


@security_bp.route('/tools-center')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def tools_center():
    """
    🔧 Tools & Integration Center - 4 في 1
    - التكامل (Integrations)
    - إدارة البريد (Email Manager)
    - إدارة الكروت (Card Vault)
    - تصدير البيانات (Data Export)
    """
    allowed_tabs = {'integrations', 'email', 'cards', 'export'}
    tab = request.args.get('tab', 'integrations')
    if tab not in allowed_tabs:
        tab = 'integrations'
    
    integrations_data = {
        'stripe': {
            'enabled': _get_setting('stripe_enabled', False),
            'public_key': _get_setting('stripe_public_key', ''),
            'secret_key': _get_setting('stripe_secret_key', ''),
        },
        'paypal': {
            'enabled': _get_setting('paypal_enabled', False),
            'mode': _get_setting('paypal_mode', 'sandbox'),
            'client_id': _get_setting('paypal_client_id', ''),
        },
        'sms': {
            'enabled': _get_setting('sms_enabled', False),
            'twilio_phone_number': _get_setting('twilio_phone_number', ''),
        },
        'email': {
            'enabled': _get_setting('email_enabled', True),
            'smtp_host': _get_setting('smtp_host', ''),
        },
    }
    
    from datetime import datetime, timedelta, timezone
    from models import NotificationLog, OnlinePayment
    
    email_query = NotificationLog.query.filter_by(type='email')
    email_stats = {
        'total': email_query.count(),
        'sent_last_day': email_query.filter(
            NotificationLog.created_at >= datetime.now(timezone.utc) - timedelta(days=1)
        ).count(),
        'failed': email_query.filter_by(status='failed').count(),
    }
    email_logs = email_query.order_by(NotificationLog.created_at.desc()).limit(5).all()
    
    card_stats = {
        'total': OnlinePayment.query.count(),
        'successful': OnlinePayment.query.filter_by(status='SUCCESS').count(),
        'pending': OnlinePayment.query.filter_by(status='PENDING').count(),
        'failed': OnlinePayment.query.filter_by(status='FAILED').count(),
    }
    recent_cards = OnlinePayment.query.order_by(OnlinePayment.created_at.desc()).limit(5).all()
    
    export_links = [
        {
            'name': 'أعمار الذمم (عملاء)',
            'format': 'CSV',
            'icon': 'fas fa-user-clock',
            'description': 'كشف الرصيد المجمع لكل عميل حسب الفترات الزمنية.',
            'url': url_for('reports_bp.export_ar_aging_csv'),
        },
        {
            'name': 'أعمار الذمم (موردون)',
            'format': 'CSV',
            'icon': 'fas fa-truck-loading',
            'description': 'متابعة مستحقات الموردين والأعمار الزمنية.',
            'url': url_for('reports_bp.export_ap_aging_csv'),
        },
    ]
    dynamic_tables = [
        {'value': 'Sale', 'label': 'المبيعات'},
        {'value': 'Invoice', 'label': 'الفواتير'},
        {'value': 'Payment', 'label': 'المدفوعات'},
        {'value': 'Customer', 'label': 'العملاء'},
        {'value': 'Supplier', 'label': 'الموردون'},
    ]
    
    stats = get_cached_security_stats()
    return render_template(
        'security/tools_center.html',
        active_tab=tab,
        integrations=integrations_data,
        stats=stats,
        email_stats=email_stats,
        email_logs=email_logs,
        card_stats=card_stats,
        recent_cards=recent_cards,
        export_links=export_links,
        dynamic_tables=dynamic_tables,
    )


def _unused_ai_config_function():
    """إعدادات AI - Groq API Keys - تم دمجها في AI Hub"""
    """تكوين AI للمساعد الذكي - دعم مفاتيح متعددة"""
    if request.method == 'POST':
        action = request.form.get('action', 'add')
        
        if action == 'add':
            api_provider = request.form.get('api_provider', 'groq')
            api_key = request.form.get('api_key', '').strip()
            key_name = request.form.get('key_name', '').strip()
            is_active = request.form.get('is_active') == 'on'
            
            if api_key:
                # قراءة المفاتيح الحالية
                keys_json = _get_system_setting('AI_API_KEYS', '[]')
                try:
                    keys = json.loads(keys_json)
                except Exception:
                    keys = []
                
                # إضافة مفتاح جديد
                new_key = {
                    'id': len(keys) + 1,
                    'name': key_name or f'مفتاح {len(keys) + 1}',
                    'provider': api_provider,
                    'key': api_key,
                    'is_active': is_active,
                    'created_at': datetime.now(timezone.utc).isoformat()
                }
                keys.append(new_key)
                
                # حفظ
                _set_system_setting('AI_API_KEYS', json.dumps(keys, ensure_ascii=False))
                flash(f'✅ تم إضافة المفتاح: {new_key["name"]}', 'success')
            else:
                flash('⚠️ مفتاح API مطلوب', 'warning')
        
        elif action == 'delete':
            key_id = int(request.form.get('key_id', 0))
            keys_json = _get_system_setting('AI_API_KEYS', '[]')
            try:
                keys = json.loads(keys_json)
                keys = [k for k in keys if k.get('id') != key_id]
                _set_system_setting('AI_API_KEYS', json.dumps(keys, ensure_ascii=False))
                flash('✅ تم حذف المفتاح', 'success')
            except Exception:
                flash('⚠️ خطأ في حذف المفتاح', 'danger')
        
        elif action == 'set_active':
            key_id = int(request.form.get('key_id', 0))
            keys_json = _get_system_setting('AI_API_KEYS', '[]')
            try:
                keys = json.loads(keys_json)
                for k in keys:
                    k['is_active'] = (k.get('id') == key_id)
                _set_system_setting('AI_API_KEYS', json.dumps(keys, ensure_ascii=False))
                flash('✅ تم تفعيل المفتاح', 'success')
            except Exception:
                flash('⚠️ خطأ في تفعيل المفتاح', 'danger')
        
        return redirect(url_for('security.ai_config'))
    
    # قراءة المفاتيح
    keys_json = _get_system_setting('AI_API_KEYS', '[]')
    try:
        keys = json.loads(keys_json)
    except Exception:
        keys = []
    
    return render_template('security/ai_config.html', keys=keys)




@security_bp.route('/card-vault')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def card_vault():
    """خزنة الكروت - عرض بيانات الفيزا كارد المشفرة"""
    from models import OnlinePayment
    cards = OnlinePayment.query.order_by(OnlinePayment.created_at.desc()).limit(100).all()
    
    stats = {
        'total_cards': OnlinePayment.query.count(),
        'successful': OnlinePayment.query.filter_by(status='SUCCESS').count(),
        'pending': OnlinePayment.query.filter_by(status='PENDING').count(),
        'failed': OnlinePayment.query.filter_by(status='FAILED').count(),
    }
    
    return render_template('security/card_vault.html', cards=cards, stats=stats)


@security_bp.route('/code-editor', methods=['GET', 'POST'])
@security_bp.route('/theme-editor', methods=['GET', 'POST'])  # Alias
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def theme_editor():
    """
    🎨 محرر الملفات الموحد - Unified File Editor
    
    📋 الوصف:
        محرر شامل 3 في 1 (CSS + HTML Templates + System Settings)
    
    📥 Parameters:
        - type (str): css|html|text (default: css)
        - file (str): اسم الملف للـ CSS (optional)
        - template (str): مسار القالب للـ HTML (optional)
        - key (str): مفتاح الإعداد للـ text (optional)
    
    📤 Response:
        HTML: templates/security/theme_editor.html
        
    🎯 الوظائف:
        ✅ CSS: تحرير ملفات الأنماط (static/css/*.css)
        ✅ HTML: تحرير قوالب Jinja2 (templates/**/*.html)
        ✅ Text: تحرير System Settings (key-value pairs)
    
    💾 العمليات المدعومة:
        - عرض شجرة الملفات/القوالب
        - تحرير محتوى الملفات
        - حفظ التغييرات
        - معاينة مباشرة (للـ CSS)
        - Syntax highlighting
    
    💡 Usage Examples:
        /theme-editor?type=css&file=style.css
        /theme-editor?type=html&template=base.html
        /theme-editor?type=text&key=company_name
        /code-editor  ← نفس الوظيفة (alias)
    
    🔒 Security:
        - Owner only
        - Path traversal protection (..)
        - File extension validation
        - UTF-8 encoding enforced
    """
    import os
    from models import SystemSettings
    
    editor_type = request.args.get('type', 'css')  # css, html, text
    
    if request.method == 'POST':
        editor_type = request.form.get('editor_type', 'css')
        
        if editor_type == 'css':
            # حفظ CSS
            css_dir = os.path.join(current_app.root_path, 'static', 'css')
            filename = request.form.get('filename', 'style.css')
            content = request.form.get('content', '')
            
            if filename.endswith('.css') and not '..' in filename:
                filepath = os.path.join(css_dir, filename)
                try:
                    with open(filepath, 'w', encoding='utf-8') as f:
                        f.write(content)
                    flash(f'✅ تم حفظ {filename} بنجاح!', 'success')
                except Exception as e:
                    flash(f'❌ خطأ: {str(e)}', 'danger')
                    
        elif editor_type == 'html':
            # حفظ HTML Template
            templates_dir = os.path.join(current_app.root_path, 'templates')
            filepath = request.form.get('filepath', '')
            content = request.form.get('content', '')
            
            if filepath and not '..' in filepath:
                full_path = os.path.join(templates_dir, filepath)
                try:
                    os.makedirs(os.path.dirname(full_path), exist_ok=True)
                    with open(full_path, 'w', encoding='utf-8') as f:
                        f.write(content)
                    flash(f'✅ تم حفظ {filepath} بنجاح!', 'success')
                except Exception as e:
                    flash(f'❌ خطأ: {str(e)}', 'danger')
                    
        elif editor_type == 'text':
            # حفظ النصوص
            key = request.form.get('key')
            value = request.form.get('value')
            
            setting = SystemSettings.query.filter_by(key=key).first()
            if setting:
                setting.value = value
            else:
                setting = SystemSettings(key=key, value=value)
                db.session.add(setting)
            
            db.session.commit()
            flash(f'✅ تم تحديث {key}', 'success')
        
        return redirect(url_for('security.theme_editor', type=editor_type))
    
    # جمع البيانات حسب النوع
    data = {}
    
    # CSS Files
    css_dir = os.path.join(current_app.root_path, 'static', 'css')
    css_files = [f for f in os.listdir(css_dir) if f.endswith('.css')]
    selected_css = request.args.get('file', 'style.css')
    css_content = ''
    if selected_css in css_files:
        try:
            with open(os.path.join(css_dir, selected_css), 'r', encoding='utf-8') as f:
                css_content = f.read()
        except Exception:
            pass
    data['css'] = {'files': css_files, 'selected': selected_css, 'content': css_content}
    
    # HTML Templates
    templates_dir = os.path.join(current_app.root_path, 'templates')
    def get_templates_tree(directory, prefix=''):
        items = []
        try:
            for item in sorted(os.listdir(directory)):
                if item.startswith('.') or item == '__pycache__':
                    continue
                full_path = os.path.join(directory, item)
                rel_path = os.path.join(prefix, item) if prefix else item
                if os.path.isdir(full_path):
                    items.append({'type': 'dir', 'name': item, 'path': rel_path})
                    items.extend(get_templates_tree(full_path, rel_path))
                elif item.endswith('.html'):
                    items.append({'type': 'file', 'name': item, 'path': rel_path})
        except Exception:
            pass
        return items
    
    templates_tree = get_templates_tree(templates_dir)
    selected_template = request.args.get('template', 'base.html')
    template_content = ''
    if selected_template and not '..' in selected_template:
        try:
            with open(os.path.join(templates_dir, selected_template), 'r', encoding='utf-8') as f:
                template_content = f.read()
        except Exception:
            pass
    data['html'] = {'tree': templates_tree, 'selected': selected_template, 'content': template_content}
    
    # Text Settings
    text_settings = SystemSettings.query.filter(
        SystemSettings.key.like('%_text%') | 
        SystemSettings.key.like('%_label%') |
        SystemSettings.key.like('%_name%')
    ).all()
    data['text'] = {'settings': text_settings}
    
    return render_template('security/theme_editor.html', 
                         data=data,
                         active_tab=editor_type)




@security_bp.route('/logo-manager', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def logo_manager():
    """مدير الشعارات - رفع وتعديل الشعارات"""
    import os
    from werkzeug.utils import secure_filename
    
    if request.method == 'POST':
        if 'logo_file' in request.files:
            file = request.files['logo_file']
            logo_type = request.form.get('logo_type', 'main')
            
            if file and file.filename:
                filename = secure_filename(file.filename)
                upload_path = os.path.join(current_app.root_path, 'static', 'img')
                
                logo_mapping = {
                    'main': 'logo_main.png',
                    'emblem': 'logo_emblem.png',
                    'white': 'logo_white.png',
                    'favicon': 'favicon.png'
                }
                
                target_name = logo_mapping.get(logo_type, 'logo_main.png')
                filepath = os.path.join(upload_path, target_name)
                
                # تحديث SystemSettings للمسار الجديد
                key_map = {
                    'main': 'custom_logo',
                    'emblem': 'custom_logo_emblem',
                    'white': 'custom_logo_white',
                    'favicon': 'custom_favicon'
                }
                
                # حفظ الملف
                try:
                    file.save(filepath)
                    
                    # تحديث الإعداد
                    setting_key = key_map.get(logo_type)
                    if setting_key:
                        db_setting = SystemSettings.query.filter_by(key=setting_key).first()
                        rel_path = f"static/img/{target_name}"
                        if db_setting:
                            db_setting.value = rel_path
                        else:
                            db.session.add(SystemSettings(key=setting_key, value=rel_path))
                        db.session.commit()
                        
                        # تحديث الكاش
                        try:
                            from extensions import cache
                            cache.delete(f"system_setting_{setting_key}")
                        except:
                            pass

                    flash(f'✅ تم رفع {target_name} بنجاح!', 'success')
                except Exception as e:
                    flash(f'❌ خطأ: {str(e)}', 'danger')
    
    # استرجاع القيم الحالية من الإعدادات أو القيم الافتراضية
    logos = {
        'main': SystemSettings.get_setting('custom_logo', 'static/img/logo.png').split('/')[-1],
        'emblem': SystemSettings.get_setting('custom_logo_emblem', 'static/img/logo_emblem.png').split('/')[-1],
        'white': SystemSettings.get_setting('custom_logo_white', 'static/img/logo_white.png').split('/')[-1],
        'favicon': SystemSettings.get_setting('custom_favicon', 'static/img/favicon.png').split('/')[-1]
    }
    
    return render_template('security/logo_manager.html', logos=logos)


@security_bp.route('/advanced-analytics')
@permission_required(SystemPermissions.VIEW_REPORTS)
def advanced_analytics():
    """تحليلات متقدمة - ذكاء اصطناعي"""
    from models import Payment, Sale, Expense, Customer, Supplier
    from sqlalchemy import func, extract
    from datetime import datetime, timedelta
    
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0)
    
    analytics = {
        'revenue_trend': [],
        'expense_trend': [],
        'top_customers': [],
        'top_products': [],
        'payment_methods': {},
        'monthly_growth': 0,
    }
    
    from decimal import Decimal
    from models import convert_amount
    
    payments_month = db.session.query(Payment).filter(
        Payment.direction == 'IN',
        Payment.status == 'COMPLETED',
        Payment.payment_date >= start_of_month
    ).all()
    
    revenue_by_day_dict = {}
    for p in payments_month:
        dt = p.payment_date.date() if p.payment_date else None
        if not dt:
            continue
        if dt not in revenue_by_day_dict:
            revenue_by_day_dict[dt] = Decimal('0.00')
        amt = Decimal(str(p.total_amount or 0))
        if p.currency == "ILS":
            revenue_by_day_dict[dt] += amt
        else:
            try:
                revenue_by_day_dict[dt] += convert_amount(amt, p.currency, "ILS", p.payment_date)
            except Exception:
                pass
    
    analytics['revenue_trend'] = [{'date': str(dt), 'amount': float(rev)} for dt, rev in sorted(revenue_by_day_dict.items())]
    
    all_payments_in = db.session.query(Payment).filter(
        Payment.direction == 'IN',
        Payment.status == 'COMPLETED'
    ).all()
    
    cust_totals = {}
    for p in all_payments_in:
        if not p.customer_id:
            continue
        if p.customer_id not in cust_totals:
            cust_totals[p.customer_id] = Decimal('0.00')
        amt = Decimal(str(p.total_amount or 0))
        if p.currency == "ILS":
            cust_totals[p.customer_id] += amt
        else:
            try:
                cust_totals[p.customer_id] += convert_amount(amt, p.currency, "ILS", p.payment_date)
            except Exception:
                pass
    
    top_customers_data = []
    for cid, total in cust_totals.items():
        cust = db.session.get(Customer, cid)
        if cust:
            top_customers_data.append({'name': cust.name, 'total': float(total)})
    top_customers_data.sort(key=lambda x: x['total'], reverse=True)
    analytics['top_customers'] = top_customers_data[:10]
    
    return render_template('security/advanced_analytics.html', analytics=analytics)


@security_bp.route('/permissions-manager', methods=['GET', 'POST'])
@permission_required(SystemPermissions.MANAGE_PERMISSIONS)
def permissions_manager():
    """إدارة الصلاحيات - إنشاء وتخصيص"""
    from models import Permission, Role
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create_permission':
            code = (request.form.get('code') or "").strip()
            name = (request.form.get('name') or "").strip()

            if not code or not name:
                flash('⚠️ يجب إدخال كود واسم للصلاحية.', 'warning')
                return redirect(url_for('security.permissions_manager'))

            # تطبيع الكود: حروف صغيرة + شرطات سفلية
            normalized_code = code.strip().lower().replace(" ", "_")

            # منع التكرار
            existing = Permission.query.filter_by(code=normalized_code).first()
            if existing:
                flash('⚠️ توجد صلاحية بنفس الكود بالفعل، يرجى اختيار كود مختلف.', 'warning')
                return redirect(url_for('security.permissions_manager'))
            
            perm = Permission(code=normalized_code, name=name)
            db.session.add(perm)
            db.session.commit()
            flash(f'✅ تم إنشاء صلاحية: {name}', 'success')
        
        return redirect(url_for('security.permissions_manager'))
    
    permissions = Permission.query.all()
    roles = Role.query.all()
    
    return render_template('security/permissions_manager.html', 
                         permissions=permissions,
                         roles=roles)


@security_bp.route('/email-manager', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def email_manager():
    """إدارة البريد - SMTP + قوالب (Deprecated -> Integrations)"""
    flash('ℹ️ تم نقل إعدادات البريد إلى صفحة التكاملات الموحدة.', 'info')
    return redirect(url_for('security.integrations'))


@security_bp.route('/invoice-designer', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def invoice_designer():
    """محرر الفواتير - تخصيص تصميم الفواتير"""
    from models import SystemSettings
    
    # قائمة الإعدادات المدعومة
    supported_settings = [
        'invoice_header_color',
        'invoice_footer_text',
        'invoice_header_title',
        'invoice_terms_text',
        'invoice_paper_size',
        'invoice_show_logo',
        'invoice_show_tax',
        'invoice_show_qr',
        'invoice_show_tax_number',
        'invoice_show_contact'
    ]

    if request.method == 'POST':
        for key in supported_settings:
            value = request.form.get(key)
            
            # معالجة الـ Checkboxes
            if key.startswith('invoice_show_'):
                value = 'True' if value == 'on' else 'False'
            
            # حفظ القيمة
            if value is not None:
                setting = SystemSettings.query.filter_by(key=key).first()
                if setting:
                    setting.value = str(value)
                else:
                    db.session.add(SystemSettings(key=key, value=str(value)))
        
        db.session.commit()
        
        # مسح الكاش إذا وجد
        try:
            from extensions import cache
            for key in supported_settings:
                cache.delete(f"system_setting_{key}")
        except:
            pass
            
        flash('✅ تم حفظ تصميم الفواتير بنجاح', 'success')
        return redirect(url_for('security.invoice_designer'))
    
    settings = {}
    for key in supported_settings:
        s = SystemSettings.query.filter_by(key=key).first()
        # تعيين قيم افتراضية ذكية
        default_val = ''
        if key == 'invoice_header_title': default_val = 'فاتورة ضريبية'
        if key == 'invoice_paper_size': default_val = 'A4'
        if key == 'invoice_header_color': default_val = '#3c8dbc'
        if key.startswith('invoice_show_'): default_val = 'True'
        
        settings[key] = s.value if s else default_val
    
    return render_template('security/invoice_designer.html', settings=settings)


@security_bp.route('/integrations', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def integrations():
    """🔌 مركز التكامل الشامل - إعداد جميع التكاملات من مكان واحد"""
    
    if request.method == 'POST':
        try:
            action = request.form.get('action')
            
            if action == 'save_stripe':
                _save_setting('stripe_enabled', request.form.get('stripe_enabled') == 'on')
                _save_setting('stripe_public_key', request.form.get('stripe_public_key', ''))
                _save_setting('stripe_secret_key', request.form.get('stripe_secret_key', ''))
                _save_setting('stripe_webhook_secret', request.form.get('stripe_webhook_secret', ''))
                flash('✅ تم حفظ إعدادات Stripe', 'success')
            
            elif action == 'save_paypal':
                _save_setting('paypal_enabled', request.form.get('paypal_enabled') == 'on')
                _save_setting('paypal_mode', request.form.get('paypal_mode', 'sandbox'))
                _save_setting('paypal_client_id', request.form.get('paypal_client_id', ''))
                _save_setting('paypal_secret', request.form.get('paypal_secret', ''))
                flash('✅ تم حفظ إعدادات PayPal', 'success')
            
            elif action == 'save_sms':
                _save_setting('sms_enabled', request.form.get('sms_enabled') == 'on')
                _save_setting('twilio_account_sid', request.form.get('twilio_account_sid', ''))
                _save_setting('twilio_auth_token', request.form.get('twilio_auth_token', ''))
                _save_setting('twilio_phone_number', request.form.get('twilio_phone_number', ''))
                _save_setting('twilio_whatsapp_number', request.form.get('twilio_whatsapp_number', ''))
                flash('✅ تم حفظ إعدادات SMS/WhatsApp', 'success')

            elif action == 'save_email':
                _save_setting('MAIL_SERVER', request.form.get('MAIL_SERVER', ''))
                _save_setting('MAIL_PORT', request.form.get('MAIL_PORT', '587'))
                _save_setting('MAIL_USERNAME', request.form.get('MAIL_USERNAME', ''))
                _save_setting('MAIL_PASSWORD', request.form.get('MAIL_PASSWORD', ''))
                _save_setting('MAIL_USE_TLS', request.form.get('MAIL_USE_TLS') == 'on')
                _save_setting('MAIL_DEFAULT_SENDER', request.form.get('MAIL_DEFAULT_SENDER', ''))
                flash('✅ تم حفظ إعدادات البريد الإلكتروني', 'success')
            
            elif action == 'save_thermal_printer':
                _save_setting('thermal_printer_enabled', request.form.get('thermal_printer_enabled') == 'on')
                _save_setting('thermal_printer_type', request.form.get('thermal_printer_type', 'network'))
                _save_setting('thermal_printer_ip', request.form.get('thermal_printer_ip', ''))
                _save_setting('thermal_printer_port', request.form.get('thermal_printer_port', '9100'))
                _save_setting('thermal_printer_usb_vendor', request.form.get('thermal_printer_usb_vendor', ''))
                _save_setting('thermal_printer_usb_product', request.form.get('thermal_printer_usb_product', ''))
                _save_setting('thermal_printer_width', request.form.get('thermal_printer_width', '80'))
                flash('✅ تم حفظ إعدادات الطابعة الحرارية', 'success')
            
            elif action == 'save_barcode_scanner':
                _save_setting('barcode_scanner_enabled', request.form.get('barcode_scanner_enabled') == 'on')
                _save_setting('barcode_scanner_type', request.form.get('barcode_scanner_type', 'web'))
                _save_setting('barcode_scanner_device', request.form.get('barcode_scanner_device', ''))
                _save_setting('barcode_auto_focus', request.form.get('barcode_auto_focus') == 'on')
                _save_setting('barcode_beep_sound', request.form.get('barcode_beep_sound') == 'on')
                flash('✅ تم حفظ إعدادات ماسح الباركود', 'success')
            
            elif action == 'save_cloud_storage':
                _save_setting('cloud_storage_enabled', request.form.get('cloud_storage_enabled') == 'on')
                _save_setting('aws_access_key', request.form.get('aws_access_key', ''))
                _save_setting('aws_secret_key', request.form.get('aws_secret_key', ''))
                _save_setting('aws_region', request.form.get('aws_region', 'eu-west-1'))
                _save_setting('aws_bucket', request.form.get('aws_bucket', ''))
                flash('✅ تم حفظ إعدادات التخزين السحابي', 'success')
            
            elif action == 'save_webhooks':
                _save_setting('webhooks_enabled', request.form.get('webhooks_enabled') == 'on')
                _save_setting('webhook_secret', request.form.get('webhook_secret', ''))
                _save_setting('webhook_retry_count', request.form.get('webhook_retry_count', '3'))
                _save_setting('webhook_timeout', request.form.get('webhook_timeout', '10'))
                flash('✅ تم حفظ إعدادات Webhooks', 'success')
            
            elif action == 'save_local_gateways':
                _save_setting('moyasar_enabled', request.form.get('moyasar_enabled') == 'on')
                _save_setting('moyasar_api_key', request.form.get('moyasar_api_key', ''))
                _save_setting('tap_enabled', request.form.get('tap_enabled') == 'on')
                _save_setting('tap_api_key', request.form.get('tap_api_key', ''))
                _save_setting('paytabs_enabled', request.form.get('paytabs_enabled') == 'on')
                _save_setting('paytabs_profile_id', request.form.get('paytabs_profile_id', ''))
                _save_setting('paytabs_server_key', request.form.get('paytabs_server_key', ''))
                flash('✅ تم حفظ إعدادات بوابات الدفع المحلية', 'success')
            
            elif action == 'save_pos_terminal':
                _save_setting('pos_terminal_enabled', request.form.get('pos_terminal_enabled') == 'on')
                _save_setting('pos_terminal_type', request.form.get('pos_terminal_type', 'verifone'))
                _save_setting('pos_terminal_ip', request.form.get('pos_terminal_ip', ''))
                _save_setting('pos_terminal_port', request.form.get('pos_terminal_port', '5000'))
                _save_setting('pos_merchant_id', request.form.get('pos_merchant_id', ''))
                flash('✅ تم حفظ إعدادات جهاز POS', 'success')
            
            elif action == 'save_obd2_reader':
                _save_setting('obd2_reader_enabled', request.form.get('obd2_reader_enabled') == 'on')
                _save_setting('obd2_reader_type', request.form.get('obd2_reader_type', 'bluetooth'))
                _save_setting('obd2_port', request.form.get('obd2_port', 'COM3'))
                _save_setting('obd2_bluetooth_address', request.form.get('obd2_bluetooth_address', ''))
                _save_setting('obd2_auto_scan', request.form.get('obd2_auto_scan') == 'on')
                flash('✅ تم حفظ إعدادات كمبيوتر السيارة', 'success')
            
            elif action == 'save_digital_scale':
                _save_setting('digital_scale_enabled', request.form.get('digital_scale_enabled') == 'on')
                _save_setting('digital_scale_type', request.form.get('digital_scale_type', 'serial'))
                _save_setting('digital_scale_port', request.form.get('digital_scale_port', 'COM4'))
                _save_setting('digital_scale_baudrate', request.form.get('digital_scale_baudrate', '9600'))
                flash('✅ تم حفظ إعدادات الميزان الإلكتروني', 'success')
            
            elif action == 'save_label_printer':
                _save_setting('label_printer_enabled', request.form.get('label_printer_enabled') == 'on')
                _save_setting('label_printer_type', request.form.get('label_printer_type', 'zebra'))
                _save_setting('label_printer_connection', request.form.get('label_printer_connection', 'usb'))
                _save_setting('label_printer_ip', request.form.get('label_printer_ip', ''))
                _save_setting('label_printer_port', request.form.get('label_printer_port', '9100'))
                _save_setting('label_printer_width', request.form.get('label_printer_width', '4'))
                flash('✅ تم حفظ إعدادات طابعة اللصاقات', 'success')
            
            elif action == 'save_cash_drawer':
                _save_setting('cash_drawer_enabled', request.form.get('cash_drawer_enabled') == 'on')
                _save_setting('cash_drawer_connection', request.form.get('cash_drawer_connection', 'printer'))
                _save_setting('cash_drawer_port', request.form.get('cash_drawer_port', 'COM1'))
                _save_setting('cash_drawer_open_code', request.form.get('cash_drawer_open_code', '27,112,0,25,250'))
                flash('✅ تم حفظ إعدادات درج النقدية', 'success')
            
            elif action == 'save_customer_display':
                _save_setting('customer_display_enabled', request.form.get('customer_display_enabled') == 'on')
                _save_setting('customer_display_type', request.form.get('customer_display_type', 'lcd'))
                _save_setting('customer_display_port', request.form.get('customer_display_port', 'COM2'))
                _save_setting('customer_display_lines', request.form.get('customer_display_lines', '2'))
                _save_setting('customer_display_chars', request.form.get('customer_display_chars', '20'))
                flash('✅ تم حفظ إعدادات شاشة العميل', 'success')
            
            elif action == 'save_fingerprint_scanner':
                _save_setting('fingerprint_scanner_enabled', request.form.get('fingerprint_scanner_enabled') == 'on')
                _save_setting('fingerprint_scanner_type', request.form.get('fingerprint_scanner_type', 'usb'))
                _save_setting('fingerprint_vendor_id', request.form.get('fingerprint_vendor_id', ''))
                _save_setting('fingerprint_product_id', request.form.get('fingerprint_product_id', ''))
                flash('✅ تم حفظ إعدادات قارئ البصمة', 'success')
            
            elif action == 'save_cctv_system':
                _save_setting('cctv_enabled', request.form.get('cctv_enabled') == 'on')
                _save_setting('cctv_type', request.form.get('cctv_type', 'hikvision'))
                _save_setting('cctv_nvr_ip', request.form.get('cctv_nvr_ip', ''))
                _save_setting('cctv_username', request.form.get('cctv_username', 'admin'))
                _save_setting('cctv_password', request.form.get('cctv_password', ''))
                _save_setting('cctv_recording', request.form.get('cctv_recording') == 'on')
                flash('✅ تم حفظ إعدادات نظام المراقبة', 'success')
            
            elif action == 'test_stripe':
                result = _test_stripe()
                if result['success']:
                    flash(f'✅ Stripe متصل! Account: {result.get("account_id")}', 'success')
                else:
                    flash(f'❌ Stripe: {result.get("error")}', 'danger')
            
            elif action == 'test_paypal':
                result = _test_paypal()
                if result['success']:
                    flash('✅ PayPal متصل بنجاح!', 'success')
                else:
                    flash(f'❌ PayPal: {result.get("error")}', 'danger')
            
            elif action == 'test_sms':
                test_number = request.form.get('test_phone_number')
                if test_number:
                    result = _test_sms(test_number)
                    if result['success']:
                        flash(f'✅ تم إرسال SMS إلى {test_number}', 'success')
                    else:
                        flash(f'❌ SMS: {result.get("error")}', 'danger')
                else:
                    flash('❌ أدخل رقم هاتف', 'warning')
            
            elif action == 'test_thermal_printer':
                result = _test_thermal()
                if result['success']:
                    flash('✅ الطابعة تعمل!', 'success')
                else:
                    flash(f'❌ الطابعة: {result.get("error")}', 'danger')
            
            elif action == 'test_pos_terminal':
                result = _test_pos()
                if result['success']:
                    flash('✅ جهاز POS متصل!', 'success')
                else:
                    flash(f'❌ POS: {result.get("error")}', 'danger')
            
            elif action == 'test_obd2_reader':
                result = _test_obd2()
                if result['success']:
                    flash(f'✅ OBD-II متصل! {result.get("vehicle_info", "")}', 'success')
                else:
                    flash(f'❌ OBD-II: {result.get("error")}', 'danger')
            
            elif action == 'test_digital_scale':
                result = _test_scale()
                if result['success']:
                    flash(f'✅ الميزان يعمل! القراءة: {result.get("weight", "0")} kg', 'success')
                else:
                    flash(f'❌ الميزان: {result.get("error")}', 'danger')
            
            elif action == 'test_cctv':
                result = _test_cctv()
                if result['success']:
                    flash(f'✅ نظام المراقبة متصل! الكاميرات: {result.get("cameras", 0)}', 'success')
                else:
                    flash(f'❌ CCTV: {result.get("error")}', 'danger')
            
            db.session.commit()
            return redirect(url_for('security.integrations'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ خطأ: {str(e)}', 'danger')
    
    # GET - جلب الإعدادات
    integrations_data = {
        'stripe': {
            'enabled': _get_setting('stripe_enabled', False),
            'public_key': _get_setting('stripe_public_key', ''),
            'secret_key': _get_setting('stripe_secret_key', ''),
            'webhook_secret': _get_setting('stripe_webhook_secret', ''),
        },
        'paypal': {
            'enabled': _get_setting('paypal_enabled', False),
            'mode': _get_setting('paypal_mode', 'sandbox'),
            'client_id': _get_setting('paypal_client_id', ''),
            'secret': _get_setting('paypal_secret', ''),
        },
        'sms': {
            'enabled': _get_setting('sms_enabled', False),
            'twilio_account_sid': _get_setting('twilio_account_sid', ''),
            'twilio_auth_token': _get_setting('twilio_auth_token', ''),
            'twilio_phone_number': _get_setting('twilio_phone_number', ''),
            'twilio_whatsapp_number': _get_setting('twilio_whatsapp_number', ''),
        },
        'email': {
            'server': _get_setting('MAIL_SERVER', ''),
            'port': _get_setting('MAIL_PORT', '587'),
            'username': _get_setting('MAIL_USERNAME', ''),
            'password': _get_setting('MAIL_PASSWORD', ''),
            'use_tls': _get_setting('MAIL_USE_TLS', True),
            'default_sender': _get_setting('MAIL_DEFAULT_SENDER', ''),
        },
        'thermal_printer': {
            'enabled': _get_setting('thermal_printer_enabled', False),
            'type': _get_setting('thermal_printer_type', 'network'),
            'ip': _get_setting('thermal_printer_ip', ''),
            'port': _get_setting('thermal_printer_port', '9100'),
            'usb_vendor': _get_setting('thermal_printer_usb_vendor', ''),
            'usb_product': _get_setting('thermal_printer_usb_product', ''),
            'width': _get_setting('thermal_printer_width', '80'),
        },
        'barcode_scanner': {
            'enabled': _get_setting('barcode_scanner_enabled', True),
            'type': _get_setting('barcode_scanner_type', 'web'),
            'device': _get_setting('barcode_scanner_device', ''),
            'auto_focus': _get_setting('barcode_auto_focus', True),
            'beep_sound': _get_setting('barcode_beep_sound', True),
        },
        'cloud_storage': {
            'enabled': _get_setting('cloud_storage_enabled', False),
            'aws_access_key': _get_setting('aws_access_key', ''),
            'aws_secret_key': _get_setting('aws_secret_key', ''),
            'aws_region': _get_setting('aws_region', 'eu-west-1'),
            'aws_bucket': _get_setting('aws_bucket', ''),
        },
        'webhooks': {
            'enabled': _get_setting('webhooks_enabled', False),
            'secret': _get_setting('webhook_secret', ''),
            'retry_count': _get_setting('webhook_retry_count', '3'),
            'timeout': _get_setting('webhook_timeout', '10'),
        },
        'local_gateways': {
            'moyasar_enabled': _get_setting('moyasar_enabled', False),
            'moyasar_api_key': _get_setting('moyasar_api_key', ''),
            'tap_enabled': _get_setting('tap_enabled', False),
            'tap_api_key': _get_setting('tap_api_key', ''),
            'paytabs_enabled': _get_setting('paytabs_enabled', False),
            'paytabs_profile_id': _get_setting('paytabs_profile_id', ''),
            'paytabs_server_key': _get_setting('paytabs_server_key', ''),
        },
        'pos_terminal': {
            'enabled': _get_setting('pos_terminal_enabled', False),
            'type': _get_setting('pos_terminal_type', 'verifone'),
            'ip': _get_setting('pos_terminal_ip', ''),
            'port': _get_setting('pos_terminal_port', '5000'),
            'merchant_id': _get_setting('pos_merchant_id', ''),
        },
        'obd2_reader': {
            'enabled': _get_setting('obd2_reader_enabled', False),
            'type': _get_setting('obd2_reader_type', 'bluetooth'),
            'port': _get_setting('obd2_port', 'COM3'),
            'bluetooth_address': _get_setting('obd2_bluetooth_address', ''),
            'auto_scan': _get_setting('obd2_auto_scan', True),
        },
        'digital_scale': {
            'enabled': _get_setting('digital_scale_enabled', False),
            'type': _get_setting('digital_scale_type', 'serial'),
            'port': _get_setting('digital_scale_port', 'COM4'),
            'baudrate': _get_setting('digital_scale_baudrate', '9600'),
        },
        'label_printer': {
            'enabled': _get_setting('label_printer_enabled', False),
            'type': _get_setting('label_printer_type', 'zebra'),
            'connection': _get_setting('label_printer_connection', 'usb'),
            'ip': _get_setting('label_printer_ip', ''),
            'port': _get_setting('label_printer_port', '9100'),
            'width': _get_setting('label_printer_width', '4'),
        },
        'cash_drawer': {
            'enabled': _get_setting('cash_drawer_enabled', False),
            'connection': _get_setting('cash_drawer_connection', 'printer'),
            'port': _get_setting('cash_drawer_port', 'COM1'),
            'open_code': _get_setting('cash_drawer_open_code', '27,112,0,25,250'),
        },
        'customer_display': {
            'enabled': _get_setting('customer_display_enabled', False),
            'type': _get_setting('customer_display_type', 'lcd'),
            'port': _get_setting('customer_display_port', 'COM2'),
            'lines': _get_setting('customer_display_lines', '2'),
            'chars_per_line': _get_setting('customer_display_chars', '20'),
        },
        'fingerprint_scanner': {
            'enabled': _get_setting('fingerprint_scanner_enabled', False),
            'type': _get_setting('fingerprint_scanner_type', 'usb'),
            'vendor_id': _get_setting('fingerprint_vendor_id', ''),
            'product_id': _get_setting('fingerprint_product_id', ''),
        },
        'cctv_system': {
            'enabled': _get_setting('cctv_enabled', False),
            'type': _get_setting('cctv_type', 'hikvision'),
            'nvr_ip': _get_setting('cctv_nvr_ip', ''),
            'username': _get_setting('cctv_username', 'admin'),
            'password': _get_setting('cctv_password', ''),
            'recording': _get_setting('cctv_recording', True),
        },
    }
    
    return render_template('security/integrations.html', integrations=integrations_data)


@cache.memoize(timeout=300)  # 5 دقائق
def get_cached_security_stats():
    """
    📊 جلب إحصائيات الأمان مع Caching
    
    التحسين: يتم حفظ النتائج لمدة 5 دقائق لتسريع التحميل
    """
    from datetime import datetime, timedelta, timezone
    
    # إحصائيات المستخدمين
    total_users = User.query.count()
    active_users = User.query.filter_by(is_active=True).count()
    blocked_users = User.query.filter_by(is_active=False).count()
    system_accounts = User.query.filter_by(is_system_account=True).count()
    
    # المتصلين الآن (آخر 15 دقيقة)
    threshold = datetime.now(timezone.utc) - timedelta(minutes=15)
    # Optimized: Filter in DB instead of loading all users
    online_users = User.query.filter(User.last_seen >= threshold).count()
    
    # محاولات فشل الدخول (آخر 24 ساعة)
    day_ago = datetime.now(timezone.utc) - timedelta(hours=24)
    from models import AuthAudit, AuthEvent
    try:
        failed_logins_24h = AuthAudit.query.filter(
            AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
            AuthAudit.created_at >= day_ago
        ).count()
    except Exception:
        failed_logins_24h = 0
    
    # Blocked IPs & Countries
    blocked_ips = 0
    blocked_countries = 0
    try:
        from models import BlockedIP, BlockedCountry
        blocked_ips = BlockedIP.query.count()
        blocked_countries = BlockedCountry.query.count()
    except Exception:
        pass
    
    # أنشطة مشبوهة
    suspicious_activities = 0
    try:
        suspicious_activities = db.session.query(
            func.count(AuthAudit.ip)
        ).filter(
            AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
            AuthAudit.created_at >= day_ago
        ).group_by(AuthAudit.ip).having(
            func.count(AuthAudit.ip) >= 5
        ).count()
    except Exception:
        pass
    
    db_size = "N/A"
    try:
        uri = current_app.config.get("SQLALCHEMY_DATABASE_URI", "")
        if "postgresql" in uri:
            db_size = db.session.execute(
                text("SELECT pg_size_pretty(pg_database_size(current_database()))")
            ).scalar() or "N/A"
    except Exception:
        pass
    
    # صحة النظام
    system_health = "ممتاز"
    if failed_logins_24h > 50:
        system_health = "تحذير"
    elif failed_logins_24h > 100:
        system_health = "خطر"
    
    # 🔄 حساب الإحصائيات (Optimized: Avoid heavy introspection)
    from sqlalchemy import inspect
    
    try:
        inspector = inspect(db.engine)
        tables = inspector.get_table_names()
        total_tables = len([t for t in tables if t != 'alembic_version'])
        
        # Estimation instead of full scan to improve performance
        total_indexes = total_tables * 2  # Approx
        total_relations = total_tables * 1 # Approx
        
    except Exception:
        total_indexes = 0
        total_tables = 0
        total_relations = 0
    
    # حساب عدد Routes (APIs) بشكل حقيقي (كل النظام)
    total_apis = len([rule for rule in current_app.url_map.iter_rules() if rule.endpoint != "static"])
    
    # Try/Catch for SystemSettings.get_setting to prevent template crashes on DB errors
    try:
        system_version = str(SystemSettings.get_setting("system_version", "1.0.0") or "1.0.0").strip()
    except Exception:
        system_version = "1.0.0"
    
    return {
        'total_users': total_users,
        'active_users': active_users,
        'blocked_users': blocked_users,
        'system_accounts': system_accounts,
        'online_users': online_users,
        'blocked_ips': blocked_ips,
        'blocked_countries': blocked_countries,
        'failed_logins_24h': failed_logins_24h,
        'suspicious_activities': suspicious_activities,
        'db_size': db_size,
        'system_health': system_health,
        'active_sessions': online_users,
        # 🔄 إحصائيات ديناميكية
        'total_services': total_tables,
        'system_version': system_version,
        'total_modules': f'{total_tables}+',
        'total_apis': total_apis,
        'total_indexes': total_indexes,
        'total_relations': total_relations
    }


@security_bp.app_context_processor
def inject_security_stats():
    return {
        'stats': get_cached_security_stats()
    }


@cache.memoize(timeout=300)  # 5 دقائق
def get_recent_suspicious_activities():
    """
    📋 جلب آخر الأنشطة المشبوهة مع Caching
    """
    from datetime import datetime, timedelta, timezone
    from models import AuthAudit, AuthEvent
    
    day_ago = datetime.now(timezone.utc) - timedelta(hours=24)
    
    try:
        return AuthAudit.query.filter(
            AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
            AuthAudit.created_at >= day_ago
        ).order_by(AuthAudit.created_at.desc()).limit(10).all()
    except Exception:
        return []


def _get_setting(key, default=None):
    """جلب إعداد من SystemSettings"""
    setting = SystemSettings.query.filter_by(key=key).first()
    if setting:
        value = setting.value
        if default is False or default is True:
            return value.lower() in ('true', '1', 'on', 'yes') if value else default
        return value if value else default
    return default


def _save_setting(key, value):
    """حفظ إعداد في SystemSettings"""
    # Use the model method for consistency and cache clearing
    SystemSettings.set_setting(key, value, commit=False)


def _test_stripe():
    """اختبار Stripe"""
    try:
        import stripe
        stripe.api_key = _get_setting('stripe_secret_key', '')
        if not stripe.api_key:
            return {'success': False, 'error': 'API Key مفقود'}
        account = stripe.Account.retrieve()
        return {'success': True, 'account_id': account.id, 'email': account.email}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _test_paypal():
    """اختبار PayPal"""
    try:
        import paypalrestsdk
        paypalrestsdk.configure({
            'mode': _get_setting('paypal_mode', 'sandbox'),
            'client_id': _get_setting('paypal_client_id', ''),
            'client_secret': _get_setting('paypal_secret', '')
        })
        payment = paypalrestsdk.Payment.find("TEST")
        return {'success': True}
    except paypalrestsdk.ResourceNotFound:
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _test_sms(phone_number):
    """اختبار SMS"""
    try:
        from twilio.rest import Client
        sid = _get_setting('twilio_account_sid', '')
        token = _get_setting('twilio_auth_token', '')
        from_num = _get_setting('twilio_phone_number', '')
        if not all([sid, token, from_num]):
            return {'success': False, 'error': 'إعدادات Twilio غير مكتملة'}
        client = Client(sid, token)
        message = client.messages.create(body='اختبار من نظام الكراج ✅', from_=from_num, to=phone_number)
        return {'success': True, 'sid': message.sid}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _test_thermal():
    """اختبار الطابعة الحرارية"""
    try:
        ptype = _get_setting('thermal_printer_type', 'network')
        if ptype == 'network':
            from escpos.printer import Network
            ip = _get_setting('thermal_printer_ip', '')
            if not ip:
                return {'success': False, 'error': 'IP مفقود'}
            printer = Network(ip, int(_get_setting('thermal_printer_port', '9100')))
            printer.text("اختبار طابعة\nTest Print\n")
            printer.cut()
            return {'success': True}
        elif ptype == 'usb':
            from escpos.printer import Usb
            vendor = _get_setting('thermal_printer_usb_vendor', '')
            product = _get_setting('thermal_printer_usb_product', '')
            if not vendor or not product:
                return {'success': False, 'error': 'USB IDs مفقودة'}
            printer = Usb(int(vendor, 16), int(product, 16))
            printer.text("اختبار\nTest\n")
            printer.cut()
            return {'success': True}
        return {'success': False, 'error': 'نوع غير معروف'}
    except ImportError:
        return {'success': False, 'error': 'pip install python-escpos'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _test_pos():
    """اختبار جهاز POS"""
    try:
        import socket
        ip = _get_setting('pos_terminal_ip', '')
        port = int(_get_setting('pos_terminal_port', '5000'))
        if not ip:
            return {'success': False, 'error': 'IP مفقود'}
        
        # محاولة الاتصال بالجهاز
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(3)
        result = sock.connect_ex((ip, port))
        sock.close()
        
        if result == 0:
            return {'success': True}
        else:
            return {'success': False, 'error': 'لا يمكن الاتصال بالجهاز'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _test_obd2():
    """اختبار كمبيوتر السيارة OBD-II"""
    try:
        import obd
        connection_type = _get_setting('obd2_reader_type', 'bluetooth')
        
        if connection_type == 'bluetooth':
            address = _get_setting('obd2_bluetooth_address', '')
            if not address:
                return {'success': False, 'error': 'عنوان Bluetooth مفقود'}
            connection = obd.OBD(portstr=address, baudrate=38400)
        else:  # Serial
            port = _get_setting('obd2_port', 'COM3')
            connection = obd.OBD(portstr=port)
        
        if connection.is_connected():
            # قراءة معلومات السيارة
            cmd = obd.commands.VIN
            response = connection.query(cmd)
            vin = response.value if response.value else 'N/A'
            connection.close()
            return {'success': True, 'vehicle_info': f'VIN: {vin}'}
        else:
            return {'success': False, 'error': 'لا يمكن الاتصال بالسيارة'}
    except ImportError:
        return {'success': False, 'error': 'pip install obd'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _test_scale():
    """اختبار الميزان الإلكتروني"""
    try:
        import serial
        port = _get_setting('digital_scale_port', 'COM4')
        baudrate = int(_get_setting('digital_scale_baudrate', '9600'))
        
        ser = serial.Serial(port, baudrate, timeout=2)
        ser.write(b'R\r\n')  # طلب قراءة
        response = ser.readline().decode('utf-8').strip()
        ser.close()
        
        if response:
            # استخراج الوزن من الاستجابة
            weight = ''.join(filter(lambda x: x.isdigit() or x == '.', response))
            return {'success': True, 'weight': weight}
        return {'success': False, 'error': 'لا توجد استجابة'}
    except ImportError:
        return {'success': False, 'error': 'pip install pyserial'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _test_cctv():
    """اختبار نظام المراقبة"""
    try:
        import requests
        from requests.auth import HTTPDigestAuth
        
        cctv_type = _get_setting('cctv_type', 'hikvision')
        ip = _get_setting('cctv_nvr_ip', '')
        username = _get_setting('cctv_username', 'admin')
        password = _get_setting('cctv_password', '')
        
        if not ip:
            return {'success': False, 'error': 'IP مفقود'}
        
        # Hikvision API
        if cctv_type == 'hikvision':
            url = f'http://{ip}/ISAPI/System/deviceInfo'
            response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=5)
            if response.status_code == 200:
                # جلب عدد الكاميرات
                cameras_url = f'http://{ip}/ISAPI/System/Video/inputs'
                cameras_resp = requests.get(cameras_url, auth=HTTPDigestAuth(username, password), timeout=5)
                # تقدير عدد الكاميرات من الاستجابة
                camera_count = cameras_resp.text.count('<VideoInputChannel>') if cameras_resp.status_code == 200 else 0
                return {'success': True, 'cameras': camera_count}
        
        # Dahua API
        elif cctv_type == 'dahua':
            url = f'http://{ip}/cgi-bin/magicBox.cgi?action=getDeviceType'
            response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=5)
            if response.status_code == 200:
                return {'success': True, 'cameras': 'متصل'}
        
        return {'success': False, 'error': 'نوع غير مدعوم'}
    except ImportError:
        return {'success': False, 'error': 'requests مثبتة مسبقاً'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ==================== خدمات التكامل الفعلية ====================

def get_stripe_service():
    """الحصول على خدمة Stripe"""
    if not _get_setting('stripe_enabled', False):
        return None
    try:
        import stripe
        stripe.api_key = _get_setting('stripe_secret_key', '')
        return stripe
    except Exception:
        return None


def get_paypal_service():
    """الحصول على خدمة PayPal"""
    if not _get_setting('paypal_enabled', False):
        return None
    try:
        import paypalrestsdk
        paypalrestsdk.configure({
            'mode': _get_setting('paypal_mode', 'sandbox'),
            'client_id': _get_setting('paypal_client_id', ''),
            'client_secret': _get_setting('paypal_secret', '')
        })
        return paypalrestsdk
    except Exception:
        return None


def send_sms(to, message):
    """إرسال SMS عبر Twilio"""
    if not _get_setting('sms_enabled', False):
        return {'success': False, 'error': 'SMS غير مفعّل'}
    try:
        from twilio.rest import Client
        client = Client(
            _get_setting('twilio_account_sid', ''),
            _get_setting('twilio_auth_token', '')
        )
        msg = client.messages.create(
            body=message,
            from_=_get_setting('twilio_phone_number', ''),
            to=to
        )
        return {'success': True, 'sid': msg.sid}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def send_whatsapp(to, message):
    """إرسال WhatsApp عبر Twilio"""
    if not _get_setting('sms_enabled', False):
        return {'success': False, 'error': 'WhatsApp غير مفعّل'}
    try:
        from twilio.rest import Client
        client = Client(
            _get_setting('twilio_account_sid', ''),
            _get_setting('twilio_auth_token', '')
        )
        msg = client.messages.create(
            body=message,
            from_=_get_setting('twilio_whatsapp_number', ''),
            to=f'whatsapp:{to}'
        )
        return {'success': True, 'sid': msg.sid}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def print_thermal_invoice(sale):
    """طباعة فاتورة على طابعة حرارية"""
    if not _get_setting('thermal_printer_enabled', False):
        return {'success': False, 'error': 'الطابعة غير مفعّلة'}
    
    try:
        ptype = _get_setting('thermal_printer_type', 'network')
        
        # الاتصال بالطابعة
        if ptype == 'network':
            from escpos.printer import Network
            printer = Network(
                _get_setting('thermal_printer_ip', ''),
                int(_get_setting('thermal_printer_port', '9100'))
            )
        else:  # USB
            from escpos.printer import Usb
            printer = Usb(
                int(_get_setting('thermal_printer_usb_vendor', ''), 16),
                int(_get_setting('thermal_printer_usb_product', ''), 16)
            )
        
        # طباعة الفاتورة
        printer.set(align='center', text_type='B', width=2, height=2)
        printer.text(f"فاتورة رقم\n{sale.sale_number}\n")
        printer.set(align='center', text_type='normal', width=1, height=1)
        printer.text("─" * 32 + "\n")
        
        printer.set(align='right')
        printer.text(f"العميل: {sale.customer.name}\n")
        printer.text(f"التاريخ: {sale.sale_date.strftime('%Y-%m-%d %H:%M')}\n")
        printer.text("─" * 32 + "\n\n")
        
        # الأصناف
        for line in sale.lines:
            printer.text(f"{line.product.name[:24]}\n")
            printer.text(f"  {line.quantity} × {line.unit_price:.2f} = {line.net_amount:.2f}\n")
        
        printer.text("\n" + "─" * 32 + "\n")
        
        # الإجمالي
        printer.set(align='right', text_type='B')
        printer.text(f"المجموع: {sale.total_amount:.2f} {sale.currency}\n")
        
        if sale.total_paid > 0:
            printer.set(text_type='normal')
            printer.text(f"المدفوع: {sale.total_paid:.2f}\n")
            printer.text(f"المتبقي: {sale.balance_due:.2f}\n")
        
        printer.text("\n")
        printer.set(align='center', text_type='normal')
        printer.text("شكراً لثقتكم\nThank You\n")
        printer.text(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # قص الفاتورة
        printer.cut()
        
        return {'success': True}
        
    except ImportError:
        return {'success': False, 'error': 'pip install python-escpos'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def upload_to_s3(file, folder='uploads'):
    """رفع ملف إلى AWS S3"""
    if not _get_setting('cloud_storage_enabled', False):
        return {'success': False, 'error': 'Cloud Storage غير مفعّل'}
    
    try:
        import boto3
        from werkzeug.utils import secure_filename
        import uuid
        
        s3 = boto3.client('s3',
            aws_access_key_id=_get_setting('aws_access_key', ''),
            aws_secret_access_key=_get_setting('aws_secret_key', ''),
            region_name=_get_setting('aws_region', 'eu-west-1')
        )
        
        bucket = _get_setting('aws_bucket', '')
        filename = secure_filename(file.filename)
        key = f"{folder}/{uuid.uuid4()}_{filename}"
        
        s3.upload_fileobj(file, bucket, key, ExtraArgs={'ACL': 'public-read'})
        url = f"https://{bucket}.s3.amazonaws.com/{key}"
        
        return {'success': True, 'url': url}
    except ImportError:
        return {'success': False, 'error': 'pip install boto3'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def trigger_webhook(event_name, data):
    """إطلاق webhook للأنظمة الخارجية"""
    if not _get_setting('webhooks_enabled', False):
        return
    
    try:
        import hmac
        import hashlib
        import requests
        
        # جلب جميع الـ webhooks المسجلة من SystemSettings
        webhooks_json = _get_setting('registered_webhooks', '[]')
        webhooks = json.loads(webhooks_json) if webhooks_json else []
        
        for webhook in webhooks:
            if event_name in webhook.get('events', []):
                payload = {
                    'event': event_name,
                    'data': data,
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'source': 'garage_manager'
                }
                
                # إنشاء signature
                secret = _get_setting('webhook_secret', '')
                signature = hmac.new(
                    secret.encode(),
                    json.dumps(payload).encode(),
                    hashlib.sha256
                ).hexdigest()
                
                headers = {
                    'Content-Type': 'application/json',
                    'X-Webhook-Signature': signature,
                    'X-Event-Type': event_name
                }
                
                timeout = int(_get_setting('webhook_timeout', '10'))
                
                requests.post(
                    webhook['url'],
                    json=payload,
                    headers=headers,
                    timeout=timeout
                )
    except Exception as e:
        current_app.logger.error(f'Webhook error: {str(e)}')


def process_card_payment(amount, currency='ILS'):
    """معالجة دفعة عبر جهاز POS"""
    if not _get_setting('pos_terminal_enabled', False):
        return {'success': False, 'error': 'جهاز POS غير مفعّل'}
    
    try:
        import socket
        import json
        
        ip = _get_setting('pos_terminal_ip', '')
        port = int(_get_setting('pos_terminal_port', '5000'))
        merchant_id = _get_setting('pos_merchant_id', '')
        
        # إنشاء طلب دفع
        payment_request = {
            'action': 'sale',
            'amount': float(amount),
            'currency': currency,
            'merchant_id': merchant_id,
            'timestamp': datetime.now().isoformat()
        }
        
        # إرسال للجهاز
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(60)  # دقيقة للدفع
        sock.connect((ip, port))
        sock.send(json.dumps(payment_request).encode('utf-8'))
        
        # استقبال الاستجابة
        response = sock.recv(4096).decode('utf-8')
        sock.close()
        
        result = json.loads(response)
        return result
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


def scan_vehicle_obd2():
    """فحص السيارة وقراءة الأعطال من OBD-II"""
    if not _get_setting('obd2_reader_enabled', False):
        return {'success': False, 'error': 'OBD-II غير مفعّل'}
    
    try:
        import obd
        
        connection_type = _get_setting('obd2_reader_type', 'bluetooth')
        
        if connection_type == 'bluetooth':
            address = _get_setting('obd2_bluetooth_address', '')
            connection = obd.OBD(portstr=address, baudrate=38400)
        else:
            port = _get_setting('obd2_port', 'COM3')
            connection = obd.OBD(portstr=port)
        
        if not connection.is_connected():
            return {'success': False, 'error': 'لا يمكن الاتصال بالسيارة'}
        
        # قراءة البيانات
        data = {
            'vin': None,
            'dtc_codes': [],
            'rpm': None,
            'speed': None,
            'coolant_temp': None,
            'engine_load': None,
            'fuel_level': None,
        }
        
        # VIN
        cmd = obd.commands.VIN
        response = connection.query(cmd)
        data['vin'] = str(response.value) if response.value else None
        
        # أكواد الأعطال
        cmd = obd.commands.GET_DTC
        response = connection.query(cmd)
        if response.value:
            data['dtc_codes'] = [(code[0], code[1]) for code in response.value]
        
        # RPM
        cmd = obd.commands.RPM
        response = connection.query(cmd)
        data['rpm'] = float(response.value.magnitude) if response.value else None
        
        # السرعة
        cmd = obd.commands.SPEED
        response = connection.query(cmd)
        data['speed'] = float(response.value.magnitude) if response.value else None
        
        # حرارة المحرك
        cmd = obd.commands.COOLANT_TEMP
        response = connection.query(cmd)
        data['coolant_temp'] = float(response.value.magnitude) if response.value else None
        
        # حمل المحرك
        cmd = obd.commands.ENGINE_LOAD
        response = connection.query(cmd)
        data['engine_load'] = float(response.value.magnitude) if response.value else None
        
        # مستوى الوقود
        cmd = obd.commands.FUEL_LEVEL
        response = connection.query(cmd)
        data['fuel_level'] = float(response.value.magnitude) if response.value else None
        
        connection.close()
        
        return {'success': True, 'data': data}
        
    except ImportError:
        return {'success': False, 'error': 'pip install obd'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def read_weight_from_scale():
    """قراءة الوزن من الميزان الإلكتروني"""
    if not _get_setting('digital_scale_enabled', False):
        return {'success': False, 'error': 'الميزان غير مفعّل'}
    
    try:
        import serial
        port = _get_setting('digital_scale_port', 'COM4')
        baudrate = int(_get_setting('digital_scale_baudrate', '9600'))
        
        ser = serial.Serial(port, baudrate, timeout=2)
        ser.write(b'R\r\n')  # طلب قراءة
        response = ser.readline().decode('utf-8').strip()
        ser.close()
        
        if response:
            # استخراج الوزن من الاستجابة
            weight = ''.join(filter(lambda x: x.isdigit() or x == '.', response))
            return {'success': True, 'weight': float(weight), 'unit': 'kg'}
        return {'success': False, 'error': 'لا توجد استجابة'}
    except ImportError:
        return {'success': False, 'error': 'pip install pyserial'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def print_product_label(product):
    """طباعة لصاقة منتج على طابعة Zebra"""
    if not _get_setting('label_printer_enabled', False):
        return {'success': False, 'error': 'طابعة اللصاقات غير مفعّلة'}
    
    try:
        from zebra import Zebra
        
        printer_type = _get_setting('label_printer_type', 'zebra')
        connection = _get_setting('label_printer_connection', 'usb')
        
        if connection == 'network':
            ip = _get_setting('label_printer_ip', '')
            port = int(_get_setting('label_printer_port', '9100'))
            z = Zebra(f'{ip}:{port}')
        else:
            z = Zebra()  # USB - default
        
        # ZPL code for label
        width = _get_setting('label_printer_width', '4')  # 4 inch
        
        zpl = f"""
^XA
^FO50,50^A0N,50,50^FD{product.name[:20]}^FS
^FO50,120^A0N,30,30^FDSKU: {product.sku}^FS
^FO50,160^A0N,40,40^FD{product.sale_price:.2f} ILS^FS
^FO50,220^BY3^BCN,100,Y,N,N^FD{product.barcode}^FS
^XZ
"""
        
        z.output(zpl)
        return {'success': True}
        
    except ImportError:
        return {'success': False, 'error': 'pip install zebra'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def open_cash_drawer():
    """فتح درج النقدية"""
    if not _get_setting('cash_drawer_enabled', False):
        return {'success': False, 'error': 'درج النقدية غير مفعّل'}
    
    try:
        connection_type = _get_setting('cash_drawer_connection', 'printer')
        
        if connection_type == 'printer':
            # فتح عبر الطابعة الحرارية
            ptype = _get_setting('thermal_printer_type', 'network')
            if ptype == 'network':
                from escpos.printer import Network
                printer = Network(
                    _get_setting('thermal_printer_ip', ''),
                    int(_get_setting('thermal_printer_port', '9100'))
                )
            else:
                from escpos.printer import Usb
                printer = Usb(
                    int(_get_setting('thermal_printer_usb_vendor', ''), 16),
                    int(_get_setting('thermal_printer_usb_product', ''), 16)
                )
            
            printer._raw(b'\x1B\x70\x00\x19\xFA')  # ESC p 0 25 250
            return {'success': True}
            
        else:  # Serial connection
            import serial
            port = _get_setting('cash_drawer_port', 'COM1')
            open_code = _get_setting('cash_drawer_open_code', '27,112,0,25,250')
            
            codes = [int(c) for c in open_code.split(',')]
            ser = serial.Serial(port, 9600, timeout=1)
            ser.write(bytes(codes))
            ser.close()
            return {'success': True}
            
    except Exception as e:
        return {'success': False, 'error': str(e)}


def update_customer_display(line1, line2=''):
    """تحديث شاشة العميل"""
    if not _get_setting('customer_display_enabled', False):
        return {'success': False, 'error': 'شاشة العميل غير مفعّلة'}
    
    try:
        import serial
        port = _get_setting('customer_display_port', 'COM2')
        display_type = _get_setting('customer_display_type', 'lcd')
        chars = int(_get_setting('customer_display_chars', '20'))
        
        # تحديد السطور بطول الشاشة
        line1 = line1[:chars].ljust(chars)
        line2 = line2[:chars].ljust(chars)
        
        ser = serial.Serial(port, 9600, timeout=1)
        
        # مسح الشاشة
        ser.write(b'\x0C')
        
        # كتابة السطر الأول
        ser.write(line1.encode('utf-8'))
        
        # الانتقال للسطر الثاني
        if line2:
            ser.write(b'\x0A')
            ser.write(line2.encode('utf-8'))
        
        ser.close()
        return {'success': True}
        
    except ImportError:
        return {'success': False, 'error': 'pip install pyserial'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def verify_fingerprint(user_id):
    """التحقق من بصمة المستخدم"""
    if not _get_setting('fingerprint_scanner_enabled', False):
        return {'success': False, 'error': 'قارئ البصمة غير مفعّل'}
    
    try:
        # هذا مثال - يعتمد على نوع القارئ المستخدم
        vendor_id = _get_setting('fingerprint_vendor_id', '')
        product_id = _get_setting('fingerprint_product_id', '')
        
        if not vendor_id or not product_id:
            return {'success': False, 'error': 'إعدادات القارئ غير مكتملة'}
        
        return {'success': False, 'error': 'قيد التطوير - يحتاج SDK خاص بالقارئ'}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


def capture_cctv_snapshot(camera_id=1):
    """التقاط صورة من كاميرا محددة"""
    if not _get_setting('cctv_enabled', False):
        return {'success': False, 'error': 'نظام المراقبة غير مفعّل'}
    
    try:
        import requests
        from requests.auth import HTTPDigestAuth
        
        cctv_type = _get_setting('cctv_type', 'hikvision')
        ip = _get_setting('cctv_nvr_ip', '')
        username = _get_setting('cctv_username', 'admin')
        password = _get_setting('cctv_password', '')
        
        if cctv_type == 'hikvision':
            url = f'http://{ip}/ISAPI/Streaming/channels/{camera_id}01/picture'
            response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=10)
            
            if response.status_code == 200:
                # حفظ الصورة
                filename = f'cctv_snapshot_{camera_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.jpg'
                filepath = f'static/uploads/cctv/{filename}'
                
                import os
                os.makedirs('static/uploads/cctv', exist_ok=True)
                
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                
                return {'success': True, 'filepath': filepath, 'url': f'/static/uploads/cctv/{filename}'}
        
        return {'success': False, 'error': 'نوع غير مدعوم'}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


@security_bp.route('/save-integration', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def save_integration():
    """حفظ إعدادات التكامل - route قديم للتوافق"""
    # إعادة توجيه للـ route الجديد
    return integrations()


@security_bp.route('/test-integration/<integration_type>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def test_integration(integration_type):
    """اختبار تكامل معين"""
    result = _test_integration_connection(integration_type)
    
    # تسجيل النشاط
    _log_integration_activity(integration_type, 'tested', result['success'])
    
    return jsonify(result)


@security_bp.route('/send-test-message/<integration_type>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def send_test_message(integration_type):
    """إرسال رسالة تجريبية"""
    result = _send_test_message(integration_type)
    
    # تسجيل النشاط
    _log_integration_activity(integration_type, 'message_sent', result['success'])
    
    return jsonify(result)


@security_bp.route('/integration-stats')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def integration_stats():
    """إحصائيات التكاملات"""
    stats = _get_integration_stats()
    return jsonify(stats)


@security_bp.route('/live-monitoring')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def live_monitoring():
    """مراقبة فورية للنظام"""
    live_data = {
        'online_users': _get_online_users_detailed(),
        'recent_actions': _get_recent_actions(50),
        'system_metrics': _get_live_metrics(),
    }
    return render_template('security/live_monitoring.html', live_data=live_data)


@security_bp.route('/user-control')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def user_control():
    """
    👑 التحكم الكامل بالمستخدمين - Owner's User Management Panel
    
    📋 الوصف:
        لوحة تحكم شاملة ومتقدمة لإدارة جميع المستخدمين
        
    📤 Response:
        HTML: templates/security/user_control.html
        
    🎯 الوظائف:
        ✅ عرض جميع المستخدمين (مع النظام المخفي)
        ✅ إحصائيات متقدمة لكل مستخدم
        ✅ Impersonation (انتحال الشخصية)
        ✅ إعادة تعيين كلمات المرور
        ✅ تفعيل/تعطيل الحسابات
        ✅ حذف المستخدمين (حذف آمن)
        ✅ عمليات جماعية (Bulk operations)
        ✅ تاريخ النشاطات لكل مستخدم
        ✅ تحليلات الأداء
        ✅ فلاتر متقدمة
        ✅ Export to CSV/Excel
    
    📊 الإحصائيات المعروضة:
        - إجمالي المستخدمين
        - نشطين / معطلين
        - متصلين الآن
        - آخر نشاط
        - عدد المبيعات / العمليات لكل مستخدم
        - معدل النجاح
        - الأخطاء
    
    🔒 Security:
        - Owner only (المالك فقط)
        - Full audit trail لكل عملية
        - حماية من الحذف الذاتي
        - تأكيد على العمليات الخطيرة
    
    💡 Usage:
        /user-control
        /user-control?filter=active|inactive|online
        /user-control?role=<role_name>
        /user-control?search=<username>
    """
    from sqlalchemy import func, or_
    from models import Sale, Payment, ServiceRequest, AuditLog
    
    # Filters
    status_filter = request.args.get('filter', 'all')  # all, active, inactive, online, system
    role_filter = request.args.get('role')
    search_query = request.args.get('search', '').strip()
    
    # Base query مع eager loading
    query = User.query.options(
        db.joinedload(User.role)
    )
    
    # تطبيق الفلاتر
    if status_filter == 'active':
        query = query.filter(User.is_active == True)
    elif status_filter == 'inactive':
        query = query.filter(User.is_active == False)
    elif status_filter == 'system':
        query = query.filter(User.is_system_account == True)
    elif status_filter == 'online':
        # متصل خلال آخر 15 دقيقة (سيتم الفلترة يدوياً بعد fetch)
        pass  # التعامل مع timezone issues
    
    if role_filter:
        query = query.join(User.role).filter(User.role.has(name=role_filter))
    
    if search_query:
        query = query.filter(
            or_(
                User.username.ilike(f'%{search_query}%'),
                User.email.ilike(f'%{search_query}%')
            )
        )
    
    users = query.order_by(User.is_system_account.desc(), User.id.asc()).all()
    
    # فلترة المتصلين الآن إذا لزم الأمر
    if status_filter == 'online':
        threshold = datetime.now(timezone.utc) - timedelta(minutes=15)
        users = [u for u in users if u.last_seen and make_aware(u.last_seen) >= threshold]
    
    from decimal import Decimal
    from models import convert_amount
    from sqlalchemy.orm import selectinload
    
    user_ids = [u.id for u in users]
    
    if user_ids:
        sales_counts = db.session.query(
            Sale.seller_id,
            func.count(Sale.id).label('count'),
            func.sum(Sale.total_amount).label('total')
        ).filter(
            Sale.seller_id.in_(user_ids),
            Sale.status == 'CONFIRMED'
        ).group_by(Sale.seller_id).all()
        
        sales_dict = {uid: {'count': 0, 'total': Decimal('0.00')} for uid in user_ids}
        for seller_id, count, total in sales_counts:
            sales_dict[seller_id] = {'count': count, 'total': Decimal(str(total or 0))}
        
        all_user_sales = db.session.query(Sale).filter(
            Sale.seller_id.in_(user_ids),
            Sale.status == 'CONFIRMED'
        ).all()
        
        for sale in all_user_sales:
            amt = Decimal(str(sale.total_amount or 0))
            if sale.currency != "ILS":
                try:
                    amt = convert_amount(amt, sale.currency, "ILS", sale.sale_date)
                except Exception:
                    amt = Decimal('0.00')
            sales_dict[sale.seller_id]['total'] += amt
        
        services_counts = db.session.query(
            ServiceRequest.mechanic_id,
            func.count(ServiceRequest.id).label('count')
        ).filter(
            ServiceRequest.mechanic_id.in_(user_ids)
        ).group_by(ServiceRequest.mechanic_id).all()
        
        services_dict = {uid: 0 for uid in user_ids}
        for mechanic_id, count in services_counts:
            services_dict[mechanic_id] = count
        
        payments_counts = db.session.query(
            Payment.created_by,
            func.count(Payment.id).label('count')
        ).filter(
            Payment.created_by.in_(user_ids)
        ).group_by(Payment.created_by).all()
        
        payments_dict = {uid: 0 for uid in user_ids}
        for created_by, count in payments_counts:
            payments_dict[created_by] = count
        
        from sqlalchemy import distinct
        
        last_audits_subq = db.session.query(
            AuditLog.user_id,
            func.max(AuditLog.created_at).label('max_created_at')
        ).filter(
            AuditLog.user_id.in_(user_ids)
        ).group_by(AuditLog.user_id).subquery()
        
        last_audits = db.session.query(
            AuditLog.user_id,
            AuditLog.action,
            AuditLog.created_at
        ).join(
            last_audits_subq,
            and_(
                AuditLog.user_id == last_audits_subq.c.user_id,
                AuditLog.created_at == last_audits_subq.c.max_created_at
            )
        ).all()
        
        audits_dict = {uid: None for uid in user_ids}
        for user_id, action, created_at in last_audits:
            audits_dict[user_id] = {'action': action, 'created_at': created_at}
    else:
        sales_dict = {}
        services_dict = {}
        payments_dict = {}
        audits_dict = {}
    
    threshold = datetime.now(timezone.utc) - timedelta(minutes=15)
    
    for user in users:
        user_data = sales_dict.get(user.id, {'count': 0, 'total': Decimal('0.00')})
        user.sales_count = user_data['count']
        user.sales_total = float(user_data['total'])
        user.services_count = services_dict.get(user.id, 0)
        user.payments_count = payments_dict.get(user.id, 0)
        
        audit = audits_dict.get(user.id)
        user.last_activity_desc = audit['action'] if audit else 'لا يوجد'
        user.last_activity_time = audit['created_at'] if audit else None
        
        if user.last_seen:
            last_seen = user.last_seen
            if last_seen.tzinfo is None:
                last_seen = last_seen.replace(tzinfo=timezone.utc)
            user.is_online = last_seen >= threshold
        else:
            user.is_online = False
    
    # إحصائيات عامة
    stats = {
        'total_users': len(users),
        'active_users': len([u for u in users if u.is_active]),
        'inactive_users': len([u for u in users if not u.is_active]),
        'online_users': len([u for u in users if u.is_online]),
        'system_accounts': len([u for u in users if u.is_system_account]),
        'total_sales': sum(u.sales_total for u in users),
        'total_operations': sum(u.sales_count + u.services_count + u.payments_count for u in users)
    }
    
    # جميع الأدوار
    from models import Role
    all_roles = Role.query.order_by(Role.name).all()
    
    return render_template('security/user_control.html', 
                         users=users,
                         stats=stats,
                         all_roles=all_roles,
                         current_filter=status_filter,
                         current_role=role_filter,
                         search_query=search_query)


@security_bp.route('/impersonate/<int:user_id>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def impersonate_user(user_id):
    """تسجيل الدخول كمستخدم آخر"""
    from flask_login import logout_user, login_user
    
    target_user = db.get_or_404(User, user_id)
    
    # منع التسجيل كنفس المستخدم
    if target_user.id == current_user.id:
        flash('⚠️ أنت بالفعل هذا المستخدم!', 'warning')
        return redirect(url_for('security.user_control'))
    
    # حفظ المستخدم الأصلي
    session['original_user_id'] = current_user.id
    session['original_username'] = current_user.username
    session['impersonating'] = True
    
    # تسجيل في AuditLog
    try:
        log = AuditLog(
            user_id=current_user.id,
            action='security.impersonate_user',
            table_name='user',
            record_id=target_user.id,
            note=f'Owner impersonated as: {target_user.username}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        pass
    
    logout_user()
    login_user(target_user)
    
    flash(f'🕵️ تم تسجيل الدخول كـ {target_user.username}', 'warning')
    return redirect(url_for('main.dashboard'))


@security_bp.route('/stop-impersonate', methods=['POST'])
def stop_impersonate():
    """إيقاف التسجيل كمستخدم آخر"""
    from flask_login import logout_user, login_user
    
    if session.get('impersonating'):
        original_user_id = session.get('original_user_id')
        if original_user_id:
            original_user = db.session.get(User, original_user_id)
            if original_user:
                logout_user()
                login_user(original_user)
                session.pop('impersonating', None)
                session.pop('original_user_id', None)
                flash('تم العودة لحسابك الأصلي', 'success')
    
    return redirect(url_for('security.ultimate_control'))


@security_bp.route('/force-reset-password/<int:user_id>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def force_reset_password(user_id):
    """إعادة تعيين كلمة مرور المستخدم"""
    from werkzeug.security import generate_password_hash
    
    user = db.get_or_404(User, user_id)
    new_password = request.form.get('new_password', '123456')
    
    user.password = generate_password_hash(new_password)
    db.session.commit()
    
    flash(f'تم إعادة تعيين كلمة مرور {user.username}', 'success')
    return redirect(url_for('security.user_control'))


@security_bp.route('/toggle-user/<int:user_id>', methods=['POST'])
@security_bp.route('/toggle_user_status/<int:user_id>', methods=['POST'])  # Alias
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def toggle_user_status(user_id):
    """
    🔄 تفعيل/تعطيل مستخدم - Toggle User Status
    
    📋 الوصف:
        تبديل حالة المستخدم (نشط ↔ معطل)
    
    📥 Parameters:
        - user_id (int): معرّف المستخدم
    
    🔒 Security:
        - Owner only
        - Audit logging
        - حماية من تعطيل الذات
    """
    user = db.get_or_404(User, user_id)
    
    # حماية من تعطيل المالك لنفسه
    if user.id == current_user.id:
        flash('⚠️ لا يمكنك تعطيل حسابك الخاص!', 'warning')
        return redirect(url_for('security.user_control'))
    
    old_status = user.is_active
    user.is_active = not user.is_active
    
    # Audit log
    try:
        log = AuditLog(
            user_id=current_user.id,
            action='UPDATE',
            model_name='User',
            record_id=user.id,
            old_data=json.dumps({'is_active': old_status}, ensure_ascii=False),
            new_data=json.dumps({
                'is_active': user.is_active,
                'note': f'Owner toggled user {user.username} status'
            }, ensure_ascii=False),
            ip_address=request.remote_addr
        )
        db.session.add(log)
    except Exception:
        pass
    
    db.session.commit()
    
    status = 'مفعل' if user.is_active else 'معطل'
    flash(f'✅ المستخدم {user.username} الآن {status}', 'success')
    return redirect(url_for('security.user_control'))


@security_bp.route('/delete-user/<int:user_id>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def delete_user(user_id):
    """
    🗑️ حذف مستخدم نهائياً - Permanent User Deletion
    
    📋 الوصف:
        حذف نهائي للمستخدم من قاعدة البيانات
    
    📥 Parameters:
        - user_id (int): معرّف المستخدم
    
    ⚠️ Warning:
        - عملية خطيرة ولا يمكن التراجع عنها
        - يجب استخدام Hard Delete للحفاظ على السجلات
    
    🔒 Security:
        - Owner only
        - حماية من الحذف الذاتي
        - حماية حسابات النظام
        - Full audit trail
    """
    user = db.get_or_404(User, user_id)
    
    # حماية من الحذف الذاتي
    if user.id == current_user.id:
        flash('⚠️ لا يمكنك حذف حسابك الخاص!', 'danger')
        return redirect(url_for('security.user_control'))
    
    # حماية حسابات النظام
    if user.is_system_account:
        flash('🔒 حسابات النظام محمية من الحذف!', 'danger')
        return redirect(url_for('security.user_control'))
    
    username = user.username
    
    # Audit log قبل الحذف
    try:
        log = AuditLog(
            user_id=current_user.id,
            action='DELETE',
            model_name='User',
            record_id=user.id,
            old_data=json.dumps({
                'username': user.username,
                'email': user.email,
                'role': user.role.name if user.role else None,
                'note': f'Owner deleted user: {username}'
            }, ensure_ascii=False),
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        pass
    
    # الحذف
    try:
        db.session.delete(user)
        db.session.commit()
        flash(f'✅ تم حذف المستخدم {username} نهائياً', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ خطأ في الحذف: {str(e)}', 'danger')
    
    return redirect(url_for('security.user_control'))


@security_bp.route('/api/users/<int:user_id>/details')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_user_details(user_id):
    """
    📊 API - الحصول على تفاصيل مستخدم كاملة
    
    📥 Parameters:
        - user_id (int): معرّف المستخدم
    
    📤 Response:
        JSON: {
            success: true/false,
            user: {معلومات المستخدم الكاملة},
            error: رسالة الخطأ (إن وُجد)
        }
    """
    from decimal import Decimal
    from models import Sale, Payment, ServiceRequest, AuditLog, convert_amount
    
    try:
        user = db.session.get(User, user_id, options=[db.joinedload(User.role)])
        if not user:
            abort(404)
        
        # إحصائيات
        sales_count = Sale.query.filter_by(seller_id=user.id).count()
        
        # حساب إجمالي المبيعات مع تحويل العملات
        user_sales = Sale.query.filter_by(seller_id=user.id).all()
        sales_total = Decimal('0.00')
        for s in user_sales:
            amt = Decimal(str(s.total_amount or 0))
            if s.currency == "ILS":
                sales_total += amt
            else:
                try:
                    sales_total += convert_amount(amt, s.currency, "ILS", s.sale_date)
                except Exception:
                    pass
        
        services_count = ServiceRequest.query.filter_by(mechanic_id=user.id).count()
        payments_count = Payment.query.filter_by(created_by=user.id).count()
        
        # آخر الأنشطة
        recent_activities = AuditLog.query.filter_by(user_id=user.id).order_by(
            AuditLog.created_at.desc()
        ).limit(10).all()
        
        from permissions_config.permissions import PermissionsRegistry
        from models import Permission
        
        all_permissions = []
        for code, info in PermissionsRegistry.get_all_permissions().items():
            all_permissions.append({
                'code': code,
                'name': info.get('name_ar', code),
                'name_ar': info.get('name_ar', code),
                'module': info.get('module', 'other'),
                'description': info.get('description', '')
            })
        
        role_permissions = []
        if user.role:
            role_perms = PermissionsRegistry.get_role_permissions(user.role.name)
            role_permissions = list(role_perms)
        
        extra_permissions = []
        if hasattr(user, 'extra_permissions') and user.extra_permissions:
            try:
                extra_perms = user.extra_permissions.all() if hasattr(user.extra_permissions, 'all') else user.extra_permissions
                extra_permissions = [p.code for p in extra_perms if hasattr(p, 'code') and p.code]
            except Exception:
                pass
        
        user_data = {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'role': user.role.name if user.role else None,
            'is_active': user.is_active,
            'is_system_account': user.is_system_account,
            'last_login': user.last_login.isoformat() if user.last_login else None,
            'last_seen': user.last_seen.isoformat() if user.last_seen else None,
            'login_count': user.login_count,
            'last_login_ip': user.last_login_ip,
            'sales_count': sales_count,
            'sales_total': float(sales_total),
            'services_count': services_count,
            'payments_count': payments_count,
            'all_permissions': all_permissions,
            'role_permissions': role_permissions,
            'extra_permissions': extra_permissions,
            'recent_activities': [
                {
                    'action': a.action,
                    'created_at': a.created_at.isoformat() if a.created_at else None,
                    'note': a.note
                }
                for a in recent_activities
            ]
        }
        
        return jsonify({'success': True, 'user': user_data})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@security_bp.route('/api/users/bulk-operation', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_users_bulk_operation():
    """
    ⚡ API - عمليات جماعية على المستخدمين
    
    📥 Parameters (JSON):
        - operation (str): activate|deactivate|delete
        - user_ids (list): قائمة معرّفات المستخدمين
    
    📤 Response:
        JSON: {
            success: true/false,
            message: رسالة النجاح,
            affected: عدد المستخدمين المتأثرين,
            error: رسالة الخطأ (إن وُجد)
        }
    """
    try:
        data = request.get_json()
        operation = data.get('operation')
        user_ids = data.get('user_ids', [])
        
        if not user_ids:
            return jsonify({'success': False, 'error': 'لم يتم تحديد أي مستخدمين'}), 400
        
        # حماية من العمليات على المالك
        user_ids = [int(uid) for uid in user_ids if int(uid) != current_user.id]
        
        if operation == 'activate':
            User.query.filter(User.id.in_(user_ids)).update(
                {'is_active': True}, synchronize_session=False
            )
            message = f'تم تفعيل {len(user_ids)} مستخدم'
            
        elif operation == 'deactivate':
            User.query.filter(User.id.in_(user_ids)).update(
                {'is_active': False}, synchronize_session=False
            )
            message = f'تم تعطيل {len(user_ids)} مستخدم'
            
        elif operation == 'delete':
            # حماية حسابات النظام
            User.query.filter(
                User.id.in_(user_ids),
                User.is_system_account == False
            ).delete(synchronize_session=False)
            message = f'تم حذف {len(user_ids)} مستخدم'
            
        else:
            return jsonify({'success': False, 'error': 'عملية غير معروفة'}), 400
        
        # Audit log
        try:
            log = AuditLog(
                user_id=current_user.id,
                action=f'security.bulk_{operation}',
                table_name='users',
                note=f'Owner performed bulk {operation} on {len(user_ids)} users',
                ip_address=request.remote_addr
            )
            db.session.add(log)
        except Exception:
            pass
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': message,
            'affected': len(user_ids)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@security_bp.route('/api/users/<int:user_id>/activity-history')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_user_activity_history(user_id):
    """
    📜 API - سجل نشاطات المستخدم الكامل
    
    📥 Parameters:
        - user_id (int): معرّف المستخدم
        - limit (int): عدد السجلات (default: 50)
    
    📤 Response:
        JSON: { success, activities: [...], total }
    """
    try:
        limit = request.args.get('limit', 50, type=int)
        
        activities = AuditLog.query.filter_by(user_id=user_id).order_by(
            AuditLog.created_at.desc()
        ).limit(limit).all()
        
        activities_data = [
            {
                'id': a.id,
                'action': a.action,
                'table_name': a.table_name,
                'record_id': a.record_id,
                'note': a.note,
                'ip_address': a.ip_address,
                'created_at': a.created_at.isoformat() if a.created_at else None
            }
            for a in activities
        ]
        
        total_count = AuditLog.query.filter_by(user_id=user_id).count()
        
        return jsonify({
            'success': True,
            'activities': activities_data,
            'total': total_count,
            'showing': len(activities_data)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@security_bp.route('/create-user', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_USERS)
def create_user():
    """
    ➕ إنشاء مستخدم جديد - Create New User
    
    📋 الوصف:
        إنشاء مستخدم جديد مباشرة من لوحة التحكم
    
    📥 Parameters (POST):
        - username (str): اسم المستخدم (unique)
        - email (str): البريد الإلكتروني (unique)
        - password (str): كلمة المرور
        - role_id (int): الدور
        - is_active (bool): الحالة
    
    🔒 Security:
        - Owner only
        - Validation على جميع الحقول
        - Audit logging
    """
    try:
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '123456')
        role_id = request.form.get('role_id', type=int)
        is_active = request.form.get('is_active') == '1'
        
        if not username or not email:
            flash('❌ اسم المستخدم والبريد مطلوبان', 'danger')
            return redirect(url_for('security.user_control'))
        
        # التحقق من عدم التكرار
        existing = User.query.filter(
            (User.username == username) | (User.email == email)
        ).first()
        
        if existing:
            flash('❌ اسم المستخدم أو البريد موجود بالفعل', 'danger')
            return redirect(url_for('security.user_control'))
        
        # إنشاء المستخدم
        from werkzeug.security import generate_password_hash
        
        new_user = User(
            username=username,
            email=email,
            password_hash=generate_password_hash(password),
            role_id=role_id,
            is_active=is_active
        )
        
        db.session.add(new_user)
        db.session.flush()
        
        # Audit log
        log = AuditLog(
            user_id=current_user.id,
            action='CREATE',
            model_name='User',
            record_id=new_user.id,
            new_data=json.dumps({
                'username': username,
                'email': email,
                'role_id': role_id,
                'note': f'Owner created new user: {username}'
            }, ensure_ascii=False),
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'✅ تم إنشاء المستخدم {username} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('security.user_control'))


@security_bp.route('/update-user-role/<int:user_id>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def update_user_role(user_id):
    """
    🎭 تحديث دور المستخدم - Update User Role
    
    📋 الوصف:
        تغيير دور/صلاحيات المستخدم
    
    📥 Parameters:
        - user_id (int): معرّف المستخدم
        - role_id (int): الدور الجديد
    
    🔒 Security:
        - Owner only
        - حماية المالك
        - Audit logging
    """
    user = db.get_or_404(User, user_id)
    
    if user.is_system_account:
        flash('🔒 لا يمكن تغيير دور حسابات النظام', 'warning')
        return redirect(url_for('security.user_control'))
    
    old_role_id = user.role_id
    new_role_id = request.form.get('role_id', type=int)
    
    user.role_id = new_role_id
    
    # Audit log
    try:
        from models import Role
        old_role = db.session.get(Role, old_role_id) if old_role_id else None
        new_role = db.session.get(Role, new_role_id) if new_role_id else None
        
        log = AuditLog(
            user_id=current_user.id,
            action='security.update_user_role',
            table_name='users',
            record_id=user.id,
            old_data=json.dumps({'role': old_role.name if old_role else None}, ensure_ascii=False),
            new_data=json.dumps({'role': new_role.name if new_role else None}, ensure_ascii=False),
            note=f'Owner changed role for {user.username}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
    except Exception:
        pass
    
    db.session.commit()
    flash(f'✅ تم تحديث دور {user.username}', 'success')
    return redirect(url_for('security.user_control'))


@security_bp.route('/update-user-permissions/<int:user_id>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def update_user_extra_permissions(user_id):
    """تحديث الصلاحيات الإضافية للمستخدم - Owner Only"""
    from models import User, Permission
    
    user = db.get_or_404(User, user_id)
    data = request.get_json()
    permission_codes = data.get('permissions', [])
    
    permissions_to_add = []
    for code in permission_codes:
        perm = Permission.query.filter_by(code=code).first()
        if perm:
            permissions_to_add.append(perm)
    
    if hasattr(user.extra_permissions, 'clear'):
        user.extra_permissions.clear()
    else:
        user.extra_permissions = []
    
    for perm in permissions_to_add:
        user.extra_permissions.append(perm)
    
    db.session.commit()
    
    utils.clear_user_permission_cache(user.id)
    
    try:
        db.session.add(AuditLog(
            model_name="User",
            record_id=int(user.id),
            customer_id=getattr(current_user, "customer_id", None),
            user_id=getattr(current_user, "id", None),
            action="UPDATE_EXTRA_PERMISSIONS",
            old_data=None,
            new_data=f"extra_permissions_count={len(permissions_to_add)}",
            ip_address=request.remote_addr,
            user_agent=request.headers.get("User-Agent"),
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
    
    return jsonify({
        'success': True,
        'message': 'تم حفظ الصلاحيات بنجاح',
        'count': len(permissions_to_add)
    })


@security_bp.route('/settings', methods=['GET', 'POST'])
@security_bp.route('/system-settings', methods=['GET', 'POST'])  # Backward compatibility
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def system_settings():
    """إعدادات النظام الموحدة - 4 في 1 (عامة + متقدمة + شركة + ثوابت أعمال)"""
    from models import SystemSettings
    
    tab = request.args.get('tab', 'general')  # general, advanced, company, business
    
    if request.method == 'POST':
        tab = request.form.get('active_tab', 'general')
        
        if tab == 'general':
            # حفظ الإعدادات العامة
            settings = {
                'maintenance_mode': request.form.get('maintenance_mode') == 'on',
                'registration_enabled': request.form.get('registration_enabled') == 'on',
                'api_enabled': request.form.get('api_enabled') == 'on',
            }
            for key, value in settings.items():
                _set_system_setting(key, value)
            flash('✅ تم حفظ الإعدادات العامة', 'success')
            
        elif tab == 'advanced':
            # حفظ التكوينات المتقدمة
            config = {
                'SESSION_TIMEOUT': request.form.get('session_timeout', 3600),
                'MAX_LOGIN_ATTEMPTS': request.form.get('max_login_attempts', 5),
                'PASSWORD_MIN_LENGTH': request.form.get('password_min_length', 8),
                'AUTO_BACKUP_ENABLED': request.form.get('auto_backup_enabled') == 'on',
                'BACKUP_INTERVAL_HOURS': request.form.get('backup_interval_hours', 24),
                'ENABLE_EMAIL_NOTIFICATIONS': request.form.get('enable_email_notifications') == 'on',
                'ENABLE_SMS_NOTIFICATIONS': request.form.get('enable_sms_notifications') == 'on',
            }
            for key, value in config.items():
                _set_system_setting(key, value)
            flash('✅ تم تحديث التكوين المتقدم', 'success')
            
        elif tab == 'company':
            # حفظ بيانات الشركة
            constants = {
                'COMPANY_NAME': request.form.get('company_name', ''),
                'COMPANY_ADDRESS': request.form.get('company_address', ''),
                'COMPANY_PHONE': request.form.get('company_phone', ''),
                'COMPANY_EMAIL': request.form.get('company_email', ''),
                'TAX_NUMBER': request.form.get('tax_number', ''),
                'CURRENCY_SYMBOL': request.form.get('currency_symbol', '$'),
                'TIMEZONE': request.form.get('timezone', 'UTC'),
                'DATE_FORMAT': request.form.get('date_format', '%Y-%m-%d'),
                'TIME_FORMAT': request.form.get('time_format', '%H:%M:%S'),
            }
            for key, value in constants.items():
                if value:
                    _set_system_setting(key, value)
            flash('✅ تم تحديث بيانات الشركة', 'success')
            
        elif tab == 'branding':
            # حفظ الهوية والمظهر
            system_name = request.form.get('system_name')
            company_name = request.form.get('company_name')
            
            if system_name: 
                _set_system_setting('system_name', system_name)
            if company_name: 
                _set_system_setting('company_name', company_name)
            
            # معالجة الشعار
            if 'custom_logo' in request.files:
                file = request.files['custom_logo']
                if file and file.filename:
                    from werkzeug.utils import secure_filename
                    import os
                    filename = secure_filename(file.filename)
                    # حفظ في static/img/uploads
                    upload_folder = os.path.join(current_app.root_path, 'static', 'img', 'uploads')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    file_path = os.path.join(upload_folder, filename)
                    file.save(file_path)
                    
                    # حفظ المسار النسبي
                    _set_system_setting('custom_logo', f'static/img/uploads/{filename}')
            
            flash('✅ تم تحديث هوية النظام بنجاح', 'success')
            
        elif tab == 'business':
            # حفظ ثوابت الأعمال (Business Constants)
            try:
                group_flags = {
                    'tax': request.form.get('enable_tax_constants') == 'on',
                    'payroll': request.form.get('enable_payroll_constants') == 'on',
                    'assets': request.form.get('enable_asset_constants') == 'on',
                    'accounting': request.form.get('enable_accounting_constants') == 'on',
                    'notifications': request.form.get('enable_notification_constants') == 'on',
                    'business_rules': request.form.get('enable_business_rules_constants') == 'on',
                    'multi_tenancy': request.form.get('enable_multi_tenancy_constants') == 'on',
                }
                
                descriptions = {
                    'tax': 'تفعيل ثوابت الضرائب',
                    'payroll': 'تفعيل ثوابت الرواتب',
                    'assets': 'تفعيل ثوابت الأصول الثابتة',
                    'accounting': 'تفعيل ثوابت المحاسبة',
                    'notifications': 'تفعيل ثوابت الإشعارات',
                    'business_rules': 'تفعيل ثوابت قواعد العمل',
                    'multi_tenancy': 'تفعيل ثوابت التعددية',
                }
                
                for group, enabled in group_flags.items():
                    SystemSettings.set_setting(
                        f'enable_{group}_constants',
                        enabled,
                        descriptions.get(group),
                        'boolean'
                    )
                
                # Tax Settings
                if group_flags['tax']:
                    SystemSettings.set_setting('default_vat_rate', request.form.get('default_vat_rate', 16.0), 
                                               'نسبة VAT الافتراضية', 'number')
                    SystemSettings.set_setting('vat_enabled', request.form.get('vat_enabled') == 'on', 
                                               'تفعيل VAT', 'boolean')
                    SystemSettings.set_setting('income_tax_rate', request.form.get('income_tax_rate', 15.0), 
                                               'ضريبة دخل الشركات', 'number')
                    SystemSettings.set_setting('withholding_tax_rate', request.form.get('withholding_tax_rate', 5.0), 
                                               'الخصم من المنبع', 'number')
                
                # Payroll Settings
                if group_flags['payroll']:
                    SystemSettings.set_setting('social_insurance_enabled', request.form.get('social_insurance_enabled') == 'on', 
                                               'تفعيل التأمينات', 'boolean')
                    SystemSettings.set_setting('social_insurance_company', request.form.get('social_insurance_company', 7.5), 
                                               'نسبة التأمين - الشركة', 'number')
                    SystemSettings.set_setting('social_insurance_employee', request.form.get('social_insurance_employee', 7.0), 
                                               'نسبة التأمين - الموظف', 'number')
                    SystemSettings.set_setting('overtime_rate_normal', request.form.get('overtime_rate_normal', 1.5), 
                                               'معدل العمل الإضافي', 'number')
                    SystemSettings.set_setting('working_hours_per_day', request.form.get('working_hours_per_day', 8), 
                                               'ساعات العمل اليومية', 'number')
                
                # Fixed Assets Settings
                if group_flags['assets']:
                    SystemSettings.set_setting('asset_auto_depreciation', request.form.get('asset_auto_depreciation') == 'on', 
                                               'استهلاك تلقائي', 'boolean')
                    SystemSettings.set_setting('asset_threshold_amount', request.form.get('asset_threshold_amount', 500), 
                                               'حد مبلغ الأصول', 'number')
                
                # Accounting Settings
                if group_flags['accounting']:
                    SystemSettings.set_setting('cost_centers_enabled', request.form.get('cost_centers_enabled') == 'on', 
                                               'تفعيل مراكز التكلفة', 'boolean')
                    SystemSettings.set_setting('budgeting_enabled', request.form.get('budgeting_enabled') == 'on', 
                                               'تفعيل الموازنات', 'boolean')
                    SystemSettings.set_setting('fiscal_year_start_month', request.form.get('fiscal_year_start_month', 1), 
                                               'بداية السنة المالية', 'number')
                
                # Notification Settings
                if group_flags['notifications']:
                    SystemSettings.set_setting('notify_on_service_complete', request.form.get('notify_on_service_complete') == 'on', 
                                               'إشعار اكتمال الصيانة', 'boolean')
                    SystemSettings.set_setting('notify_on_payment_due', request.form.get('notify_on_payment_due') == 'on', 
                                               'إشعار استحقاق الدفعات', 'boolean')
                    SystemSettings.set_setting('notify_on_low_stock', request.form.get('notify_on_low_stock') == 'on', 
                                               'تنبيه انخفاض المخزون', 'boolean')
                    SystemSettings.set_setting('payment_reminder_days', request.form.get('payment_reminder_days', 3), 
                                               'التذكير قبل الاستحقاق', 'number')
                
                # Business Rules
                if group_flags['business_rules']:
                    SystemSettings.set_setting('allow_negative_stock', request.form.get('allow_negative_stock') == 'on', 
                                               'السماح بالمخزون السالب', 'boolean')
                    SystemSettings.set_setting('require_approval_for_sales_above', request.form.get('require_approval_for_sales_above', 10000), 
                                               'طلب موافقة للمبيعات الكبيرة', 'number')
                    SystemSettings.set_setting('discount_max_percent', request.form.get('discount_max_percent', 50), 
                                               'الحد الأقصى للخصم', 'number')
                    SystemSettings.set_setting('credit_limit_check', request.form.get('credit_limit_check') == 'on', 
                                               'فحص حد الائتمان', 'boolean')
                
                # Multi-Tenancy Settings  
                if group_flags['multi_tenancy']:
                    SystemSettings.set_setting('multi_tenancy_enabled', request.form.get('multi_tenancy_enabled') == 'on', 
                                               'تفعيل تعدد المستأجرين', 'boolean')
                    SystemSettings.set_setting('trial_period_days', request.form.get('trial_period_days', 30), 
                                               'مدة التجريبي', 'number')
                
                db.session.commit()
                flash('✅ تم حفظ ثوابت الأعمال', 'success')
                
            except Exception as e:
                db.session.rollback()
                flash(f'❌ خطأ: {str(e)}', 'danger')
        
        return redirect(url_for('security.system_settings', tab=tab))
    
    # قراءة جميع الإعدادات
    data = {
        'general': {
            'maintenance_mode': _get_system_setting('maintenance_mode', False),
            'registration_enabled': _get_system_setting('registration_enabled', True),
            'api_enabled': _get_system_setting('api_enabled', True),
        },
        'advanced': {
            'SESSION_TIMEOUT': _get_system_setting('SESSION_TIMEOUT', 3600),
            'MAX_LOGIN_ATTEMPTS': _get_system_setting('MAX_LOGIN_ATTEMPTS', 5),
            'PASSWORD_MIN_LENGTH': _get_system_setting('PASSWORD_MIN_LENGTH', 8),
            'AUTO_BACKUP_ENABLED': _get_system_setting('AUTO_BACKUP_ENABLED', True),
            'BACKUP_INTERVAL_HOURS': _get_system_setting('BACKUP_INTERVAL_HOURS', 24),
            'ENABLE_EMAIL_NOTIFICATIONS': _get_system_setting('ENABLE_EMAIL_NOTIFICATIONS', True),
            'ENABLE_SMS_NOTIFICATIONS': _get_system_setting('ENABLE_SMS_NOTIFICATIONS', False),
        },
        'company': {
            'COMPANY_NAME': _get_system_setting('COMPANY_NAME', 'اسم الشركة'),
            'COMPANY_ADDRESS': _get_system_setting('COMPANY_ADDRESS', ''),
            'COMPANY_PHONE': _get_system_setting('COMPANY_PHONE', ''),
            'COMPANY_EMAIL': _get_system_setting('COMPANY_EMAIL', ''),
            'TAX_NUMBER': _get_system_setting('TAX_NUMBER', ''),
            'CURRENCY_SYMBOL': _get_system_setting('CURRENCY_SYMBOL', '$'),
            'TIMEZONE': _get_system_setting('TIMEZONE', 'UTC'),
            'DATE_FORMAT': _get_system_setting('DATE_FORMAT', '%Y-%m-%d'),
            'TIME_FORMAT': _get_system_setting('TIME_FORMAT', '%H:%M:%S'),
        },
        'business': {
            'enable_tax_constants': SystemSettings.get_setting('enable_tax_constants', True),
            'enable_payroll_constants': SystemSettings.get_setting('enable_payroll_constants', True),
            'enable_asset_constants': SystemSettings.get_setting('enable_asset_constants', True),
            'enable_accounting_constants': SystemSettings.get_setting('enable_accounting_constants', True),
            'enable_notification_constants': SystemSettings.get_setting('enable_notification_constants', True),
            'enable_business_rules_constants': SystemSettings.get_setting('enable_business_rules_constants', True),
            'enable_multi_tenancy_constants': SystemSettings.get_setting('enable_multi_tenancy_constants', True),
            # Tax
            'default_vat_rate': SystemSettings.get_setting('default_vat_rate', 16.0),
            'vat_enabled': SystemSettings.get_setting('vat_enabled', True),
            'income_tax_rate': SystemSettings.get_setting('income_tax_rate', 15.0),
            'withholding_tax_rate': SystemSettings.get_setting('withholding_tax_rate', 5.0),
            # Payroll
            'social_insurance_enabled': SystemSettings.get_setting('social_insurance_enabled', False),
            'social_insurance_company': SystemSettings.get_setting('social_insurance_company', 7.5),
            'social_insurance_employee': SystemSettings.get_setting('social_insurance_employee', 7.0),
            'overtime_rate_normal': SystemSettings.get_setting('overtime_rate_normal', 1.5),
            'working_hours_per_day': SystemSettings.get_setting('working_hours_per_day', 8),
            # Assets
            'asset_auto_depreciation': SystemSettings.get_setting('asset_auto_depreciation', True),
            'asset_threshold_amount': SystemSettings.get_setting('asset_threshold_amount', 500),
            # Accounting
            'cost_centers_enabled': SystemSettings.get_setting('cost_centers_enabled', False),
            'budgeting_enabled': SystemSettings.get_setting('budgeting_enabled', False),
            'fiscal_year_start_month': SystemSettings.get_setting('fiscal_year_start_month', 1),
            # Notifications
            'notify_on_service_complete': SystemSettings.get_setting('notify_on_service_complete', True),
            'notify_on_payment_due': SystemSettings.get_setting('notify_on_payment_due', True),
            'notify_on_low_stock': SystemSettings.get_setting('notify_on_low_stock', True),
            'payment_reminder_days': SystemSettings.get_setting('payment_reminder_days', 3),
            # Business Rules
            'allow_negative_stock': SystemSettings.get_setting('allow_negative_stock', False),
            'require_approval_for_sales_above': SystemSettings.get_setting('require_approval_for_sales_above', 10000),
            'discount_max_percent': SystemSettings.get_setting('discount_max_percent', 50),
            'credit_limit_check': SystemSettings.get_setting('credit_limit_check', True),
            # Multi-Tenancy
            'multi_tenancy_enabled': SystemSettings.get_setting('multi_tenancy_enabled', False),
            'trial_period_days': SystemSettings.get_setting('trial_period_days', 30),
        },
        'branding': {
            'system_name': _get_system_setting('system_name', 'نظام الحازم'),
            'company_name': _get_system_setting('company_name', 'شركة الحازم للأنظمة الذكية'),
            'company_logo': _get_system_setting('custom_logo', ''),
        }
    }
    
    # 🔄 إضافة إحصائيات ديناميكية
    stats = get_cached_security_stats()
    
    return render_template('security/system_settings.html', data=data, active_tab=tab, stats=stats)


@security_bp.route('/emergency-tools')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def emergency_tools():
    """أدوات الطوارئ"""
    from models import User, AuditLog
    
    emergency_data = {
        'maintenance_mode': _get_setting('maintenance_mode', 'false') == 'true',
        'total_users': User.query.count(),
        'active_users': User.query.filter_by(is_active=True).count(),
        'blocked_users': User.query.filter_by(is_active=False).count(),
        'recent_errors': AuditLog.query.filter(
            AuditLog.action.like('%error%') | AuditLog.action.like('%failed%')
        ).order_by(AuditLog.created_at.desc()).limit(10).count(),
    }
    
    return render_template('security/emergency_tools.html', emergency_data=emergency_data)


@security_bp.route('/emergency/maintenance-mode', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def toggle_maintenance_mode():
    """تفعيل/تعطيل وضع الصيانة"""
    current = _get_system_setting('maintenance_mode', False)
    _set_system_setting('maintenance_mode', not current)
    
    status = 'مفعل' if not current else 'معطل'
    flash(f'وضع الصيانة الآن {status}', 'warning')
    return redirect(url_for('security.emergency_tools'))


@security_bp.route('/emergency/clear-cache', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def clear_system_cache():
    """مسح الكاش بالكامل"""
    from extensions import cache
    cache.clear()
    flash('تم مسح الكاش بالكامل', 'success')
    return redirect(url_for('security.emergency_tools'))


@security_bp.route('/emergency/kill-sessions', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def kill_all_sessions():
    """إنهاء جميع الجلسات"""
    # إنهاء جميع الجلسات النشطة
    _kill_all_user_sessions()
    flash('تم إنهاء جميع الجلسات', 'warning')
    return redirect(url_for('security.emergency_tools'))


@security_bp.route('/data-export')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def data_export():
    """تصدير البيانات"""
    tables = _get_all_tables()
    return render_template('security/data_export.html', tables=tables)


@security_bp.route('/export-table/<table_name>')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def export_table_csv(table_name):
    """تصدير جدول كـ CSV"""
    import csv
    from io import StringIO
    
    data, columns = _browse_table(table_name, limit=10000)
    
    si = StringIO()
    writer = csv.DictWriter(si, fieldnames=columns)
    writer.writeheader()
    writer.writerows(data)
    
    output = si.getvalue()
    
    from flask import make_response
    response = make_response(output)
    response.headers["Content-Disposition"] = f"attachment; filename={table_name}.csv"
    response.headers["Content-type"] = "text/csv"
    
    return response




@security_bp.route('/performance-monitor')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def performance_monitor():
    """مراقبة الأداء"""
    performance = {
        'db_queries': _get_slow_queries(),
        'response_times': _get_avg_response_times(),
        'memory_usage': _get_memory_usage(),
        'cpu_usage': _get_cpu_usage(),
    }
    return render_template('security/performance_monitor.html', performance=performance)


@security_bp.route('/system-branding', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def system_branding():
    """تخصيص هوية النظام - الألوان والشعارات"""
    from models import SystemSettings
    
    if request.method == 'POST':
        updated = []
        
        # 1. تحديث الألوان
        colors = {
            'primary_color': request.form.get('primary_color'),
            'secondary_color': request.form.get('secondary_color'),
            'sidebar_bg': request.form.get('sidebar_bg'),
            'sidebar_text': request.form.get('sidebar_text'),
        }
        
        for key, value in colors.items():
            if value:
                SystemSettings.set_setting(key, value) # Use direct model method
                updated.append('الألوان')
        
        # 2. تحديث النصوص (اسم النظام واسم الشركة)
        texts = {
            'system_name': request.form.get('system_name'),
            'company_name': request.form.get('company_name'),
            'login_title': request.form.get('login_title'),
            'login_subtitle': request.form.get('login_subtitle'),
            'footer_text': request.form.get('footer_text'),
        }
        
        for key, value in texts.items():
            if value:
                SystemSettings.set_setting(key, value)
                updated.append('النصوص')

        # الشعار
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                # التحقق من نوع الملف
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
                file_ext = logo_file.filename.rsplit('.', 1)[1].lower() if '.' in logo_file.filename else ''
                
                if file_ext in allowed_extensions:
                    filename = secure_filename(logo_file.filename)
                    os.makedirs('static/img', exist_ok=True)
                    logo_path = f'static/img/custom_logo_{filename}'
                    logo_file.save(logo_path)
                    SystemSettings.set_setting('custom_logo', logo_path)
                    updated.append('الشعار')
                else:
                    flash('⚠️ نوع ملف الشعار غير مدعوم (استخدم: png, jpg, jpeg, gif, webp)', 'warning')
        
        # الأيقونة
        if 'favicon' in request.files:
            favicon_file = request.files['favicon']
            if favicon_file and favicon_file.filename:
                allowed_extensions = {'png', 'ico'}
                file_ext = favicon_file.filename.rsplit('.', 1)[1].lower() if '.' in favicon_file.filename else ''
                
                if file_ext in allowed_extensions:
                    filename = secure_filename(favicon_file.filename)
                    favicon_path = f'static/favicon_custom_{filename}'
                    favicon_file.save(favicon_path)
                    SystemSettings.set_setting('custom_favicon', favicon_path)
                    updated.append('الأيقونة')
                else:
                    flash('⚠️ نوع ملف الأيقونة غير مدعوم (استخدم: png, ico)', 'warning')
        
        if updated:
            flash(f'✅ تم تحديث: {", ".join(updated)} بنجاح!', 'success')
            
            # تسجيل في AuditLog
            try:
                import json
                log = AuditLog(
                    user_id=current_user.id,
                    action='UPDATE',
                    model_name='SystemSettings',
                    new_data=json.dumps({'updated_fields': updated, 'note': 'Updated branding settings'}),
                    ip_address=request.remote_addr
                )
                db.session.add(log)
                db.session.commit()
            except Exception as e:
                # Log error but don't stop the process
                current_app.logger.error(f"Audit log failed: {e}")
                pass
        else:
            flash('ℹ️ لم يتم تحديث أي شيء', 'info')
        
        return redirect(url_for('security.system_branding'))
    
    # قراءة الإعدادات الحالية
    branding = {
        'system_name': _get_system_setting('system_name', 'نظام إدارة الكراج'),
        'company_name': _get_system_setting('company_name', 'شركة ازاد للأنظمة الذكية'),
        'login_title': _get_system_setting('login_title', 'مرحباً بك'),
        'login_subtitle': _get_system_setting('login_subtitle', 'سجل دخولك للمتابعة'),
        'footer_text': _get_system_setting('footer_text', 'جميع الحقوق محفوظة'),
        'primary_color': _get_system_setting('primary_color', '#06b6d4'),
        'secondary_color': _get_system_setting('secondary_color', '#1f2937'),
        'sidebar_bg': _get_system_setting('sidebar_bg', '#111827'),
        'sidebar_text': _get_system_setting('sidebar_text', '#f9fafb'),
        'custom_logo': _get_system_setting('custom_logo', ''),
        'custom_favicon': _get_system_setting('custom_favicon', ''),
    }
    
    return render_template('security/system_branding.html', branding=branding)









@security_bp.route('/db-editor/add-column/<table_name>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_add_column(table_name):
    """إضافة عمود جديد"""
    _assert_valid_table(table_name)
    column_name = request.form.get('column_name', '').strip()
    column_type = request.form.get('column_type', 'TEXT').strip()
    default_value = request.form.get('default_value', '')
    
    if not column_name or not _is_safe_identifier(column_name):
        flash('اسم العمود مطلوب', 'danger')
        return redirect(url_for('security.database_manager', tab='edit', table=table_name))

    normalized_type = _normalize_column_type(column_type)
    if not normalized_type:
        flash('نوع العمود غير صالح', 'danger')
        return redirect(url_for('security.database_manager', tab='edit', table=table_name))
    
    try:
        params = {}
        if default_value:
            sql = text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {normalized_type} DEFAULT :default_value")
            params['default_value'] = default_value
        else:
            sql = text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {normalized_type}")
        db.session.execute(sql, params)
        db.session.commit()
        
        flash(f'تم إضافة العمود {column_name} بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('security.database_manager', tab='edit', table=table_name))


@security_bp.route('/db-editor/update-cell/<table_name>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_update_cell(table_name):
    """تحديث خلية واحدة مباشرة - للتعديل السريع"""
    try:
        _assert_valid_table(table_name)
        data = request.get_json() or {}
        row_id = data.get('row_id')
        column = data.get('column')
        value = data.get('value')
        
        if not all([row_id, column]):
            return jsonify({'success': False, 'error': 'معلومات ناقصة'}), 400
        
        if not _is_safe_identifier(column):
            return jsonify({'success': False, 'error': 'اسم عمود غير صالح'}), 400

        primary_key = _get_primary_key_column(table_name) or 'id'
        
        # تحديث الخلية
        if not _is_safe_identifier(primary_key):
            return jsonify({'success': False, 'error': 'مفتاح أساسي غير صالح'}), 400

        sql = text(f"UPDATE {table_name} SET {column} = :value WHERE {primary_key} = :row_id")
        
        result = db.session.execute(sql, {'value': value, 'row_id': row_id})
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'تم تحديث {column} بنجاح',
            'rows_affected': result.rowcount
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@security_bp.route('/db-editor/edit-row/<table_name>/<int:row_id>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_edit_row(table_name, row_id):
    """تعديل صف في الجدول"""
    try:
        _assert_valid_table(table_name)
        allowed_columns = set(_get_table_columns_safe(table_name))
        # الحصول على جميع الحقول من الفورم
        updates = []
        for key, value in request.form.items():
            if key in ['csrf_token', 'id']:
                continue
            if key not in allowed_columns or not _is_safe_identifier(key):
                continue
            updates.append((key, value))
        
        if updates:
            set_parts = [f"{col} = :val_{i}" for i, (col, _) in enumerate(updates)]
            params = {f"val_{i}": v for i, (_, v) in enumerate(updates)}
            params['row_id'] = row_id
            sql = text(f"UPDATE {table_name} SET {', '.join(set_parts)} WHERE id = :row_id")
            db.session.execute(sql, params)
            db.session.commit()
            flash('تم التحديث بنجاح', 'success')
        else:
            flash('لا توجد تغييرات', 'warning')
    
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('security.database_manager', tab='edit', table=table_name))


@security_bp.route('/db-editor/delete-row/<table_name>/<row_id>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_delete_row(table_name, row_id):
    """حذف صف من الجدول"""
    try:
        _assert_valid_table(table_name)
        primary_key = _get_primary_key_column(table_name) or 'id'
        
        # حذف الصف
        if not _is_safe_identifier(primary_key):
            flash('❌ مفتاح أساسي غير صالح', 'danger')
            return redirect(url_for('security.database_manager', tab='edit', table=table_name))
        sql = text(f"DELETE FROM {table_name} WHERE {primary_key} = :row_id")
        
        result = db.session.execute(sql, {'row_id': row_id})
        db.session.commit()
        flash(f'✅ تم حذف الصف #{row_id} بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ خطأ في الحذف: {str(e)}', 'danger')
    
    return redirect(url_for('security.database_manager', tab='edit', table=table_name))

@security_bp.route('/db-editor/delete-column/<table_name>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_delete_column(table_name):
    """حذف عمود كامل من الجدول"""
    _assert_valid_table(table_name)
    column_name = request.form.get('column_name', '').strip()
    
    if not column_name:
        flash('❌ اسم العمود مطلوب', 'danger')
        return redirect(url_for('security.database_manager', tab='edit', table=table_name))
    
    # حماية من حذف الأعمدة الحرجة
    protected_columns = ['id', 'created_at', 'updated_at']
    if (column_name.lower() in protected_columns) or (not _is_safe_identifier(column_name)):
        flash(f'❌ لا يمكن حذف العمود {column_name} (محمي)', 'danger')
        return redirect(url_for('security.database_manager', tab='edit', table=table_name))
    
    try:
        cols = _get_table_columns_safe(table_name)
        if column_name not in cols:
            flash('❌ عمود غير موجود', 'danger')
            return redirect(url_for('security.database_manager', tab='edit', table=table_name))
        db.session.execute(text(f"ALTER TABLE {table_name} DROP COLUMN {column_name}"))
        db.session.commit()
        flash(f'✅ تم حذف العمود {column_name} بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ خطأ في حذف العمود: {str(e)}', 'danger')
    
    return redirect(url_for('security.database_manager', tab='edit', table=table_name))


@security_bp.route('/db-editor/add-row/<table_name>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_add_row(table_name):
    """إضافة صف جديد"""
    try:
        _assert_valid_table(table_name)
        allowed_columns = set(_get_table_columns_safe(table_name))
        cols = []
        params = {}
        placeholders = []
        idx = 0
        for key, value in request.form.items():
            if key == 'csrf_token':
                continue
            if key not in allowed_columns or not _is_safe_identifier(key):
                continue
            bind = f"v_{idx}"
            cols.append(key)
            placeholders.append(f":{bind}")
            params[bind] = value
            idx += 1
        if cols:
            sql = text(f"INSERT INTO {table_name} ({', '.join(cols)}) VALUES ({', '.join(placeholders)})")
            db.session.execute(sql, params)
            db.session.commit()
            flash('تم الإضافة بنجاح', 'success')
        else:
            flash('لا توجد بيانات', 'warning')
    
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('security.database_manager', tab='edit', table=table_name))


@security_bp.route('/db-editor/bulk-update/<table_name>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_bulk_update(table_name):
    """تحديث جماعي للبيانات"""
    _assert_valid_table(table_name)
    column = request.form.get('column', '')
    old_value = request.form.get('old_value', '')
    new_value = request.form.get('new_value', '')
    
    if not column:
        flash('اسم العمود مطلوب', 'danger')
        return redirect(url_for('security.database_manager', tab='edit', table=table_name))
    
    try:
        if not _is_safe_identifier(column) or column not in _get_table_columns_safe(table_name):
            flash('اسم عمود غير صالح', 'danger')
            return redirect(url_for('security.database_manager', tab='edit', table=table_name))
        if old_value:
            sql = text(f"UPDATE {table_name} SET {column} = :new_value WHERE {column} = :old_value")
            result = db.session.execute(sql, {"new_value": new_value, "old_value": old_value})
        else:
            sql = text(f"UPDATE {table_name} SET {column} = :new_value WHERE {column} IS NULL OR {column} = ''")
            result = db.session.execute(sql, {"new_value": new_value})
        db.session.commit()
        
        flash(f'تم تحديث {result.rowcount} صف بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('security.database_manager', tab='edit', table=table_name))


@security_bp.route('/db-editor/fill-missing/<table_name>', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_fill_missing(table_name):
    """ملء البيانات الناقصة"""
    _assert_valid_table(table_name)
    column = request.form.get('column', '')
    fill_value = request.form.get('fill_value', '')
    
    if not column:
        flash('اسم العمود مطلوب', 'danger')
        return redirect(url_for('security.database_manager', tab='edit', table=table_name))
    
    try:
        if not _is_safe_identifier(column) or column not in _get_table_columns_safe(table_name):
            flash('اسم عمود غير صالح', 'danger')
            return redirect(url_for('security.database_manager', tab='edit', table=table_name))
        sql = text(f"UPDATE {table_name} SET {column} = :fill_value WHERE {column} IS NULL OR {column} = ''")
        result = db.session.execute(sql, {"fill_value": fill_value})
        db.session.commit()
        
        flash(f'تم ملء {result.rowcount} حقل ناقص بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('security.database_manager', tab='edit', table=table_name))


@security_bp.route('/db-editor/schema/<table_name>')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def db_schema_editor(table_name):
    """Redirect to database_manager - schema tab"""
    return redirect(url_for('security.database_manager', tab='schema', table=table_name))



def _get_blocked_ips_count():
    """عدد IPs المحظورة"""
    blocked = cache.get('blocked_ips') or []
    return len(blocked)

def _get_blocked_countries_count():
    """عدد الدول المحظورة"""
    blocked = cache.get('blocked_countries') or []
    return len(blocked)

def _get_failed_logins_count(hours=24):
    """عدد محاولات تسجيل الدخول الفاشلة"""
    since = datetime.now(timezone.utc) - timedelta(hours=hours)
    return AuditLog.query.filter(
        AuditLog.action.in_(['login.failed', 'login.blocked']),
        AuditLog.created_at >= since
    ).count()

def _get_suspicious_activities_count(hours=24):
    """عدد الأنشطة المشبوهة"""
    since = datetime.now(timezone.utc) - timedelta(hours=hours)
    return AuditLog.query.filter(
        AuditLog.action.like('%suspicious%'),
        AuditLog.created_at >= since
    ).count()

def _get_recent_suspicious_activities(limit=10):
    """آخر الأنشطة المشبوهة"""
    return AuditLog.query.filter(
        AuditLog.action.like('%suspicious%')
    ).order_by(AuditLog.created_at.desc()).limit(limit).all()

def _block_ip(ip, reason, duration):
    """حظر IP"""
    blocked = cache.get('blocked_ips') or []
    
    # إضافة IP للقائمة
    blocked_entry = {
        'ip': ip,
        'reason': reason,
        'duration': duration,
        'blocked_at': datetime.now(timezone.utc).isoformat(),
        'blocked_by': current_user.id
    }
    
    blocked.append(blocked_entry)
    
    # حفظ في Cache
    if duration == 'permanent':
        cache.set('blocked_ips', blocked, timeout=0)  # لا ينتهي
    else:
        timeout = _parse_duration(duration)
        cache.set('blocked_ips', blocked, timeout=timeout)
    
    # تسجيل في Audit
    AuditLog(
        model_name='Security',
        action='IP_BLOCKED',
        user_id=current_user.id,
        old_data=json.dumps({'ip': ip, 'reason': reason}, ensure_ascii=False),
        ip_address=request.remote_addr
    )
    db.session.commit()

def _unblock_ip(ip):
    """إلغاء حظر IP"""
    blocked = cache.get('blocked_ips') or []
    blocked = [b for b in blocked if b.get('ip') != ip]
    cache.set('blocked_ips', blocked, timeout=0)

def _get_all_blocked_ips():
    """الحصول على جميع IPs المحظورة"""
    return cache.get('blocked_ips') or []

def _block_country(country_code, reason):
    """حظر دولة"""
    blocked = cache.get('blocked_countries') or []
    
    blocked_entry = {
        'country_code': country_code,
        'reason': reason,
        'blocked_at': datetime.now(timezone.utc).isoformat(),
        'blocked_by': current_user.id
    }
    
    blocked.append(blocked_entry)
    cache.set('blocked_countries', blocked, timeout=0)

def _get_all_blocked_countries():
    """الحصول على جميع الدول المحظورة"""
    return cache.get('blocked_countries') or []

def _get_cleanable_tables():
    """
    🔄 الجداول القابلة للتنظيف - تحديث تلقائي من قاعدة البيانات
    
    ✅ يتحدث تلقائياً بناءً على التهجيرات (Migrations)
    🛡️ يستثني الجداول الحساسة (alembic_version, system_settings)
    ⚠️ يحدد مستوى الخطورة بذكاء حسب نوع الجدول
    """
    from sqlalchemy import inspect, text
    
    # الحصول على جميع الجداول من قاعدة البيانات
    inspector = inspect(db.engine)
    all_tables = inspector.get_table_names()
    
    # الجداول التي لا يجب حذفها أبداً
    SYSTEM_TABLES = {
        'alembic_version',      # تاريخ التهجيرات
        'system_settings',      # إعدادات النظام
        'branches',             # الفروع
        'currencies',           # العملات
        'accounts',             # دليل الحسابات
    }
    
    # قواعد تحديد مستوى الخطورة والأيقونات
    DANGER_RULES = {
        # خطر عالي جداً - بيانات أساسية
        'high': {
            'keywords': ['customer', 'supplier', 'partner', 'user', 'payment', 'sale', 'invoice', 
                        'gl_', 'check', 'warehouse', 'product', 'shipment', 'stock_level', 'employee'],
            'icon': '🔴'
        },
        # خطر متوسط - بيانات مهمة
        'medium': {
            'keywords': ['service_', 'stock_adjustment', 'preorder', 'expense', 'settlement', 
                        'note', 'category', 'type', 'loan', 'partner'],
            'icon': '🟡'
        },
        # خطر منخفض - سجلات ولوجات
        'low': {
            'keywords': ['log', 'audit', 'notification', 'cart', 'rating', 'helpful'],
            'icon': '🟢'
        }
    }
    
    # قاموس الأيقونات حسب نوع الجدول
    TABLE_ICONS = {
        'user': '👤', 'role': '🎭', 'customer': '👥', 'supplier': '🏭', 'partner': '🤝',
        'payment': '💰', 'check': '📝', 'expense': '📤', 'sale': '🛍️', 'invoice': '📄',
        'product': '📦', 'warehouse': '🏪', 'stock': '📊', 'shipment': '🚚',
        'service': '🔧', 'cart': '🛒', 'preorder': '📅', 'settlement': '💼',
        'log': '📋', 'audit': '🔍', 'notification': '🔔', 'note': '📝',
        'gl_': '📖', 'account': '💼', 'category': '🏷️', 'type': '📂',
        'employee': '👔', 'advance': '💵', 'deduction': '➖'
    }
    
    def get_danger_level(table_name):
        """تحديد مستوى الخطورة بناءً على اسم الجدول"""
        table_lower = table_name.lower()
        for level, rules in DANGER_RULES.items():
            if any(keyword in table_lower for keyword in rules['keywords']):
                return level
        return 'medium'  # افتراضي
    
    def get_icon(table_name):
        """الحصول على الأيقونة المناسبة للجدول"""
        table_lower = table_name.lower()
        for keyword, icon in TABLE_ICONS.items():
            if keyword in table_lower:
                return icon
        return '📊'  # أيقونة افتراضية
    
    def get_arabic_name(table_name):
        """توليد اسم عربي للجدول"""
        # قاموس الترجمات
        translations = {
            'users': 'المستخدمين', 'roles': 'الأدوار', 'customers': 'العملاء',
            'suppliers': 'الموردين', 'partners': 'الشركاء', 'employees': 'الموظفين',
            'payments': 'المدفوعات', 'checks': 'الشيكات', 'expenses': 'المصاريف',
            'sales': 'المبيعات', 'invoices': 'الفواتير', 'products': 'المنتجات',
            'warehouses': 'المخازن', 'shipments': 'الشحنات', 'stock_levels': 'مستويات المخزون',
            'service_requests': 'طلبات الصيانة', 'service_parts': 'قطع الصيانة',
            'service_tasks': 'مهام الصيانة', 'audit_logs': 'سجلات التدقيق',
            'deletion_logs': 'سجلات الحذف', 'notes': 'الملاحظات',
            'online_carts': 'سلات التسوق', 'preorders': 'الحجوزات المسبقة',
            'gl_batches': 'دفعات القيود', 'gl_entries': 'القيود المحاسبية',
            'product_categories': 'فئات المنتجات', 'expense_types': 'أنواع المصاريف',
            'payment_splits': 'تقسيمات الدفع', 'sale_lines': 'بنود المبيعات',
            'invoice_lines': 'بنود الفواتير', 'shipment_items': 'بنود الشحنات',
            'stock_adjustments': 'تعديلات المخزون', 'exchange_transactions': 'معاملات التبادل',
            'supplier_settlements': 'تسويات الموردين', 'partner_settlements': 'تسويات الشركاء',
            'product_supplier_loans': 'قروض الموردين', 'utility_accounts': 'حسابات المرافق',
            'equipment_types': 'أنواع المركبات', 'online_payments': 'المدفوعات الإلكترونية',
            'online_preorders': 'الحجوزات الإلكترونية', 'product_partners': 'ربط المنتجات بالشركاء',
            'shipment_partners': 'ربط الشحنات بالشركاء', 'notifications': 'الإشعارات',
            'product_ratings': 'تقييمات المنتجات', 'employee_advances': 'سلف الموظفين',
            'employee_deductions': 'استقطاعات الموظفين', 'saas_subscriptions': 'اشتراكات SaaS',
            'auth_audit': 'سجلات المصادقة', 'archives': 'الأرشيفات',
            'import_runs': 'عمليات الاستيراد', 'user_branches': 'فروع المستخدمين',
            'user_permissions': 'صلاحيات المستخدمين', 'role_permissions': 'صلاحيات الأدوار',
            'permissions': 'الصلاحيات', 'sites': 'المواقع', 'transfers': 'التحويلات',
            'sale_returns': 'مرتجعات المبيعات', 'sale_return_lines': 'بنود المرتجعات',
            'customer_loyalty': 'برنامج الولاء', 'customer_loyalty_points': 'نقاط الولاء',
            'warehouse_partner_shares': 'حصص الشركاء بالمخازن',
            'employee_advance_installments': 'أقساط سلف الموظفين',
            'supplier_loan_settlements': 'تسويات قروض الموردين',
            'partner_settlement_lines': 'بنود تسويات الشركاء',
            'supplier_settlement_lines': 'بنود تسويات الموردين',
            'online_cart_items': 'محتويات السلة', 'online_preorder_items': 'بنود الحجوزات',
            'product_rating_helpful': 'تقييمات مفيدة', 'saas_plans': 'خطط SaaS',
            'saas_invoices': 'فواتير SaaS'
        }
        
        return translations.get(table_name, table_name.replace('_', ' ').title())
    
    # بناء قائمة الجداول
    cleanable_tables = []
    
    # إضافة خيار خاص للمستخدمين (حذف الكل ما عدا أول مدير نظام)
    cleanable_tables.append({
        'name': 'users_except_first_super',
        'display': '👤 المستخدمين (ما عدا أول مدير نظام)',
        'danger': 'high',
        'category': 'المستخدمين والأدوار'
    })
    
    # تصنيف الجداول حسب النوع
    categories = {
        'المستخدمين والأدوار': [],
        'السجلات واللوجات': [],
        'العمليات المالية': [],
        'المبيعات والصيانة': [],
        'المخزون والمنتجات': [],
        'الجهات': [],
        'التسوق الإلكتروني': [],
        'العمليات المحاسبية': [],
        'أخرى': []
    }
    
    def get_category(table_name):
        """تحديد التصنيف"""
        if 'user' in table_name or 'role' in table_name or 'permission' in table_name:
            return 'المستخدمين والأدوار'
        elif 'log' in table_name or 'audit' in table_name or 'notification' in table_name:
            return 'السجلات واللوجات'
        elif 'payment' in table_name or 'check' in table_name or 'expense' in table_name:
            return 'العمليات المالية'
        elif 'sale' in table_name or 'service' in table_name or 'invoice' in table_name:
            return 'المبيعات والصيانة'
        elif 'stock' in table_name or 'product' in table_name or 'warehouse' in table_name:
            return 'المخزون والمنتجات'
        elif 'customer' in table_name or 'supplier' in table_name or 'partner' in table_name or 'employee' in table_name:
            return 'الجهات'
        elif 'online' in table_name or 'cart' in table_name or 'preorder' in table_name:
            return 'التسوق الإلكتروني'
        elif 'gl_' in table_name or 'account' in table_name:
            return 'العمليات المحاسبية'
        else:
            return 'أخرى'
    
    # إضافة الجداول من قاعدة البيانات
    for table_name in sorted(all_tables):
        # تجاوز الجداول الحساسة
        if table_name in SYSTEM_TABLES:
            continue
        
        # تجاوز جدول المستخدمين (تمت إضافته يدوياً)
        if table_name == 'users':
            continue
        
        danger_level = get_danger_level(table_name)
        icon = get_icon(table_name)
        arabic_name = get_arabic_name(table_name)
        category = get_category(table_name)
        
        cleanable_tables.append({
            'name': table_name,
            'display': f'{icon} {arabic_name}',
            'danger': danger_level,
            'category': category
        })
    
    return cleanable_tables

def _cleanup_tables(tables):
    """تنظيف الجداول المحددة"""
    cleaned = 0
    errors = []
    
    for table in tables:
        try:
            # معالجة خاصة للمستخدمين - حذف الكل (حتى الأدمنز) ما عدا أول مدير نظام
            if table == 'users_except_first_super':
                from models import User
                # البحث عن أول مدير نظام (الأقدم)
                first_super = User.query.filter_by(is_super_admin=True).order_by(User.id.asc()).first()
                
                if first_super:
                    first_super_id = first_super.id
                    # حذف جميع المستخدمين (بما فيهم الأدمنز الآخرين) ما عدا أول مدير نظام
                    deleted_count = db.session.execute(
                        text("DELETE FROM users WHERE id != :super_id"), 
                        {'super_id': first_super_id}
                    ).rowcount
                    db.session.commit()
                    current_app.logger.info(f"[INFO] Deleted {deleted_count} users, kept first system admin (ID: {first_super_id})")
                    cleaned += 1
                else:
                    # إذا لم يوجد مدير نظام، لا نحذف شيء للحماية
                    errors.append(f"تخطي {table}: لا يوجد مدير نظام!")
                    continue
            else:
                # تنظيف عادي للجداول الأخرى
                try:
                    deleted_count = db.session.execute(text(f"DELETE FROM {table}")).rowcount
                    db.session.commit()
                    current_app.logger.info(f"[INFO] Cleaned table '{table}': {deleted_count} rows deleted")
                    cleaned += 1
                except Exception as delete_error:
                    # قد لا يكون الجدول موجوداً
                    db.session.rollback()
                    current_app.logger.warning(f"[WARNING] Table '{table}' not found or error: {str(delete_error)}")
                    continue
            
            # تسجيل في Audit (إذا لم يتم حذف audit_logs نفسه)
            if table != 'audit_logs':
                try:
                    db.session.add(AuditLog(
                        model_name='Security',
                        action='TABLE_CLEANED',
                        user_id=current_user.id,
                        old_data=json.dumps({'table': table}, ensure_ascii=False),
                        ip_address=request.remote_addr
                    ))
                    db.session.commit()
                except Exception:
                    pass  # إذا تم حذف audit_logs، نتجاوز
                    
        except Exception as e:
            db.session.rollback()
            error_msg = f"Failed to clean table {table}: {str(e)}"
            current_app.logger.error(f"[ERROR] {error_msg}")
            errors.append(error_msg)
            continue
    
    return {'cleaned': cleaned, 'total': len(tables), 'errors': errors}

def _parse_duration(duration):
    """تحويل المدة إلى ثواني"""
    if duration == '1h':
        return 3600
    elif duration == '24h':
        return 86400
    elif duration == '7d':
        return 604800
    elif duration == '30d':
        return 2592000
    else:
        return 0  # permanent



def _ai_security_analysis(query):
    """تحليل أمني بالذكاء الاصطناعي"""
    analysis = {
        'query': query,
        'type': 'security_analysis',
        'findings': [],
        'recommendations': [],
        'threat_level': 'low'
    }
    
    query_lower = query.lower()
    
    # تحليل ذكي بناءً على السؤال
    if 'ip' in query_lower or 'عنوان' in query_lower:
        analysis['findings'].append('فحص IPs المشبوهة...')
        analysis['findings'].append(f'عدد IPs المحظورة: {_get_blocked_ips_count()}')
        analysis['recommendations'].append('مراقبة IPs من دول معينة')
    
    if 'login' in query_lower or 'دخول' in query_lower:
        failed = _get_failed_logins_count(24)
        analysis['findings'].append(f'محاولات فاشلة (24h): {failed}')
        if failed > 10:
            analysis['threat_level'] = 'medium'
            analysis['recommendations'].append('تفعيل CAPTCHA أو تقليل rate limit')
    
    if 'user' in query_lower or 'مستخدم' in query_lower:
        analysis['findings'].append(f'إجمالي المستخدمين: {User.query.count()}')
        analysis['findings'].append(f'المستخدمين النشطين: {User.query.filter_by(is_active=True).count()}')
    
    return analysis


def _get_ai_suggestions():
    """اقتراحات ذكية من AI"""
    suggestions = []
    
    # فحص محاولات فاشلة
    failed = _get_failed_logins_count(24)
    if failed > 10:
        suggestions.append({
            'type': 'warning',
            'title': f'محاولات دخول فاشلة كثيرة ({failed})',
            'action': 'تفعيل CAPTCHA أو حظر IPs',
            'priority': 'high'
        })
    
    # فحص مستخدمين غير نشطين
    inactive = User.query.filter_by(is_active=False).count()
    if inactive > 5:
        suggestions.append({
            'type': 'info',
            'title': f'مستخدمين محظورين ({inactive})',
            'action': 'مراجعة المستخدمين المحظورين',
            'priority': 'low'
        })
    
    return suggestions


def _get_all_tables():
    """الحصول على جميع جداول قاعدة البيانات"""
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        return [t for t in inspector.get_table_names()]
    except Exception:
        try:
            uri = current_app.config.get("SQLALCHEMY_DATABASE_URI", "")
            if "postgresql" in uri:
                rows = db.session.execute(text("SELECT tablename FROM pg_catalog.pg_tables WHERE schemaname NOT IN ('pg_catalog','information_schema') ORDER BY tablename")).fetchall()
                return [r[0] for r in rows]
            return []
        except Exception:
            return []


def _browse_table(table_name, limit=100):
    """تصفح جدول معين"""
    try:
        _assert_valid_table(table_name)
        try:
            limit = int(limit)
        except Exception:
            limit = 100
        limit = max(1, min(limit, 10000))
        result = db.session.execute(text(f"SELECT * FROM {table_name} LIMIT {limit}"))
        columns = list(result.keys()) if hasattr(result, "keys") else []
        data = [dict(row) for row in result.mappings().all()]
        return data, columns
    except Exception:
        return [], []


def _get_table_info(table_name):
    """الحصول على معلومات الجدول (الأعمدة والأنواع)"""
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        columns = inspector.get_columns(table_name)
        pk_constraint = inspector.get_pk_constraint(table_name)
        pk_cols = pk_constraint.get('constrained_columns', [])
        
        info = []
        for i, col in enumerate(columns):
            info.append({
                'cid': i,
                'name': col['name'],
                'type': str(col['type']),
                'notnull': 1 if not col['nullable'] else 0,
                'default': str(col['default']) if col['default'] else None,
                'pk': 1 if col['name'] in pk_cols else 0
            })
        return info
    except Exception as e:
        current_app.logger.error(f"Error getting table info: {e}")
        return []


def _decrypt_data(encrypted_data, decrypt_type):
    """فك تشفير البيانات"""
    result = {
        'success': False,
        'decrypted': None,
        'method': decrypt_type,
        'error': None
    }
    
    try:
        if decrypt_type == 'base64':
            import base64
            result['decrypted'] = base64.b64decode(encrypted_data).decode('utf-8')
            result['success'] = True
        
        elif decrypt_type == 'fernet':
            from cryptography.fernet import Fernet
            key = current_app.config.get('CARD_ENC_KEY', '').encode()
            if key:
                f = Fernet(key)
                result['decrypted'] = f.decrypt(encrypted_data.encode()).decode('utf-8')
                result['success'] = True
            else:
                result['error'] = 'CARD_ENC_KEY غير موجود'
        
        elif decrypt_type == 'auto':
            # محاولة جميع الطرق
            for method in ['base64', 'fernet']:
                try:
                    temp_result = _decrypt_data(encrypted_data, method)
                    if temp_result['success']:
                        result = temp_result
                        result['method'] = f'auto ({method})'
                        break
                except Exception:
                    continue
    
    except Exception as e:
        result['error'] = str(e)
    
    return result


def _analyze_user_behavior():
    """تحليل سلوك المستخدمين"""
    return {
        'most_active': _get_most_active_users(5),
        'login_patterns': _analyze_login_patterns(),
        'suspicious_users': _detect_suspicious_users()
    }


def _detect_security_patterns():
    """كشف أنماط أمنية"""
    return {
        'failed_login_ips': _get_top_failed_ips(10),
        'attack_patterns': _detect_attack_patterns(),
        'time_patterns': _analyze_time_patterns()
    }


def _detect_anomalies():
    """كشف الشذوذات"""
    anomalies = []
    
    # محاولات دخول غير عادية
    failed_count = _get_failed_logins_count(1)  # آخر ساعة
    if failed_count > 5:
        anomalies.append({
            'type': 'login_spike',
            'severity': 'high',
            'description': f'محاولات دخول فاشلة غير عادية: {failed_count} في الساعة الأخيرة'
        })
    
    return anomalies


def _ai_recommendations():
    """توصيات ذكية"""
    recommendations = []
    
    # توصيات بناءً على التحليل
    failed = _get_failed_logins_count(24)
    if failed > 20:
        recommendations.append('تفعيل 2FA للمستخدمين')
        recommendations.append('تقليل rate limit على /login')
    
    return recommendations


def _calculate_threat_level():
    """حساب مستوى التهديد"""
    score = 0
    
    # محاولات فاشلة
    failed = _get_failed_logins_count(24)
    score += min(failed, 50)
    
    # مستخدمين محظورين
    blocked = User.query.filter_by(is_active=False).count()
    score += blocked * 2
    
    if score < 10:
        return {'level': 'low', 'color': 'success', 'label': 'منخفض'}
    elif score < 30:
        return {'level': 'medium', 'color': 'warning', 'label': 'متوسط'}
    else:
        return {'level': 'high', 'color': 'danger', 'label': 'عالي'}


def _detect_suspicious_patterns():
    """كشف الأنماط المشبوهة"""
    patterns = []
    
    # IPs مع محاولات فاشلة متعددة
    suspicious_ips = _get_top_failed_ips(10)
    for ip_data in suspicious_ips:
        if ip_data['count'] > 5:
            patterns.append({
                'type': 'suspicious_ip',
                'ip': ip_data['ip'],
                'count': ip_data['count'],
                'severity': 'high' if ip_data['count'] > 10 else 'medium'
            })
    
    return patterns


def _get_most_active_users(limit=5):
    """المستخدمين الأكثر نشاطاً"""
    return User.query.filter_by(is_active=True).order_by(
        User.login_count.desc()
    ).limit(limit).all()


def _analyze_login_patterns():
    """تحليل أنماط تسجيل الدخول"""
    # تحليل الأوقات
    return {'peak_hours': [9, 10, 11, 14, 15], 'off_hours': [0, 1, 2, 3, 4, 5]}


def _detect_suspicious_users():
    """كشف المستخدمين المشبوهين"""
    suspicious = []
    
    # مستخدمين مع محاولات فاشلة كثيرة
    users_with_fails = AuditLog.query.filter(
        AuditLog.action == 'login.failed',
        AuditLog.created_at >= datetime.now(timezone.utc) - timedelta(days=7)
    ).all()
    
    return suspicious


def _get_top_failed_ips(limit=10):
    """أكثر IPs مع محاولات فاشلة"""
    failed_ips = {}
    
    logs = AuditLog.query.filter(
        AuditLog.action.in_(['login.failed', 'login.blocked']),
        AuditLog.created_at >= datetime.now(timezone.utc) - timedelta(hours=24)
    ).all()
    
    for log in logs:
        ip = log.ip_address
        if ip:
            failed_ips[ip] = failed_ips.get(ip, 0) + 1
    
    sorted_ips = sorted(failed_ips.items(), key=lambda x: x[1], reverse=True)
    return [{'ip': ip, 'count': count} for ip, count in sorted_ips[:limit]]


def _detect_attack_patterns():
    """كشف أنماط الهجوم"""
    return ['brute_force', 'sql_injection_attempt', 'xss_attempt']


def _analyze_time_patterns():
    """تحليل أنماط الوقت"""
    return {'suspicious_hours': [2, 3, 4], 'normal_hours': [9, 10, 11, 14, 15]}




def _kill_all_user_sessions():
    """إنهاء جميع جلسات المستخدمين"""
    # تحديث last_seen لجميع المستخدمين
    User.query.update({'last_seen': datetime.now(timezone.utc) - timedelta(days=30)})
    db.session.commit()


def _get_active_users():
    """الحصول على المستخدمين النشطين"""
    threshold = datetime.now(timezone.utc) - timedelta(minutes=5)
    all_users = User.query.filter(User.last_seen.isnot(None)).all()
    return [u for u in all_users if make_aware(u.last_seen) >= threshold]


def _get_users_online():
    """عدد المستخدمين المتصلين"""
    threshold = datetime.now(timezone.utc) - timedelta(minutes=5)
    all_users = User.query.filter(User.last_seen.isnot(None)).all()
    return sum(1 for u in all_users if make_aware(u.last_seen) >= threshold)


def _get_system_setting(key, default=None):
    """توجيه لدالة get_system_setting من ai_service"""
    return get_system_setting(key, default)


def _get_recent_actions(limit=50):
    """آخر الإجراءات"""
    return AuditLog.query.order_by(AuditLog.created_at.desc()).limit(limit).all()


def _get_live_metrics():
    """مقاييس حية"""
    import psutil
    return {
        'cpu': psutil.cpu_percent(interval=1),
        'memory': psutil.virtual_memory().percent,
        'disk': psutil.disk_usage('/').percent,
    }


def _set_system_setting(key, value):
    """حفظ إعداد نظام"""
    from models import SystemSettings
    setting = SystemSettings.query.filter_by(key=key).first()
    if setting:
        setting.value = str(value)
    else:
        setting = SystemSettings(key=key, value=str(value))
        db.session.add(setting)
    db.session.commit()


def _get_db_size():
    """حجم قاعدة البيانات"""
    try:
        res = db.session.execute(text("SELECT pg_database_size(current_database())"))
        size_bytes = int(res.scalar() or 0)
        if size_bytes <= 0:
            return "N/A"
        return f"{size_bytes / (1024*1024):.2f} MB"
    except Exception:
        return "N/A"


def _get_system_health():
    """صحة النظام"""
    try:
        # فحص قاعدة البيانات
        db.session.execute(text("SELECT 1"))
        return "ممتاز"
    except Exception:
        return "خطأ"


def _get_active_sessions_count():
    """عدد الجلسات النشطة"""
    threshold = datetime.now(timezone.utc) - timedelta(hours=24)
    all_users = User.query.filter(User.last_login.isnot(None)).all()
    return sum(1 for u in all_users if make_aware(u.last_login) >= threshold)


def _get_online_users_detailed():
    """تفاصيل المستخدمين المتصلين"""
    threshold = datetime.now(timezone.utc) - timedelta(minutes=5)
    all_users = User.query.filter(User.last_seen.isnot(None)).all()
    return [u for u in all_users if make_aware(u.last_seen) >= threshold]


def _get_available_backups():
    """قائمة النسخ الاحتياطية"""
    import os
    backup_dir = 'instance/backups'
    backups = []
    
    if os.path.exists(backup_dir):
        for f in os.listdir(backup_dir):
            if f.endswith('.dump'):
                full_path = os.path.join(backup_dir, f)
                backups.append({
                    'name': f,
                    'size': f"{os.path.getsize(full_path) / (1024*1024):.2f} MB",
                    'date': datetime.fromtimestamp(os.path.getmtime(full_path))
                })
    
    return sorted(backups, key=lambda x: x['date'], reverse=True)


def _get_slow_queries():
    """استعلامات بطيئة"""
    # محاكاة - في الواقع تحتاج لـ query profiling
    return []


def _get_avg_response_times():
    """متوسط أوقات الاستجابة"""
    return {'avg': '120ms', 'min': '50ms', 'max': '500ms'}


def _get_memory_usage():
    """استخدام الذاكرة"""
    import psutil
    return psutil.virtual_memory().percent


def _get_cpu_usage():
    """استخدام المعالج"""
    import psutil
    return psutil.cpu_percent(interval=1)


def _safe_count_table(table_name):
    """عد صفوف جدول بشكل آمن"""
    try:
        result = db.session.execute(text(f"SELECT COUNT(*) FROM {table_name}"))
        return result.scalar()
    except Exception:
        return 0


def _get_available_log_files():
    """الحصول على ملفات اللوج المتاحة"""
    import os
    
    log_files = []
    
    files = {
        'error': 'logs/error.log',
        'server': 'logs/server_error.log',
        'audit': 'instance/audit.log',
        'access': 'logs/access.log',
        'security': 'logs/security.log',
        'performance': 'logs/performance.log',
    }
    
    for log_type, log_path in files.items():
        if os.path.exists(log_path):
            size = os.path.getsize(log_path)
            log_files.append({
                'type': log_type,
                'path': log_path,
                'size': f"{size / 1024:.2f} KB",
                'modified': datetime.fromtimestamp(os.path.getmtime(log_path))
            })
    
    return log_files


def _test_integration_connection(integration_type):
    """اختبار اتصال التكامل"""
    from models import SystemSettings
    import requests
    import smtplib
    from email.mime.text import MIMEText
    
    try:
        if integration_type == 'whatsapp':
            phone = SystemSettings.query.filter_by(key='whatsapp_phone').first()
            token = SystemSettings.query.filter_by(key='whatsapp_token').first()
            url = SystemSettings.query.filter_by(key='whatsapp_url').first()
            
            if not all([phone, token, url]):
                return {'success': False, 'error': 'بيانات واتساب ناقصة'}
            
            # اختبار API واتساب
            test_url = f"{url.value}/status"
            headers = {'Authorization': f'Bearer {token.value}'}
            response = requests.get(test_url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                return {'success': True, 'message': 'واتساب متصل بنجاح'}
            else:
                return {'success': False, 'error': f'خطأ واتساب: {response.status_code}'}
        
        elif integration_type == 'email':
            server = SystemSettings.query.filter_by(key='smtp_server').first()
            port = SystemSettings.query.filter_by(key='smtp_port').first()
            username = SystemSettings.query.filter_by(key='smtp_username').first()
            password = SystemSettings.query.filter_by(key='smtp_password').first()
            
            if not all([server, port, username, password]):
                return {'success': False, 'error': 'بيانات البريد ناقصة'}
            
            # اختبار SMTP
            smtp = smtplib.SMTP(server.value, int(port.value))
            smtp.starttls()
            smtp.login(username.value, password.value)
            smtp.quit()
            
            return {'success': True, 'message': 'البريد الإلكتروني متصل بنجاح'}
        
        elif integration_type == 'api_keys':
            openai_key = SystemSettings.query.filter_by(key='openai_key').first()
            google_maps_key = SystemSettings.query.filter_by(key='google_maps_key').first()
            
            if openai_key and openai_key.value:
                # اختبار OpenAI
                headers = {'Authorization': f'Bearer {openai_key.value}'}
                response = requests.get('https://api.openai.com/v1/models', headers=headers, timeout=10)
                if response.status_code != 200:
                    return {'success': False, 'error': 'مفتاح OpenAI غير صالح'}
            
            if google_maps_key and google_maps_key.value:
                # اختبار Google Maps
                test_url = f"https://maps.googleapis.com/maps/api/geocode/json?address=test&key={google_maps_key.value}"
                response = requests.get(test_url, timeout=10)
                if response.status_code != 200:
                    return {'success': False, 'error': 'مفتاح Google Maps غير صالح'}
            
            return {'success': True, 'message': 'مفاتيح API صحيحة'}
        
        else:
            return {'success': True, 'message': 'التكامل محفوظ'}
    
    except requests.exceptions.RequestException as e:
        return {'success': False, 'error': f'خطأ في الشبكة: {str(e)}'}
    except smtplib.SMTPException as e:
        return {'success': False, 'error': f'خطأ في البريد: {str(e)}'}
    except Exception as e:
        return {'success': False, 'error': f'خطأ عام: {str(e)}'}


def _send_test_message(integration_type):
    """إرسال رسالة تجريبية"""
    from models import SystemSettings
    import requests
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    try:
        if integration_type == 'whatsapp':
            phone = SystemSettings.query.filter_by(key='whatsapp_phone').first()
            token = SystemSettings.query.filter_by(key='whatsapp_token').first()
            url = SystemSettings.query.filter_by(key='whatsapp_url').first()
            
            if not all([phone, token, url]):
                return {'success': False, 'error': 'بيانات واتساب ناقصة'}
            
            # إرسال رسالة تجريبية
            message_data = {
                'to': phone.value,
                'message': '🧪 رسالة تجريبية من نظام إدارة الكراج - التكامل يعمل بنجاح! ✅'
            }
            
            headers = {'Authorization': f'Bearer {token.value}', 'Content-Type': 'application/json'}
            response = requests.post(f"{url.value}/send", json=message_data, headers=headers, timeout=10)
            
            if response.status_code == 200:
                return {'success': True, 'message': 'تم إرسال رسالة واتساب تجريبية'}
            else:
                return {'success': False, 'error': f'فشل الإرسال: {response.status_code}'}
        
        elif integration_type == 'email':
            server = SystemSettings.query.filter_by(key='smtp_server').first()
            port = SystemSettings.query.filter_by(key='smtp_port').first()
            username = SystemSettings.query.filter_by(key='smtp_username').first()
            password = SystemSettings.query.filter_by(key='smtp_password').first()
            
            if not all([server, port, username, password]):
                return {'success': False, 'error': 'بيانات البريد ناقصة'}
            
            # إرسال بريد تجريبي
            msg = MIMEMultipart()
            msg['From'] = username.value
            msg['To'] = username.value  # إرسال لنفسه
            msg['Subject'] = '🧪 رسالة تجريبية - نظام إدارة الكراج'
            
            body = '''
            <h2>🧪 رسالة تجريبية</h2>
            <p>هذه رسالة تجريبية من نظام إدارة الكراج</p>
            <p><strong>التكامل يعمل بنجاح! ✅</strong></p>
            <p>الوقت: {}</p>
            '''.format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            msg.attach(MIMEText(body, 'html'))
            
            smtp = smtplib.SMTP(server.value, int(port.value))
            smtp.starttls()
            smtp.login(username.value, password.value)
            smtp.send_message(msg)
            smtp.quit()
            
            return {'success': True, 'message': 'تم إرسال بريد تجريبي'}
        
        else:
            return {'success': False, 'error': 'نوع التكامل غير مدعوم للإرسال'}
    
    except Exception as e:
        return {'success': False, 'error': f'خطأ في الإرسال: {str(e)}'}


def _get_integration_stats():
    """إحصائيات التكاملات الحقيقية"""
    from models import SystemSettings
    
    # فحص التكوين الحقيقي
    whatsapp_configured = bool(SystemSettings.query.filter_by(key='whatsapp_token').first())
    email_configured = bool(SystemSettings.query.filter_by(key='smtp_server').first())
    api_configured = bool(SystemSettings.query.filter_by(key='openai_key').first())
    
    # إحصائيات حقيقية من قاعدة البيانات
    stats = {
        'whatsapp': {
            'configured': whatsapp_configured,
            'last_test': _get_last_integration_activity('whatsapp'),
            'messages_sent': _count_integration_usage('whatsapp'),
            'status': 'active' if whatsapp_configured else 'inactive'
        },
        'email': {
            'configured': email_configured,
            'last_test': _get_last_integration_activity('email'),
            'emails_sent': _count_integration_usage('email'),
            'status': 'active' if email_configured else 'inactive'
        },
        'api_keys': {
            'configured': api_configured,
            'last_test': _get_last_integration_activity('api'),
            'requests_made': _count_integration_usage('api'),
            'status': 'active' if api_configured else 'inactive'
        }
    }
    
    return stats


def _get_last_integration_activity(integration_type):
    """الحصول على آخر نشاط للتكامل"""
    try:
        # البحث في سجل التدقيق
        from models import AuditLog
        last_activity = AuditLog.query.filter(
            AuditLog.action.like(f'%{integration_type}%')
        ).order_by(AuditLog.timestamp.desc()).first()
        
        if last_activity:
            return last_activity.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        else:
            return 'لم يتم الاستخدام بعد'
    except Exception:
        return 'غير متاح'


def _count_integration_usage(integration_type):
    """عد استخدام التكامل"""
    try:
        from models import AuditLog
        count = AuditLog.query.filter(
            AuditLog.action.like(f'%{integration_type}%')
        ).count()
        return count
    except Exception:
        return 0


def _log_integration_activity(integration_type, action, success):
    """تسجيل نشاط التكامل"""
    try:
        from models import AuditLog
        from flask_login import current_user
        
        activity = AuditLog(
            user_id=current_user.id,
            action=f'{integration_type}_{action}',
            details=f'Integration {action}: {integration_type} - Success: {success}',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', ''),
            timestamp=datetime.now(timezone.utc)
        )
        
        db.session.add(activity)
        db.session.commit()
    except Exception as e:
        current_app.logger.error(f"Error logging integration activity: {e}")


def _get_recent_errors(limit=100):
    """الحصول على آخر الأخطاء"""
    import os
    
    errors = []
    
    if os.path.exists('error.log'):
        try:
            with open('error.log', 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
                # آخر الأخطاء
                for line in lines[-limit:]:
                    if line.strip():
                        errors.append({
                            'message': line.strip(),
                            'timestamp': datetime.now(timezone.utc)
                        })
        except Exception:
            pass
    
    return errors


def _get_error_statistics():
    """إحصائيات الأخطاء"""
    import os
    
    stats = {
        'total_errors': 0,
        'today_errors': 0,
        'critical_errors': 0,
        'warning_errors': 0,
    }
    
    if os.path.exists('error.log'):
        try:
            with open('error.log', 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
                stats['total_errors'] = len(lines)
                
                # تحليل بسيط
                for line in lines:
                    if 'CRITICAL' in line or 'ERROR' in line:
                        stats['critical_errors'] += 1
                    elif 'WARNING' in line:
                        stats['warning_errors'] += 1
        except Exception:
            pass
    
    return stats


def _get_security_notifications():
    """الحصول على الإشعارات الأمنية"""
    notifications = []
    
    # فحص محاولات فاشلة
    failed = _get_failed_logins_count(1)
    if failed > 5:
        notifications.append({
            'severity': 'danger',
            'icon': 'exclamation-triangle',
            'title': 'محاولات دخول فاشلة',
            'message': f'{failed} محاولة فاشلة في الساعة الأخيرة',
            'timestamp': datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')
        })
    
    # فحص وضع الصيانة
    if _get_system_setting('maintenance_mode', False):
        notifications.append({
            'severity': 'warning',
            'icon': 'tools',
            'title': 'وضع الصيانة مفعل',
            'message': 'النظام في وضع الصيانة - المستخدمون لا يمكنهم الدخول',
            'timestamp': datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')
        })
    
    return notifications


@security_bp.route('/monitoring-dashboard')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def monitoring_dashboard():
    """لوحة مراقبة الأداء الشاملة (Grafana-like)"""
    from models import User, Sale, Payment, ServiceRequest
    
    dashboard_data = {
        'active_users': User.query.filter(User.last_seen >= datetime.now(timezone.utc) - timedelta(minutes=30)).count(),
        'total_sales_today': Sale.query.filter(Sale.created_at >= datetime.now(timezone.utc).replace(hour=0, minute=0, second=0)).count(),
        'total_payments_today': Payment.query.filter(Payment.created_at >= datetime.now(timezone.utc).replace(hour=0, minute=0, second=0)).count(),
        'pending_services': ServiceRequest.query.filter_by(status='pending').count(),
    }
    
    try:
        import psutil
        system_metrics = {
            'cpu': round(psutil.cpu_percent(interval=1), 2),
            'memory': round(psutil.virtual_memory().percent, 2),
            'disk': round(psutil.disk_usage('/').percent, 2),
        }
    except Exception:
        system_metrics = {'cpu': 0, 'memory': 0, 'disk': 0}
    
    return render_template('security/monitoring_dashboard.html',
                         title='لوحة المراقبة الشاملة',
                         dashboard_data=dashboard_data,
                         system_metrics=system_metrics)


@security_bp.route('/dark-mode-settings', methods=['GET', 'POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def dark_mode_settings():
    """إعدادات الوضع الليلي (Dark Mode)"""
    if request.method == 'POST':
        # حفظ الإعدادات
        flash('✅ تم حفظ إعدادات الوضع الليلي', 'success')
        return redirect(url_for('security.dark_mode_settings'))
    
    return render_template('security/dark_mode_settings.html',
                         title='إعدادات الوضع الليلي')


@security_bp.route('/grafana-setup')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def grafana_setup():
    """إعداد وتثبيت Grafana + Prometheus"""
    return render_template('security/grafana_setup.html',
                         title='إعداد Grafana + Prometheus')


@security_bp.route('/prometheus-metrics')
def prometheus_metrics():
    """Prometheus metrics endpoint"""
    try:
        from services.prometheus_service import get_all_metrics
        return get_all_metrics()
    except Exception:
        ts = int(datetime.now(timezone.utc).timestamp())
        db_up = 0
        try:
            db.session.execute(text("SELECT 1")).scalar()
            db_up = 1
        except Exception:
            db_up = 0

        body = "\n".join(
            [
                "# HELP garage_manager_up Application up indicator",
                "# TYPE garage_manager_up gauge",
                "garage_manager_up 1",
                "# HELP garage_manager_db_up Database connectivity indicator",
                "# TYPE garage_manager_db_up gauge",
                f"garage_manager_db_up {db_up}",
                "# HELP garage_manager_timestamp_seconds Current server timestamp",
                "# TYPE garage_manager_timestamp_seconds gauge",
                f"garage_manager_timestamp_seconds {ts}",
                "",
            ]
        )
        return body, 200, {"Content-Type": "text/plain; version=0.0.4; charset=utf-8"}


@security_bp.route('/api/live-metrics')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_live_metrics():
    """API للحصول على المتريكات الحية"""
    from services.prometheus_service import get_live_metrics_json
    return jsonify(get_live_metrics_json())


@security_bp.route('/api/indexes/create', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_create_index():
    """إنشاء فهرس جديد"""
    try:
        data = request.get_json()
        table_name = data.get('table')
        index_name = data.get('index_name')
        columns = data.get('columns')
        unique = data.get('unique', False)
        method = str(data.get('method') or data.get('index_method') or '').strip().upper() or None
        
        if not all([table_name, index_name, columns]):
            return jsonify({'success': False, 'message': 'بيانات ناقصة'}), 400
        
        if isinstance(columns, str):
            columns = [columns]
        
        _assert_valid_table(table_name)
        if not _is_safe_identifier(index_name):
            return jsonify({'success': False, 'message': 'اسم فهرس غير صالح'}), 400
        allowed_columns = set(_get_table_columns_safe(table_name))

        safe_cols_parts: list[str] = []
        if isinstance(columns, list) and columns and isinstance(columns[0], dict):
            for item in columns:
                col = (item or {}).get('name')
                opclass = (item or {}).get('opclass')
                if not col or not _is_safe_identifier(col) or col not in allowed_columns:
                    continue
                if opclass:
                    opclass = str(opclass).strip()
                    if not _is_safe_qualified_identifier(opclass):
                        continue
                    safe_cols_parts.append(f"{col} {opclass}")
                else:
                    safe_cols_parts.append(col)
        else:
            safe_cols_parts = [c for c in columns if _is_safe_identifier(c) and c in allowed_columns]

        if not safe_cols_parts:
            return jsonify({'success': False, 'message': 'أعمدة غير صالحة'}), 400

        unique_str = "UNIQUE" if unique else ""
        cols_str = ", ".join(safe_cols_parts)

        if method:
            allowed_pg_methods = {"BTREE", "GIN", "GIST", "HASH", "BRIN", "SPGIST"}
            if not _is_postgresql():
                return jsonify({'success': False, 'message': 'نوع الفهرس مدعوم فقط على PostgreSQL'}), 400
            if method not in allowed_pg_methods:
                return jsonify({'success': False, 'message': 'نوع فهرس غير مدعوم'}), 400
            if unique and method != "BTREE":
                return jsonify({'success': False, 'message': 'UNIQUE مدعوم فقط مع B-TREE'}), 400
            if method == "BTREE":
                sql = f"CREATE {unique_str} INDEX {index_name} ON {table_name} ({cols_str})"
            else:
                sql = f"CREATE {unique_str} INDEX {index_name} ON {table_name} USING {method} ({cols_str})"
        else:
            sql = f"CREATE {unique_str} INDEX {index_name} ON {table_name} ({cols_str})"

        db.session.execute(text(sql))
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'✅ تم إنشاء الفهرس {index_name} بنجاح'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/indexes/drop', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_drop_index():
    """حذف فهرس"""
    try:
        data = request.get_json()
        index_name = data.get('index_name')
        table_name = data.get('table')
        
        if not index_name:
            return jsonify({'success': False, 'message': 'اسم الفهرس مطلوب'}), 400
        
        if not _is_safe_identifier(index_name):
            return jsonify({'success': False, 'message': 'اسم فهرس غير صالح'}), 400
        db.session.execute(text(f"DROP INDEX IF EXISTS {index_name}"))
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'✅ تم حذف الفهرس {index_name} بنجاح'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/indexes/auto-optimize', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_auto_optimize_indexes():
    """تحسين تلقائي للفهارس"""
    try:
        from sqlalchemy import inspect
        
        inspector = inspect(db.engine)
        tables = inspector.get_table_names()
        
        created_indexes = []
        skipped_indexes = []
        
        optimization_rules = {
            'customers': ['name', 'phone', 'email', 'is_active', 'created_at'],
            'suppliers': ['name', 'phone', 'created_at'],
            'partners': ['name', 'phone_number', 'created_at'],
            'products': ['name', 'barcode', 'sku', 'category_id', 'is_active', 'created_at'],
            'sales': ['customer_id', 'seller_id', 'status', 'sale_date', 'created_at', 'payment_status'],
            'sale_lines': ['sale_id', 'product_id', 'warehouse_id'],
            'payments': ['entity_type', 'customer_id', 'supplier_id', 'partner_id', 'status', 'direction', 'payment_date', 'receipt_number'],
            'service_requests': ['customer_id', 'status', 'priority', 'created_at', 'service_number'],
            'shipments': ['destination_id', 'status', 'shipment_date', 'created_at'],
            'shipment_items': ['shipment_id', 'product_id'],
            'invoices': ['customer_id', 'status', 'invoice_number', 'invoice_date', 'due_date', 'source'],
            'expenses': ['type_id', 'employee_id', 'date', 'created_at'],
            'stock_levels': ['product_id', 'warehouse_id'],
            'audit_logs': ['user_id', 'action', 'model_name', 'record_id', 'created_at'],
            'checks': ['customer_id', 'supplier_id', 'partner_id', 'check_number', 'check_date', 'check_due_date', 'status'],
            'users': ['username', 'email', 'is_active', 'role_id'],
            'warehouses': ['name', 'warehouse_type', 'is_active'],
            'notes': ['entity_type', 'entity_id', 'author_id', 'created_at']
        }
        
        for table, columns_to_index in optimization_rules.items():
            if table not in tables:
                continue
            
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=SAWarning, message="Skipped unsupported reflection*")
                existing_indexes = inspector.get_indexes(table)
            existing_index_names = {idx['name'] for idx in existing_indexes}
            
            for column in columns_to_index:
                index_name = f"ix_{table}_{column}"
                
                if index_name in existing_index_names:
                    skipped_indexes.append(index_name)
                    continue
                
                table_columns = inspector.get_columns(table)
                column_names = [c['name'] for c in table_columns]
                
                if column not in column_names:
                    continue
                
                try:
                    sql = f"CREATE INDEX {index_name} ON {table} ({column})"
                    db.session.execute(text(sql))
                    db.session.commit()
                    created_indexes.append(index_name)
                except Exception:
                    db.session.rollback()
        
        composite_indexes = [
            ('sales', ['customer_id', 'sale_date'], 'ix_sales_customer_date'),
            ('sales', ['status', 'sale_date'], 'ix_sales_status_date'),
            ('payments', ['customer_id', 'payment_date'], 'ix_payments_customer_date'),
            ('service_requests', ['customer_id', 'status'], 'ix_service_requests_customer_status'),
            ('service_requests', ['status', 'created_at'], 'ix_service_requests_status_date'),
            ('audit_logs', ['user_id', 'created_at'], 'ix_audit_logs_user_date'),
            ('stock_levels', ['product_id', 'warehouse_id'], 'ix_stock_levels_product_warehouse'),
        ]
        
        for table, columns, index_name in composite_indexes:
            if table not in tables:
                continue
            
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=SAWarning, message="Skipped unsupported reflection*")
                existing_indexes = inspector.get_indexes(table)
            existing_index_names = {idx['name'] for idx in existing_indexes}
            
            if index_name in existing_index_names:
                skipped_indexes.append(index_name)
                continue
            
            try:
                cols_str = ", ".join(columns)
                unique_str = "UNIQUE" if 'product_warehouse' in index_name else ""
                sql = f"CREATE {unique_str} INDEX {index_name} ON {table} ({cols_str})"
                db.session.execute(text(sql))
                db.session.commit()
                created_indexes.append(index_name)
            except Exception:
                db.session.rollback()
        
        return jsonify({
            'success': True,
            'message': f'✅ تم إنشاء {len(created_indexes)} فهرس جديد',
            'created': created_indexes,
            'skipped': len(skipped_indexes),
            'total': len(created_indexes) + len(skipped_indexes)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/indexes/clean-and-rebuild', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_clean_rebuild_indexes():
    """تنظيف وإعادة بناء الفهارس"""
    try:
        from sqlalchemy import inspect
        
        inspector = inspect(db.engine)
        tables = inspector.get_table_names()
        
        dropped_count = 0
        created_count = 0
        
        for table in tables:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=SAWarning, message="Skipped unsupported reflection*")
                indexes = inspector.get_indexes(table)
            
            for idx in indexes:
                if idx['name'] and idx['name'].startswith('ix_'):
                    try:
                        db.session.execute(text(f"DROP INDEX {idx['name']}"))
                        db.session.commit()
                        dropped_count += 1
                    except Exception:
                        db.session.rollback()
        
        optimization_rules = {
            'customers': ['name', 'phone', 'email', 'is_active', 'created_at'],
            'suppliers': ['name', 'phone', 'created_at'],
            'partners': ['name', 'phone_number', 'created_at'],
            'products': ['name', 'barcode', 'sku', 'category_id', 'is_active', 'created_at'],
            'sales': ['customer_id', 'seller_id', 'status', 'sale_date', 'created_at', 'payment_status'],
            'sale_lines': ['sale_id', 'product_id', 'warehouse_id'],
            'payments': ['entity_type', 'customer_id', 'supplier_id', 'partner_id', 'status', 'direction', 'payment_date', 'receipt_number'],
            'service_requests': ['customer_id', 'status', 'priority', 'created_at', 'service_number'],
            'shipments': ['destination_id', 'status', 'shipment_date', 'created_at'],
            'shipment_items': ['shipment_id', 'product_id'],
            'invoices': ['customer_id', 'status', 'invoice_number', 'invoice_date', 'due_date', 'source'],
            'expenses': ['type_id', 'employee_id', 'date', 'created_at'],
            'stock_levels': ['product_id', 'warehouse_id'],
            'audit_logs': ['user_id', 'action', 'model_name', 'record_id', 'created_at'],
            'checks': ['customer_id', 'supplier_id', 'partner_id', 'check_number', 'check_date', 'check_due_date', 'status'],
            'users': ['username', 'email', 'is_active', 'role_id'],
            'warehouses': ['name', 'warehouse_type', 'is_active'],
            'notes': ['entity_type', 'entity_id', 'author_id', 'created_at']
        }
        
        for table, columns_to_index in optimization_rules.items():
            if table not in tables:
                continue
            
            table_columns = inspector.get_columns(table)
            column_names = [c['name'] for c in table_columns]
            
            for column in columns_to_index:
                if column not in column_names:
                    continue
                
                index_name = f"ix_{table}_{column}"
                try:
                    sql = f"CREATE INDEX {index_name} ON {table} ({column})"
                    db.session.execute(text(sql))
                    db.session.commit()
                    created_count += 1
                except Exception:
                    db.session.rollback()
        
        composite_indexes = [
            ('sales', ['customer_id', 'sale_date'], 'ix_sales_customer_date'),
            ('sales', ['status', 'sale_date'], 'ix_sales_status_date'),
            ('payments', ['customer_id', 'payment_date'], 'ix_payments_customer_date'),
            ('service_requests', ['customer_id', 'status'], 'ix_service_requests_customer_status'),
            ('service_requests', ['status', 'created_at'], 'ix_service_requests_status_date'),
            ('audit_logs', ['user_id', 'created_at'], 'ix_audit_logs_user_date'),
            ('stock_levels', ['product_id', 'warehouse_id'], 'ix_stock_levels_product_warehouse'),
        ]
        
        for table, columns, index_name in composite_indexes:
            if table not in tables:
                continue
            
            try:
                cols_str = ", ".join(columns)
                unique_str = "UNIQUE" if 'product_warehouse' in index_name else ""
                sql = f"CREATE {unique_str} INDEX {index_name} ON {table} ({cols_str})"
                db.session.execute(text(sql))
                db.session.commit()
                created_count += 1
            except Exception:
                db.session.rollback()
        
        return jsonify({
            'success': True,
            'message': f'✅ تم حذف {dropped_count} فهرس وإنشاء {created_count} فهرس جديد',
            'dropped': dropped_count,
            'created': created_count
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/indexes/analyze-table', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_analyze_table():
    """تحليل جدول واقتراح فهارس"""
    try:
        from sqlalchemy import inspect
        
        data = request.get_json()
        table_name = data.get('table')
        
        if not table_name:
            return jsonify({'success': False, 'message': 'اسم الجدول مطلوب'}), 400
        
        inspector = inspect(db.engine)
        
        if table_name not in inspector.get_table_names():
            return jsonify({'success': False, 'message': 'الجدول غير موجود'}), 404
        
        columns = inspector.get_columns(table_name)
        indexes = inspector.get_indexes(table_name)
        foreign_keys = inspector.get_foreign_keys(table_name)
        
        indexed_columns = set()
        for idx in indexes:
            indexed_columns.update(idx['column_names'])
        
        suggestions = []
        
        for col in columns:
            col_name = col['name']
            col_type = str(col['type'])
            
            if col_name in indexed_columns:
                continue
            
            priority = 'low'
            reason = ''
            
            if col_name.endswith('_id'):
                priority = 'high'
                reason = 'Foreign Key - يسرع عمليات JOIN'
            elif 'status' in col_name.lower():
                priority = 'high'
                reason = 'حقل حالة - يستخدم كثيراً في الفلترة'
            elif 'date' in col_name.lower() or 'time' in col_name.lower():
                priority = 'medium'
                reason = 'حقل تاريخ - يستخدم في الفرز والفلترة'
            elif col_name in ['name', 'email', 'phone', 'username']:
                priority = 'high'
                reason = 'حقل بحث رئيسي'
            elif 'number' in col_name.lower():
                priority = 'medium'
                reason = 'حقل رقمي - قد يستخدم في البحث'
            elif col_name.startswith('is_'):
                priority = 'low'
                reason = 'حقل boolean - قد يفيد في الفلترة'
            
            if priority != 'low' or len(suggestions) < 20:
                suggestions.append({
                    'column': col_name,
                    'type': col_type,
                    'priority': priority,
                    'reason': reason,
                    'index_name': f"ix_{table_name}_{col_name}"
                })
        
        suggestions.sort(key=lambda x: {'high': 0, 'medium': 1, 'low': 2}[x['priority']])
        
        return jsonify({
            'success': True,
            'table': table_name,
            'total_columns': len(columns),
            'indexed_columns': len(indexed_columns),
            'suggestions': suggestions[:15],
            'foreign_keys': [fk['constrained_columns'] for fk in foreign_keys]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/indexes/batch-create', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_batch_create_indexes():
    """إنشاء عدة فهارس دفعة واحدة"""
    try:
        data = request.get_json()
        indexes = data.get('indexes', [])
        
        if not indexes:
            return jsonify({'success': False, 'message': 'لا توجد فهارس للإنشاء'}), 400
        
        created = []
        failed = []
        
        for idx in indexes:
            table_name = idx.get('table')
            index_name = idx.get('index_name')
            columns = idx.get('columns')
            unique = idx.get('unique', False)
            
            if not all([table_name, index_name, columns]):
                failed.append({'index': index_name, 'reason': 'بيانات ناقصة'})
                continue
            
            if isinstance(columns, str):
                columns = [columns]
            
            try:
                _assert_valid_table(table_name)
                if not _is_safe_identifier(index_name):
                    raise ValueError('اسم فهرس غير صالح')
                allowed_columns = set(_get_table_columns_safe(table_name))
                safe_cols = [c for c in columns if _is_safe_identifier(c) and c in allowed_columns]
                if not safe_cols:
                    raise ValueError('أعمدة غير صالحة')
                unique_str = "UNIQUE" if unique else ""
                cols_str = ", ".join(safe_cols)
                sql = f"CREATE {unique_str} INDEX {index_name} ON {table_name} ({cols_str})"
                db.session.execute(text(sql))
                db.session.commit()
                created.append(index_name)
            except Exception as e:
                db.session.rollback()
                failed.append({'index': index_name, 'reason': str(e)})
        
        return jsonify({
            'success': True,
            'message': f'✅ تم إنشاء {len(created)} فهرس من أصل {len(indexes)}',
            'created': created,
            'failed': failed
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/maintenance/vacuum', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_maintenance_vacuum():
    """تنفيذ VACUUM على قاعدة البيانات"""
    try:
        dialect = getattr(db.engine, "dialect", None)
        dialect_name = (getattr(dialect, "name", "") or "").lower()
        if dialect_name != "postgresql":
            return jsonify({
                'success': False,
                'message': 'هذه العملية مدعومة فقط على PostgreSQL'
            }), 400

        try:
            db.session.rollback()
        except Exception:
            pass

        with db.engine.connect().execution_options(isolation_level="AUTOCOMMIT") as conn:
            conn.execute(text("VACUUM"))
        return jsonify({
            'success': True,
            'message': '✅ تم تنفيذ VACUUM بنجاح - تم تنظيف قاعدة البيانات'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/maintenance/analyze', methods=['POST'])
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def api_maintenance_analyze():
    """تنفيذ ANALYZE على جميع الجداول"""
    try:
        db.session.execute(text('ANALYZE'))
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '✅ تم تنفيذ ANALYZE بنجاح - تم تحليل جميع الجداول'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/maintenance/checkpoint', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SYSTEM_HEALTH)
def api_maintenance_checkpoint():
    """تنفيذ Checkpoint لدمج WAL"""
    try:
        dialect = getattr(db.engine, "dialect", None)
        dialect_name = (getattr(dialect, "name", "") or "").lower()
        if dialect_name != "postgresql":
            return jsonify({
                'success': False,
                'message': 'هذه العملية مدعومة فقط على PostgreSQL'
            }), 400

        try:
            db.session.rollback()
        except Exception:
            pass

        with db.engine.connect().execution_options(isolation_level="AUTOCOMMIT") as conn:
            conn.execute(text("CHECKPOINT"))
        return jsonify({'success': True, 'message': '✅ تم تنفيذ CHECKPOINT بنجاح'})
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/api/maintenance/db-info', methods=['GET'])
@permission_required(SystemPermissions.MANAGE_SYSTEM_HEALTH)
def api_maintenance_db_info():
    """الحصول على معلومات قاعدة البيانات"""
    try:
        dialect = (getattr(getattr(db, "engine", None), "dialect", None) and db.engine.dialect.name) or ""
        if str(dialect).lower() != "postgresql":
            page_size = db.session.execute(text("PRAGMA page_size")).scalar()
            page_count = db.session.execute(text("PRAGMA page_count")).scalar()
            journal_mode = db.session.execute(text("PRAGMA journal_mode")).scalar()
            size_bytes = int(page_size or 0) * int(page_count or 0)
            return jsonify({
                'success': True,
                'db_size': f"{size_bytes} bytes" if size_bytes else 'N/A',
                'wal_mode': journal_mode,
                'page_size': f"{page_size} bytes" if page_size else 'N/A'
            })
        size_row = db.session.execute(text("SELECT pg_size_pretty(pg_database_size(current_database()))")).fetchone()
        block_row = db.session.execute(text("SHOW block_size")).fetchone()
        return jsonify({
            'success': True,
            'db_size': size_row[0] if size_row else 'N/A',
            'wal_mode': None,
            'page_size': f"{block_row[0]} bytes" if block_row else 'N/A'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'❌ خطأ: {str(e)}'
        }), 500


@security_bp.route('/data-quality-center', methods=['GET', 'POST'])
@permission_required(SystemPermissions.MANAGE_SYSTEM_HEALTH)
def data_quality_center():
    """
    مركز متقدم لفحص وتحسين جودة البيانات
    يفحص جميع الحقول الإجبارية والاختيارية ويقترح إصلاحات
    """
    from models import (
        Check, Payment, PaymentSplit, Customer, Supplier, Partner,
        Sale, Invoice, ServiceRequest, Shipment, Expense, Account, GLEntry,
        PaymentMethod
    )
    from datetime import timedelta
    from decimal import Decimal
    
    if request.method == 'GET':
        # جمع إحصائيات شاملة
        issues = {
            'critical': [],
            'warning': [],
            'info': []
        }
        
        # فحص الشيكات
        checks_no_entity = Check.query.filter(
            Check.customer_id == None,
            Check.supplier_id == None,
            Check.partner_id == None
        ).count()
        
        checks_no_bank = Check.query.filter(
            db.or_(Check.check_bank == None, Check.check_bank == '')
        ).count()
        
        # فحص الدفعات
        payments_no_bank = Payment.query.filter(
            Payment.method == PaymentMethod.CHEQUE.value,
            db.or_(Payment.check_bank == None, Payment.check_bank == '')
        ).count()
        
        payments_no_due_date = Payment.query.filter(
            Payment.method == PaymentMethod.CHEQUE.value,
            Payment.check_due_date == None
        ).count()
        
        # فحص الأرصدة
        customers_null_balance = 0
        suppliers_null_balance = 0
        partners_null_balance = 0
        
        # الإحصائيات العامة
        total_checks = Check.query.count()
        total_payments = Payment.query.count()
        total_customers = Customer.query.count()
        total_suppliers = Supplier.query.count()
        total_partners = Partner.query.count()
        
        stats = {
            'checks': {
                'total': total_checks,
                'no_entity': checks_no_entity,
                'no_bank': checks_no_bank
            },
            'payments': {
                'total': total_payments,
                'no_bank': payments_no_bank,
                'no_due_date': payments_no_due_date
            },
            'balances': {
                'customers_null': customers_null_balance,
                'suppliers_null': suppliers_null_balance,
                'partners_null': partners_null_balance
            },
            'entities': {
                'customers': total_customers,
                'suppliers': total_suppliers,
                'partners': total_partners
            }
        }
        
        total_issues = (checks_no_entity + checks_no_bank + payments_no_bank + 
                       payments_no_due_date + customers_null_balance + 
                       suppliers_null_balance + partners_null_balance)
        
        return render_template('security/data_quality_center.html',
                             stats=stats,
                             total_issues=total_issues)
    
    # POST - تنفيذ الإصلاح
    try:
        action = request.form.get('action', 'all')
        fixed_count = 0
        
        if action in ['all', 'checks']:
            # إصلاح الشيكات
            from datetime import timedelta
            
            # ربط الشيكات بالجهات
            checks_without_entity = Check.query.filter(
                Check.customer_id == None,
                Check.supplier_id == None,
                Check.partner_id == None
            ).all()
            
            for check in checks_without_entity:
                payment = None
                
                if check.reference_number:
                    if check.reference_number.startswith('PMT-SPLIT-'):
                        split_id = int(check.reference_number.replace('PMT-SPLIT-', ''))
                        split = db.session.get(PaymentSplit, split_id)
                        if split:
                            payment = split.payment
                    elif check.reference_number.startswith('PMT-'):
                        try:
                            payment_id = int(check.reference_number.replace('PMT-', ''))
                            payment = db.session.get(Payment, payment_id)
                        except Exception:
                            pass
                
                if not payment and check.check_number:
                    payment = Payment.query.filter(
                        Payment.check_number == check.check_number
                    ).first()
                
                if payment:
                    if payment.customer_id:
                        check.customer_id = payment.customer_id
                        fixed_count += 1
                    elif payment.supplier_id:
                        check.supplier_id = payment.supplier_id
                        fixed_count += 1
                    elif payment.partner_id:
                        check.partner_id = payment.partner_id
                        fixed_count += 1
                    elif payment.sale_id:
                        sale = db.session.get(Sale, payment.sale_id)
                        if sale and sale.customer_id:
                            check.customer_id = sale.customer_id
                            fixed_count += 1
        
        if action in ['all', 'payments']:
            # إصلاح الدفعات
            check_payments = Payment.query.filter(
                Payment.method == PaymentMethod.CHEQUE.value
            ).all()
            
            for payment in check_payments:
                if not payment.check_bank:
                    check_record = Check.query.filter(
                        Check.reference_number == f'PMT-{payment.id}'
                    ).first()
                    if check_record and check_record.check_bank:
                        payment.check_bank = check_record.check_bank
                    else:
                        payment.check_bank = 'غير محدد'
                    fixed_count += 1
                
                if not payment.check_due_date:
                    check_record = Check.query.filter(
                        Check.reference_number == f'PMT-{payment.id}'
                    ).first()
                    if check_record and check_record.check_due_date:
                        payment.check_due_date = check_record.check_due_date
                    else:
                        payment.check_due_date = (payment.payment_date or datetime.now(timezone.utc)) + timedelta(days=30)
                    fixed_count += 1
        
        if action in ['all', 'balances']:
            pass
        
        db.session.commit()
        
        utils.log_audit("System", None, "DATA_QUALITY_FIX", 
                       details=f"تم إصلاح {fixed_count} مشكلة")
        
        flash(f'✅ تم إصلاح {fixed_count} مشكلة بنجاح!', 'success')
        return redirect(url_for('security.data_quality_center'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ خطأ: {str(e)}', 'danger')
        return redirect(url_for('security.data_quality_center'))


@security_bp.route('/advanced-check-linking', methods=['GET', 'POST'])
@permission_required(SystemPermissions.MANAGE_SYSTEM_HEALTH)
def advanced_check_linking():
    """
    ربط متقدم للشيكات بالجهات من خلال تتبع المبيعات والعلاقات
    يستخدم في حالة حدوث خلل في ربط الشيكات
    """
    from models import (
        Check, Payment, PaymentSplit, Customer, Supplier, Partner,
        Sale, Invoice, ServiceRequest, Shipment, Expense
    )
    
    if request.method == 'GET':
        # عرض الصفحة مع إحصائيات
        checks_without_entity = Check.query.filter(
            Check.customer_id == None,
            Check.supplier_id == None,
            Check.partner_id == None
        ).count()
        
        total_checks = Check.query.count()
        
        return render_template('security/advanced_check_linking.html',
                             checks_without_entity=checks_without_entity,
                             total_checks=total_checks)
    
    # POST - تنفيذ الربط
    try:
        fixed_count = 0
        errors = []
        
        checks_without_entity = Check.query.filter(
            Check.customer_id == None,
            Check.supplier_id == None,
            Check.partner_id == None
        ).all()
        
        for check in checks_without_entity:
            try:
                payment = None
                entity_found = False
                
                # محاولة البحث من reference_number
                if check.reference_number:
                    if check.reference_number.startswith('PMT-SPLIT-'):
                        split_id = int(check.reference_number.replace('PMT-SPLIT-', ''))
                        split = db.session.get(PaymentSplit, split_id)
                        if split:
                            payment = split.payment
                    elif check.reference_number.startswith('PMT-'):
                        try:
                            payment_id = int(check.reference_number.replace('PMT-', ''))
                            payment = db.session.get(Payment, payment_id)
                        except Exception:
                            pass
                    elif check.reference_number.startswith('SPLIT-'):
                        split_id = int(check.reference_number.replace('SPLIT-', ''))
                        split = db.session.get(PaymentSplit, split_id)
                        if split:
                            payment = split.payment
                
                # البحث برقم الشيك
                if not payment and check.check_number:
                    payment = Payment.query.filter(
                        Payment.check_number == check.check_number
                    ).first()
                
                # البحث بتاريخ الشيك والمبلغ
                if not payment and check.amount and check.check_date:
                    payment = Payment.query.filter(
                        Payment.total_amount == check.amount,
                        func.date(Payment.payment_date) == check.check_date.date()
                    ).first()
                
                if payment:
                    # الجهة المباشرة
                    if payment.customer_id:
                        check.customer_id = payment.customer_id
                        entity_found = True
                    elif payment.supplier_id:
                        check.supplier_id = payment.supplier_id
                        entity_found = True
                    elif payment.partner_id:
                        check.partner_id = payment.partner_id
                        entity_found = True
                    
                    # من المبيعة
                    if not entity_found and payment.sale_id:
                        sale = db.session.get(Sale, payment.sale_id)
                        if sale and sale.customer_id:
                            check.customer_id = sale.customer_id
                            entity_found = True
                    
                    # من الفاتورة
                    if not entity_found and payment.invoice_id:
                        invoice = db.session.get(Invoice, payment.invoice_id)
                        if invoice and invoice.customer_id:
                            check.customer_id = invoice.customer_id
                            entity_found = True
                    
                    # من الخدمة
                    if not entity_found and payment.service_id:
                        service = db.session.get(ServiceRequest, payment.service_id)
                        if service and service.customer_id:
                            check.customer_id = service.customer_id
                            entity_found = True
                    
                    # من الشحنة
                    if not entity_found and payment.shipment_id:
                        shipment = db.session.get(Shipment, payment.shipment_id)
                        if shipment and shipment.supplier_id:
                            check.supplier_id = shipment.supplier_id
                            entity_found = True
                    
                    # من المصروف
                    if not entity_found and payment.expense_id:
                        expense = db.session.get(Expense, payment.expense_id)
                        if expense and expense.supplier_id:
                            check.supplier_id = expense.supplier_id
                            entity_found = True
                    
                    # تحديث البيانات الأخرى
                    if entity_found:
                        if not check.currency:
                            check.currency = payment.currency or 'ILS'
                        if not check.direction:
                            check.direction = payment.direction
                        if not check.amount or check.amount == 0:
                            check.amount = payment.total_amount
                        fixed_count += 1
                
                if not entity_found:
                    errors.append(f"الشيك {check.check_number}")
                    
            except Exception as e:
                errors.append(f"الشيك {check.check_number}: {str(e)}")
        
        db.session.commit()
        
        # تسجيل في الـ audit
        utils.log_audit("System", None, "ADVANCED_CHECK_LINKING", 
                       details=f"تم ربط {fixed_count} شيك")
        
        flash(f'✅ تم ربط {fixed_count} شيك بنجاح!', 'success')
        if errors:
            flash(f'⚠️ فشل ربط {len(errors)} شيك', 'warning')
        
        return redirect(url_for('security.advanced_check_linking'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ خطأ: {str(e)}', 'danger')
        return redirect(url_for('security.advanced_check_linking'))




@security_bp.route('/help')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def help_page():
    """
    ❓ مركز المساعدة - Help Center
    
    دليل شامل للوحدة السرية:
    - شرح جميع المراكز والتبويبات
    - اختصارات لوحة المفاتيح
    - الأسئلة الشائعة
    - حل المشاكل
    """
    help_data = {
        'total_centers': 7,
        'total_features': 41,
        'total_routes': 97,
        'version': '5.0.0'
    }
    return render_template('security/help.html', help_data=help_data)


@security_bp.route('/sitemap')
@permission_required(SystemPermissions.ACCESS_OWNER_DASHBOARD)
def sitemap():
    """
    🗺️ خريطة الموقع - Site Map
    
    عرض جميع روابط الوحدة السرية في شكل شجرة:
    - 7 مراكز موحدة + تبويباتها
    - الأدوات المستقلة
    - أدوات الأمان والحظر
    - روابط سريعة
    """
    sitemap_data = {
        'centers': [
            {'name': 'Database Control Center', 'url': 'security.database_manager', 'tabs': 11},
            {'name': 'Users & Permissions Center', 'url': 'security.users_center', 'tabs': 2},
            {'name': 'Settings & Customization Center', 'url': 'security.settings_center', 'tabs': 8},
            {'name': 'Reports & Performance Center', 'url': 'security.reports_center', 'tabs': 4},
            {'name': 'Tools & Integration Center', 'url': 'security.tools_center', 'tabs': 5},
            {'name': 'Security & Monitoring Center', 'url': 'security.security_center', 'tabs': 4},
            {'name': 'Ledger Control', 'url': 'ledger_control.index', 'tabs': 0},
        ],
        'total_routes': 97,
        'total_tabs': 34
    }
    return render_template('security/sitemap.html', sitemap_data=sitemap_data)


@security_bp.route('/api/system-constants')
@permission_required(SystemPermissions.MANAGE_SYSTEM_CONFIG)
def api_system_constants():
    """
    🔧 API للحصول على ثوابت النظام (للاستخدام في JavaScript)
    
    Returns:
        JSON مع جميع ثوابت الأعمال
    
    Example:
        GET /security/api/system-constants
        
        Response:
        {
            "success": true,
            "data": {
                "tax": {"default_vat_rate": 16.0, ...},
                "payroll": {...}
            }
        }
    """
    try:
        from utils import get_all_business_constants
        constants = get_all_business_constants()
        return jsonify({
            'success': True,
            'data': constants
        })
    except Exception as e:
        current_app.logger.error(f"⚠️ فشل جلب الثوابت: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== وحدة الإشعارات - Notifications Center ====================

@security_bp.route('/notifications', methods=['GET'])
@permission_required(SystemPermissions.MANAGE_SYSTEM_CONFIG)
def notifications_log():
    """
    📧 مركز الإشعارات - Notifications Center
    
    إدارة شاملة:
    - سجل الإشعارات
    - إحصائيات
    - اختبار الإرسال
    """
    from models import NotificationLog
    from sqlalchemy import func, desc
    from datetime import datetime, timedelta
    
    # الإحصائيات
    stats = {
        'total': NotificationLog.query.count(),
        'sent': NotificationLog.query.filter_by(status='sent').count(),
        'failed': NotificationLog.query.filter_by(status='failed').count(),
        'today': NotificationLog.query.filter(
            NotificationLog.created_at >= datetime.now().date()
        ).count()
    }
    
    # آخر 50 إشعار
    recent_logs = NotificationLog.query.order_by(
        desc(NotificationLog.created_at)
    ).limit(50).all()
    
    # توزيع حسب النوع
    type_stats_list = db.session.query(
        NotificationLog.type,
        func.count(NotificationLog.id).label('count')
    ).group_by(NotificationLog.type).all()
    
    type_stats = dict(type_stats_list) if type_stats_list else {}
    
    return render_template(
        'security/notifications.html',
        stats=stats,
        recent_logs=recent_logs,
        type_stats=type_stats
    )


@security_bp.route('/notifications/test', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_SYSTEM_CONFIG)
def test_notification():
    """اختبار إرسال إشعار"""
    from utils import send_notification_sms, send_notification_email
    
    notification_type = request.form.get('type')
    recipient = request.form.get('recipient')
    
    if notification_type == 'sms':
        result = send_notification_sms(
            to=recipient,
            message='🧪 رسالة اختبار من النظام',
            metadata={'type': 'test'}
        )
    elif notification_type == 'email':
        result = send_notification_email(
            to=recipient,
            subject='🧪 رسالة اختبار',
            body_html='<h2>رسالة اختبار من النظام</h2><p>النظام يعمل بنجاح!</p>',
            metadata={'type': 'test'}
        )
    else:
        return jsonify({'success': False, 'error': 'Invalid type'}), 400
    
    return jsonify(result)


# ==================== وحدة الضرائب - Tax Module ====================

@security_bp.route('/tax-reports', methods=['GET'])
@permission_required(SystemPermissions.MANAGE_REPORTS)
def tax_reports():
    """
    💰 تقارير الضرائب - Tax Reports
    
    - ملخص VAT
    - تقارير شهرية/سنوية
    - الإقرارات الضريبية
    """
    from models import TaxEntry, Sale, SaleStatus
    from utils import get_tax_summary
    from sqlalchemy import func
    from datetime import datetime
    from calendar import monthrange
    
    # الفترة
    period = request.args.get('period', datetime.now().strftime('%Y-%m'))
    
    # ملخص الفترة
    summary = get_tax_summary(period)
    
    # السجلات التفصيلية
    entries = TaxEntry.query.filter_by(tax_period=period).order_by(
        TaxEntry.created_at.desc()
    ).limit(100).all()
    
    # إحصائيات سنوية
    year = period.split('-')[0]
    yearly_stats = db.session.query(
        TaxEntry.tax_period,
        TaxEntry.entry_type,
        func.sum(TaxEntry.tax_amount).label('total')
    ).filter(
        TaxEntry.fiscal_year == int(year)
    ).group_by(
        TaxEntry.tax_period,
        TaxEntry.entry_type
    ).all()
    
    # تنظيم البيانات السنوية
    yearly_data = {}
    for period_key, entry_type, total in yearly_stats:
        if period_key not in yearly_data:
            yearly_data[period_key] = {}
        yearly_data[period_key][entry_type] = float(total or 0)

    confirmed_sales_count = 0
    confirmed_sales_without_vat = 0
    try:
        y_str, m_str = str(period).split("-", 1)
        y = int(y_str)
        m = int(m_str)
        last_day = monthrange(y, m)[1]
        start_dt = datetime(y, m, 1)
        end_dt = datetime(y, m, last_day, 23, 59, 59, 999999)
        confirmed_sales_count = Sale.query.filter(
            Sale.status == SaleStatus.CONFIRMED.value,
            Sale.sale_date >= start_dt,
            Sale.sale_date <= end_dt,
        ).count()
        confirmed_sales_without_vat = Sale.query.filter(
            Sale.status == SaleStatus.CONFIRMED.value,
            Sale.sale_date >= start_dt,
            Sale.sale_date <= end_dt,
            (Sale.tax_rate.is_(None)) | (Sale.tax_rate <= 0),
        ).count()
    except Exception:
        confirmed_sales_count = 0
        confirmed_sales_without_vat = 0
    
    return render_template(
        'security/tax_reports.html',
        period=period,
        summary=summary,
        entries=entries,
        yearly_data=yearly_data,
        current_year=year,
        confirmed_sales_count=confirmed_sales_count,
        confirmed_sales_without_vat=confirmed_sales_without_vat,
    )


def _vat_backfill_sales_for_period(period: str) -> dict:
    from datetime import datetime
    from calendar import monthrange
    from decimal import Decimal, ROUND_HALF_UP
    from sqlalchemy.orm import joinedload
    from models import (
        db,
        TaxEntry,
        Sale,
        SaleStatus,
        ServiceRequest,
        ServiceStatus,
        GLBatch,
        GLEntry,
        SystemSettings,
        _ensure_account_exists,
        convert_amount,
    )

    try:
        year_str, month_str = str(period or "").split("-", 1)
        year = int(year_str)
        month = int(month_str)
        if month < 1 or month > 12:
            raise ValueError("invalid month")
    except Exception:
        raise ValueError("invalid_period")

    vat_enabled_row = SystemSettings.query.filter_by(key="vat_enabled").first()
    vat_enabled = str(getattr(vat_enabled_row, "value", "") or "").strip().lower() in ("1", "true", "yes", "on")
    if not vat_enabled:
        return {
            "period": period,
            "vat_enabled": False,
            "created_count": 0,
            "missing_vat_count": 0,
            "posted_batch": False,
            "vat_total_ils": 0.0,
            "message": "⚠️ VAT غير مفعّل في إعدادات النظام.",
            "message_level": "warning",
        }

    vat_rate_row = SystemSettings.query.filter_by(key="default_vat_rate").first()
    try:
        vat_rate = float(str(getattr(vat_rate_row, "value", "") or "16").strip() or 16.0)
    except Exception:
        vat_rate = 16.0
    if vat_rate <= 0:
        vat_rate = 16.0

    start_dt = datetime(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_dt = datetime(year, month, last_day, 23, 59, 59, 999999)

    existing_sale_ids = {
        int(x[0])
        for x in db.session.query(TaxEntry.transaction_id).filter(
            TaxEntry.tax_period == period,
            TaxEntry.transaction_type == "SALE",
            TaxEntry.entry_type == "OUTPUT_VAT",
        ).all()
        if x and x[0] is not None
    }

    sales = (
        Sale.query.options(joinedload(Sale.lines))
        .filter(
            Sale.status == SaleStatus.CONFIRMED.value,
            Sale.sale_date >= start_dt,
            Sale.sale_date <= end_dt,
        )
        .all()
    )

    created_count = 0
    missing_vat_count = 0
    vat_total_ils = Decimal("0.00")
    tolerance = Decimal("0.50")
    created_service_count = 0
    updated_service_count = 0

    for s in sales:
        try:
            sid = int(getattr(s, "id", 0) or 0)
        except Exception:
            sid = 0
        if not sid or sid in existing_sale_ids:
            continue

        total = Decimal(str(getattr(s, "total_amount", 0) or 0))
        if total <= Decimal("0"):
            continue

        sale_rate = Decimal(str(getattr(s, "tax_rate", 0) or 0))
        if sale_rate <= Decimal("0"):
            sale_rate = Decimal(str(vat_rate))

        subtotal_net = Decimal("0.00")
        for ln in (getattr(s, "lines", None) or []):
            try:
                subtotal_net += Decimal(str(getattr(ln, "net_amount", 0) or 0))
            except Exception:
                continue

        disc = Decimal(str(getattr(s, "discount_total", 0) or 0))
        ship = Decimal(str(getattr(s, "shipping_cost", 0) or 0))
        base_for_tax = (subtotal_net - disc + ship).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if base_for_tax < Decimal("0.00"):
            base_for_tax = Decimal("0.00")

        expected_with_vat = (base_for_tax * (Decimal("1") + (sale_rate / Decimal("100")))).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        expected_no_vat = base_for_tax

        if abs(total - expected_with_vat) > tolerance and abs(total - expected_with_vat) > abs(total - expected_no_vat):
            missing_vat_count += 1
            continue

        denom = (Decimal("1") + (sale_rate / Decimal("100")))
        base = (total / denom).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        tax = (total - base).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if tax <= Decimal("0.00"):
            continue

        currency = (getattr(s, "currency", None) or "ILS").upper()
        try:
            if currency == "ILS":
                vat_ils = tax
            else:
                vat_ils = convert_amount(tax, currency, "ILS", getattr(s, "sale_date", None))
        except Exception:
            try:
                fx_raw = Decimal(str(getattr(s, "fx_rate_used", 0) or 0))
                if fx_raw > Decimal("0"):
                    vat_ils = (tax * fx_raw).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                else:
                    vat_ils = tax
            except Exception:
                vat_ils = tax

        vat_total_ils += vat_ils.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        te = TaxEntry(
            entry_type="OUTPUT_VAT",
            transaction_type="SALE",
            transaction_id=sid,
            transaction_reference=getattr(s, "sale_number", None),
            tax_rate=float(sale_rate),
            base_amount=float(base),
            tax_amount=float(tax),
            total_amount=float(total),
            currency=currency,
            fiscal_year=year,
            fiscal_month=month,
            tax_period=period,
            customer_id=getattr(s, "customer_id", None),
            notes=f"تصحيح VAT لبيع: {getattr(s, 'sale_number', None) or sid}",
        )
        db.session.add(te)
        created_count += 1

    existing_service_entries = {
        int(r.transaction_id): (
            Decimal(str(r.base_amount or 0)),
            Decimal(str(r.tax_amount or 0)),
            Decimal(str(r.total_amount or 0)),
        )
        for r in db.session.query(TaxEntry).filter(
            TaxEntry.tax_period == period,
            TaxEntry.transaction_type == "SERVICE",
            TaxEntry.entry_type == "OUTPUT_VAT",
        ).all()
        if getattr(r, "transaction_id", None) is not None
    }

    services = (
        ServiceRequest.query.options(joinedload(ServiceRequest.parts), joinedload(ServiceRequest.tasks))
        .filter(
            ServiceRequest.status.in_([ServiceStatus.COMPLETED.value, getattr(ServiceStatus, "CLOSED", ServiceStatus.COMPLETED).value]),
            ServiceRequest.received_at >= start_dt,
            ServiceRequest.received_at <= end_dt,
        )
        .all()
    )

    for srv in services:
        try:
            sid = int(getattr(srv, "id", 0) or 0)
        except Exception:
            sid = 0
        if not sid:
            continue

        base = Decimal(str(getattr(srv, "total_amount", 0) or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if base <= Decimal("0.00"):
            continue

        srv_rate = Decimal(str(getattr(srv, "tax_rate", 0) or 0))
        if srv_rate <= Decimal("0"):
            srv_rate = Decimal(str(vat_rate))

        tax = (base * srv_rate / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if tax <= Decimal("0.00"):
            continue

        total = (base + tax).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        existing = existing_service_entries.get(sid)
        if existing:
            eb, et, etot = existing
            if abs(eb - base) <= Decimal("0.01") and abs(et - tax) <= Decimal("0.01") and abs(etot - total) <= Decimal("0.01"):
                continue

        currency = (getattr(srv, "currency", None) or "ILS").upper()
        te_q = TaxEntry.query.filter_by(transaction_type="SERVICE", transaction_id=sid, entry_type="OUTPUT_VAT").all()
        for old in te_q:
            db.session.delete(old)

        te = TaxEntry(
            entry_type="OUTPUT_VAT",
            transaction_type="SERVICE",
            transaction_id=sid,
            transaction_reference=getattr(srv, "service_number", None),
            tax_rate=float(srv_rate),
            base_amount=float(base),
            tax_amount=float(tax),
            total_amount=float(total),
            currency=currency,
            fiscal_year=year,
            fiscal_month=month,
            tax_period=period,
            customer_id=getattr(srv, "customer_id", None),
            notes=f"تصحيح VAT لصيانة: {getattr(srv, 'service_number', None) or sid}",
        )
        db.session.add(te)
        if existing:
            updated_service_count += 1
        else:
            created_service_count += 1

    vat_total_ils = vat_total_ils.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    posted_batch = False
    if vat_total_ils > Decimal("0.00"):
        yyyymm = int(year * 100 + month)
        existing_batch = GLBatch.query.filter_by(source_type="VAT_BACKFILL", source_id=yyyymm).first()
        if not existing_batch:
            conn = db.session.connection()
            acc_sales = "4000_SALES"
            acc_vat = "2100_VAT_PAYABLE"
            _ensure_account_exists(conn, acc_sales)
            _ensure_account_exists(conn, acc_vat)

            batch = GLBatch(
                source_type="VAT_BACKFILL",
                source_id=yyyymm,
                purpose="VAT_BACKFILL",
                posted_at=end_dt,
                currency="ILS",
                memo=f"VAT Backfill {period}",
                status="POSTED",
                entity_type=None,
                entity_id=None,
            )
            db.session.add(batch)
            db.session.flush()

            ref_number = f"VAT-BACKFILL-{period}"
            db.session.add(
                GLEntry(
                    batch_id=batch.id,
                    account=acc_sales,
                    debit=float(vat_total_ils),
                    credit=0,
                    currency="ILS",
                    ref=ref_number,
                )
            )
            db.session.add(
                GLEntry(
                    batch_id=batch.id,
                    account=acc_vat,
                    debit=0,
                    credit=float(vat_total_ils),
                    currency="ILS",
                    ref=ref_number,
                )
            )
            posted_batch = True

    db.session.commit()

    msg = f"✅ تم إنشاء {created_count} سجل VAT للفترة {period}."
    if missing_vat_count:
        msg += f" (تم تجاوز {missing_vat_count} مبيعات يبدو أنها بدون VAT)"
    if created_service_count or updated_service_count:
        msg += f" وتم إنشاء {created_service_count} وتحديث {updated_service_count} للصيانة."
    if posted_batch:
        msg += f" وتم إنشاء قيد دفتر الأستاذ بمبلغ {float(vat_total_ils):.2f} ₪."

    return {
        "period": period,
        "vat_enabled": True,
        "created_count": created_count,
        "missing_vat_count": missing_vat_count,
        "created_service_count": created_service_count,
        "updated_service_count": updated_service_count,
        "posted_batch": posted_batch,
        "vat_total_ils": float(vat_total_ils),
        "message": msg,
        "message_level": "success",
    }


@security_bp.route('/tax-reports/reconcile/<period>', methods=['POST'])
@permission_required(SystemPermissions.MANAGE_REPORTS)
def reconcile_tax_report(period):
    from datetime import datetime
    from flask import redirect, url_for

    try:
        result = _vat_backfill_sales_for_period(period)
    except ValueError:
        valid_period = datetime.now().strftime("%Y-%m")
        flash("❌ فترة غير صالحة.", "danger")
        return redirect(url_for("security.tax_reports", period=valid_period))
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f"❌ فشل تصحيح VAT: {e}", "danger")
        return redirect(url_for("security.tax_reports", period=period))

    flash(result.get("message") or "تمت العملية.", result.get("message_level") or "success")
    return redirect(url_for("security.tax_reports", period=period))


@security_bp.route('/tax-reports/export/<period>')
@permission_required(SystemPermissions.MANAGE_REPORTS)
def export_tax_report(period):
    from flask import Response, send_file, redirect, url_for
    from models import TaxEntry
    from decimal import Decimal
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    from datetime import datetime
    
    try:
        year, month = period.split('-', 1)
        int(year)
        int(month)
    except Exception:
        valid_period = datetime.now().strftime('%Y-%m')
        return redirect(url_for('security.tax_reports', period=valid_period))
    
    output_entries = TaxEntry.query.filter(
        TaxEntry.entry_type == 'OUTPUT_VAT',
        TaxEntry.tax_period == period
    ).all()
    
    input_entries = TaxEntry.query.filter(
        TaxEntry.entry_type == 'INPUT_VAT',
        TaxEntry.tax_period == period
    ).all()
    
    total_sales = sum(Decimal(str(e.base_amount or 0)) for e in output_entries)
    output_vat = sum(Decimal(str(e.tax_amount or 0)) for e in output_entries)
    total_purchases = sum(Decimal(str(e.base_amount or 0)) for e in input_entries)
    input_vat = sum(Decimal(str(e.tax_amount or 0)) for e in input_entries)
    net_vat = output_vat - input_vat
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"VAT-{period}"
    
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws['A1'] = 'إقرار ضريبة القيمة المضافة - فلسطين'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f'الفترة: {month}/{year}'
    ws['A2'].font = Font(bold=True, size=12)
    
    ws['A4'] = 'البند'
    ws['B4'] = 'الوصف'
    ws['C4'] = 'المبلغ الأساسي (₪)'
    ws['D4'] = 'ضريبة القيمة المضافة (₪)'
    for cell in ['A4', 'B4', 'C4', 'D4']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].border = border
        ws[cell].alignment = Alignment(horizontal='center')
    
    ws['A5'] = '1'
    ws['B5'] = 'المبيعات الخاضعة للضريبة'
    ws['C5'] = float(total_sales)
    ws['D5'] = float(output_vat)
    
    ws['A6'] = '2'
    ws['B6'] = 'المشتريات القابلة للخصم'
    ws['C6'] = float(total_purchases)
    ws['D6'] = float(input_vat)
    
    ws['A8'] = '3'
    ws['B8'] = 'صافي الضريبة المستحقة/المستردة'
    ws['D8'] = float(net_vat)
    ws['D8'].font = Font(bold=True, color="FF0000" if net_vat > 0 else "00AA00")
    
    for row in range(5, 9):
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = border
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'VAT_Declaration_{period}.xlsx'
    )


# ==================== وحدة سجل التدقيق - Audit Log ====================

@security_bp.route('/audit-log')
@permission_required(SystemPermissions.VIEW_AUDIT_LOGS)
def audit_log_viewer():
    from flask import Response
    import csv
    import io
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    export = request.args.get('export', '').strip()
    
    model_name = request.args.get('model', '').strip()
    user_id = request.args.get('user', type=int)
    action = request.args.get('action', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    search = request.args.get('search', '').strip()
    
    query = AuditLog.query
    
    if model_name:
        query = query.filter_by(model_name=model_name)
    if user_id:
        query = query.filter_by(user_id=user_id)
    if action:
        query = query.filter_by(action=action.upper())
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(AuditLog.created_at >= start_dt)
        except Exception:
            pass
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(AuditLog.created_at <= end_dt)
        except Exception:
            pass
    if search:
        query = query.filter(
            db.or_(
                AuditLog.model_name.ilike(f'%{search}%'),
                AuditLog.action.ilike(f'%{search}%')
            )
        )
    
    if export == 'csv':
        logs_data = query.order_by(AuditLog.created_at.desc()).limit(10000).all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['الوقت', 'المستخدم', 'الجدول', 'رقم السجل', 'الإجراء', 'IP'])
        
        for log in logs_data:
            user_name = (db.session.get(User, log.user_id).username if log.user_id else 'نظام')
            writer.writerow([
                log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
                user_name,
                log.model_name,
                log.record_id or '',
                log.action,
                log.ip_address or ''
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=audit_log_{datetime.now().strftime("%Y%m%d")}.csv'}
        )
    
    logs = query.order_by(AuditLog.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    total_logs = AuditLog.query.count()
    last_24h = AuditLog.query.filter(
        AuditLog.created_at >= datetime.now(timezone.utc) - timedelta(days=1)
    ).count()
    
    by_action = db.session.query(
        AuditLog.action,
        func.count(AuditLog.id)
    ).group_by(AuditLog.action).order_by(func.count(AuditLog.id).desc()).limit(10).all()
    
    top_users = db.session.query(
        User.username,
        func.count(AuditLog.id).label('count')
    ).join(User, User.id == AuditLog.user_id).group_by(User.username).order_by(func.count(AuditLog.id).desc()).limit(10).all()
    
    all_models = db.session.query(AuditLog.model_name).distinct().order_by(AuditLog.model_name).all()
    all_actions = db.session.query(AuditLog.action).distinct().order_by(AuditLog.action).all()
    all_users = User.query.order_by(User.username).all()
    
    stats = {
        'total_logs': total_logs,
        'last_24h': last_24h,
        'by_action': by_action,
        'top_users': top_users
    }
    
    return render_template('security/audit_log.html',
                         logs=logs,
                         stats=stats,
                         all_models=[m[0] for m in all_models],
                         all_actions=[a[0] for a in all_actions],
                         all_users=all_users,
                         filters={
                             'model': model_name,
                             'user': user_id,
                             'action': action,
                             'start_date': start_date,
                             'end_date': end_date,
                             'search': search
                         })

@security_bp.route('/audit-log/<int:log_id>')
@permission_required(SystemPermissions.VIEW_AUDIT_LOGS)
def audit_log_detail(log_id):
    log = db.session.get(AuditLog, log_id)
    if not log:
        if request.accept_mimetypes.best == 'application/json':
            return jsonify({'success': False, 'error': 'السجل غير موجود'}), 404
        flash('السجل غير موجود', 'danger')
        return redirect(url_for('security.audit_log_viewer'))
    
    old_data = {}
    new_data = {}
    changes = []
    
    if log.old_data:
        try:
            old_data = json.loads(log.old_data)
        except Exception:
            old_data = {}
    
    if log.new_data:
        try:
            new_data = json.loads(log.new_data)
        except Exception:
            new_data = {}
    
    all_keys = set(old_data.keys()) | set(new_data.keys())
    for key in sorted(all_keys):
        old_val = old_data.get(key, '')
        new_val = new_data.get(key, '')
        if old_val != new_val:
            changes.append({
                'field': key,
                'old': str(old_val) if old_val else '-',
                'new': str(new_val) if new_val else '-'
            })
    
    user_name = (db.session.get(User, log.user_id).username if log.user_id else 'نظام')
    
    if request.accept_mimetypes.best == 'application/json':
        log_dict = {
            'id': log.id,
            'model_name': log.model_name,
            'record_id': log.record_id,
            'action': log.action,
            'user_id': log.user_id,
            'ip_address': log.ip_address,
            'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else None
        }
        
        return jsonify({
            'success': True,
            'log': log_dict,
            'old_data': old_data,
            'new_data': new_data,
            'changes': changes,
            'user': user_name
        })
    
    return render_template('security/audit_log_detail.html',
                         log=log,
                         user_name=user_name,
                         old_data=old_data,
                         new_data=new_data,
                         changes=changes)


# ==================== وحدة التكامل مع الأجهزة والأنظمة ====================


@security_bp.route('/security-audit-report', methods=['GET'])
@permission_required(SystemPermissions.VIEW_AUDIT_LOGS)
def security_audit_report():
    """Security Audit Report شامل - تقرير أمني شامل"""
    from datetime import datetime, timedelta, timezone
    from models import AuthAudit, AuthEvent
    
    cache_key = "security_audit_report"
    cached_report = cache.get(cache_key)
    
    if cached_report is None:
        try:
            now = datetime.now(timezone.utc)
            last_24h = now - timedelta(hours=24)
            last_7d = now - timedelta(days=7)
            last_30d = now - timedelta(days=30)
            
            failed_logins_24h = AuthAudit.query.filter(
                AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
                AuthAudit.created_at >= last_24h
            ).count()
            
            failed_logins_7d = AuthAudit.query.filter(
                AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
                AuthAudit.created_at >= last_7d
            ).count()
            
            successful_logins_24h = AuthAudit.query.filter(
                AuthAudit.event == AuthEvent.LOGIN_SUCCESS.value,
                AuthAudit.created_at >= last_24h
            ).count()
            
            security_actions = AuditLog.query.filter(
                AuditLog.action.like('security.%'),
                AuditLog.created_at >= last_30d
            ).count()
            
            blocked_ips_count = 0
            try:
                from models import BlockedIP
                blocked_ips_count = BlockedIP.query.count()
            except Exception:
                pass
            
            suspicious_activities = AuditLog.query.filter(
                AuditLog.action.in_(['security.blocked_ip', 'security.failed_login', 'security.unauthorized_access']),
                AuditLog.created_at >= last_7d
            ).count()
            
            top_failed_ips = db.session.query(
                AuthAudit.ip,
                func.count(AuthAudit.id).label('count')
            ).filter(
                AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
                AuthAudit.created_at >= last_7d
            ).group_by(AuthAudit.ip).order_by(func.count(AuthAudit.id).desc()).limit(10).all()
            
            report = {
                'summary': {
                    'failed_logins_24h': failed_logins_24h,
                    'failed_logins_7d': failed_logins_7d,
                    'successful_logins_24h': successful_logins_24h,
                    'security_actions_30d': security_actions,
                    'blocked_ips': blocked_ips_count,
                    'suspicious_activities_7d': suspicious_activities
                },
                'top_failed_ips': [{'ip': ip, 'count': count} for ip, count in top_failed_ips],
                'generated_at': now.isoformat()
            }
            
            cache.set(cache_key, report, timeout=600)
        except Exception as e:
            current_app.logger.error(f"Error generating security audit report: {e}")
            report = {
                'summary': {},
                'top_failed_ips': [],
                'error': str(e),
                'generated_at': datetime.now(timezone.utc).isoformat()
            }
    else:
        report = cached_report
    
    return render_template('security/security_audit_report.html', report=report)


@security_bp.route('/api/security-audit/stats', methods=['GET'])
@permission_required(SystemPermissions.VIEW_AUDIT_LOGS)
def api_security_audit_stats():
    """API للحصول على إحصائيات الأمان"""
    from datetime import datetime, timedelta, timezone
    from models import AuthAudit, AuthEvent
    
    try:
        now = datetime.now(timezone.utc)
        last_24h = now - timedelta(hours=24)
        
        failed_logins = AuthAudit.query.filter(
            AuthAudit.event == AuthEvent.LOGIN_FAIL.value,
            AuthAudit.created_at >= last_24h
        ).count()
        
        successful_logins = AuthAudit.query.filter(
            AuthAudit.event == AuthEvent.LOGIN_SUCCESS.value,
            AuthAudit.created_at >= last_24h
        ).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'failed_logins_24h': failed_logins,
                'successful_logins_24h': successful_logins,
                'timestamp': now.isoformat()
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
